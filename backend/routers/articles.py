from fastapi import APIRouter, Query, HTTPException, WebSocket, WebSocketDisconnect, Depends, status
import asyncio
import csv
import io
from typing import Optional, AsyncGenerator
from datetime import date, datetime
from sqlalchemy import select, func, or_, and_, update, desc, text, delete
from db.database import get_db, Article, ScrapeJob
from .auth_utils import get_auth_user as get_current_user, TokenData
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter()

@router.get("/")
async def get_articles(
    sector: Optional[str] = None,
    region: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    job_id: Optional[str] = None,
    search: Optional[str] = None,
    has_body: Optional[bool] = None,
    page: int = 1,
    page_size: int = 25,
    current_user: TokenData = Depends(get_current_user)
):
    """Fetch articles with filtering and pagination."""
    offset = (page - 1) * page_size
    async with get_db() as db:
        stmt = select(Article).where(Article.user_id == current_user.id)
        
        if sector: stmt = stmt.where(Article.sector == sector)
        if region: stmt = stmt.where(Article.region == region)
        if date_from: stmt = stmt.where(Article.published_at >= date_from)
        if date_to: stmt = stmt.where(Article.published_at <= date_to)
        if job_id: stmt = stmt.where(Article.scrape_job_id == job_id)
        if search:
            stmt = stmt.where(or_(
                Article.title.ilike(f"%{search}%"),
                Article.full_body.ilike(f"%{search}%")
            ))
        if has_body is True:
            stmt = stmt.where(and_(Article.full_body != None, func.length(Article.full_body) > 150))
        elif has_body is False:
            stmt = stmt.where(or_(Article.full_body == None, func.length(Article.full_body) <= 150))

        # Count total
        count_stmt = select(func.count()).select_from(stmt.subquery())
        res_total = await db.execute(count_stmt)
        total = res_total.scalar()

        # Get results
        stmt = stmt.order_by(Article.published_at.desc()).offset(offset).limit(page_size)
        res_articles = await db.execute(stmt)
        articles = res_articles.scalars().all()

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": -(-total // page_size) if total else 0,
        "articles": articles
    }

async def _fetch_stats_logic(user_id: str):
    """Internal stats logic."""
    async with get_db() as db:
        # Total
        total_res = await db.execute(select(func.count(Article.id)).where(Article.user_id == user_id))
        total = total_res.scalar() or 0
        
        # With Body
        body_res = await db.execute(select(func.count(Article.id)).where(Article.user_id == user_id, Article.full_body != None, func.length(Article.full_body) > 150))
        with_body = body_res.scalar() or 0
        
        # With Summary
        sum_res = await db.execute(select(func.count(Article.id)).where(Article.user_id == user_id, Article.summary != None))
        with_summary = sum_res.scalar() or 0
        
        # By Sector
        sect_res = await db.execute(select(Article.sector, func.count(Article.id)).where(Article.user_id == user_id).group_by(Article.sector).order_by(desc(func.count(Article.id))))
        by_sector = [{"sector": r[0], "count": r[1]} for r in sect_res.all()]
        
        # Jobs
        jobs_res = await db.execute(select(ScrapeJob.status, func.count(ScrapeJob.id)).where(ScrapeJob.user_id == user_id).group_by(ScrapeJob.status))
        jobs_by_status = [{"status": r[0], "count": r[1]} for r in jobs_res.all()]

    return {
        "total_articles": total,
        "articles_with_body": with_body,
        "articles_with_summary": with_summary,
        "body_coverage_pct": round((with_body / total * 100), 1) if total else 0,
        "by_sector": by_sector,
        "jobs_by_status": jobs_by_status
    }

@router.get("/stats/summary")
async def get_stats(current_user: TokenData = Depends(get_current_user)):
    return await _fetch_stats_logic(current_user.id)

@router.get("/export/csv")
async def export_csv(
    job_id: str,
    current_user: TokenData = Depends(get_current_user)
):
    async def generate():
        yield b'\xef\xbb\xbf' # BOM for Excel
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Title", "URL", "Agency", "Published_At", "Summary"])
        yield output.getvalue().encode("utf-8")
        
        async with get_db() as db:
            stmt = select(Article).where(Article.scrape_job_id == job_id, Article.user_id == current_user.id)
            res = await db.execute(stmt)
            for a in res.scalars():
                out = io.StringIO()
                cw = csv.writer(out)
                cw.writerow([a.title, a.url, a.agency, a.published_at, a.summary])
                yield out.getvalue().encode("utf-8")

    return StreamingResponse(generate(), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename=export_{job_id}.csv"})

@router.get("/export/xlsx")
async def export_xlsx(
    job_id: str,
    current_user: TokenData = Depends(get_current_user)
):
    """Excel export with formatting (Blueprint Phase 5)."""
    async with get_db() as db:
        # Get job for naming
        job_res = await db.execute(select(ScrapeJob).where(ScrapeJob.id == job_id, ScrapeJob.user_id == current_user.id))
        job = job_res.scalar_one_or_none()
        if not job:
            raise HTTPException(404, "Job not found")
        
        # Get articles
        count_stmt = select(func.count()).where(Article.scrape_job_id == job_id, Article.user_id == current_user.id)
        total_count = (await db.execute(count_stmt)).scalar() or 0
        
        if total_count == 0:
            raise HTTPException(400, "No articles found for this job. Ensure discovery is complete.")
            
        if total_count > 5000:
            # Hard limit for XLSX to prevent OOM
            raise HTTPException(400, "Job too large for XLSX (Max 5,000 articles). Use CSV export instead.")

        stmt = select(Article).where(Article.scrape_job_id == job_id, Article.user_id == current_user.id).order_by(Article.published_at.desc())
        res = await db.execute(stmt)
        articles = res.scalars().all()

        wb = Workbook()
        ws = wb.active
        brand_name = job.sector.replace(" ", "_")
        ws.title = f"Nexus_{brand_name}"

        # Columns: Title, Resolved URL, Publisher/Agency, Author, Summary, Published At, Source Feed, Keyword Matched
        headers = ["Title", "Resolved URL", "Publisher/Agency", "Author", "Summary", "Published At", "Source Feed", "Keyword Matched"]
        ws.append(headers)

        # Formatting Header
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Freeze top row
        ws.freeze_panes = "A2"

        # Data Rows
        alternate_fill = PatternFill(start_color="F0F4FA", end_color="F0F4FA", fill_type="solid")
        for i, a in enumerate(articles, start=2):
            published_str = a.published_at.strftime("%d %b %Y %H:%M") if a.published_at else ""
            row_data = [
                a.title,
                a.resolved_url or a.url,
                a.agency,
                a.author or "Staff Reporter",
                a.summary,
                published_str,
                a.source_feed or "google_news",
                job.sector
            ]
            ws.append(row_data)
            
            # Alternate row shading
            if i % 2 == 0:
                for cell in ws[i]:
                    cell.fill = alternate_fill

            # Hyperlinks for URL
            url_cell = ws.cell(row=i, column=2)
            url_cell.hyperlink = a.resolved_url or a.url
            url_cell.font = Font(color="0000FF", underline="single")

        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except: pass
            adjusted_width = min(max(max_length + 2, 15), 60)
            ws.column_dimensions[column].width = adjusted_width

        # Stream back
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"NEXUS_{brand_name}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

@router.get("/{article_id}")
async def get_article(article_id: int, current_user: TokenData = Depends(get_current_user)):
    async with get_db() as db:
        res = await db.execute(select(Article).where(Article.id == article_id, Article.user_id == current_user.id))
        art = res.scalar_one_or_none()
        if not art: raise HTTPException(404)
        return art

@router.delete("/{article_id}")
async def delete_article(article_id: int, current_user: TokenData = Depends(get_current_user)):
    async with get_db() as db:
        res = await db.execute(select(Article).where(Article.id == article_id, Article.user_id == current_user.id))
        art = res.scalar_one_or_none()
        if not art: raise HTTPException(404, "Article not found or access denied")
        
        await db.execute(delete(Article).where(Article.id == article_id))
        await db.commit()
        return {"status": "success", "message": "Article deleted permanently"}
@router.websocket("/ws/stats")
async def websocket_stats(websocket: WebSocket, token: Optional[str] = Query(None)):
    await websocket.accept()
    if not token:
        await websocket.close(code=status.WS_1008_POLICY_VIOLATION)
        return
    
    from .auth_utils import get_current_user
    try:
        user_data = await get_current_user(token)
    except Exception:
        await websocket.close(code=status.WS_1008_POLICY_VIOLATION)
        return

    try:
        # Send initial stats immediately
        initial_stats = await _fetch_stats_logic(user_data.user_id)
        await websocket.send_json(initial_stats)
        
        while True:
            await asyncio.sleep(10)
            stats = await _fetch_stats_logic(user_data.user_id)
            await websocket.send_json(stats)
    except WebSocketDisconnect:
        pass
    except Exception as e:
        # Ensure we don't crash the worker thread
        print(f"WS Error: {e}")
        pass
