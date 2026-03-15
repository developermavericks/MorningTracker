import uuid
import io
import csv
from datetime import date, timedelta, datetime
from typing import List, Optional
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, func, update, delete, and_
from db.database import get_db, WatchedBrand, Article, ScrapeJob
from .auth_utils import get_auth_user as get_current_user, TokenData
from celery_app import celery_app

router = APIRouter()

class BrandRequest(BaseModel):
    name: str
    keywords: Optional[str] = None
    region: Optional[str] = "india"

@router.get("/")
async def get_brands(current_user: TokenData = Depends(get_current_user)):
    """List all watched brands with article counts."""
    async with get_db() as db:
        # Subquery for article count
        count_stmt = (
            select(func.count(Article.id))
            .where(Article.sector == WatchedBrand.name)
            .where(Article.user_id == current_user.id)
            .label("article_count")
        )
        
        stmt = (
            select(WatchedBrand, count_stmt)
            .where(WatchedBrand.user_id == current_user.id)
            .order_by(WatchedBrand.name.asc())
        )
        
        res = await db.execute(stmt)
        results = []
        for brand, count in res.all():
            brand_dict = {
                "id": brand.id,
                "name": brand.name,
                "keywords": brand.keywords,
                "region": brand.region,
                "created_at": brand.created_at,
                "last_scraped": brand.last_scraped,
                "article_count": count
            }
            results.append(brand_dict)
        return results

@router.post("/")
async def add_brand(req: BrandRequest, current_user: TokenData = Depends(get_current_user)):
    """Add a new brand to watch list."""
    async with get_db() as db:
        try:
            new_brand = WatchedBrand(
                name=req.name,
                user_id=current_user.id,
                keywords=req.keywords,
                region=req.region or "india"
            )
            db.add(new_brand)
            await db.commit()
            return {"status": "success", "brand": req.name}
        except Exception:
            await db.rollback()
            raise HTTPException(400, "Brand already being watched or error occurred")

@router.put("/{name}")
async def update_brand(name: str, req: BrandRequest, current_user: TokenData = Depends(get_current_user)):
    """Update brand keywords and region."""
    async with get_db() as db:
        stmt = (
            update(WatchedBrand)
            .where(WatchedBrand.name == name)
            .where(WatchedBrand.user_id == current_user.id)
            .values(keywords=req.keywords, region=req.region or "india")
        )
        await db.execute(stmt)
        await db.commit()
        return {"status": "updated", "brand": name}

@router.delete("/{name}")
async def delete_brand(name: str, current_user: TokenData = Depends(get_current_user)):
    """Stop watching a brand."""
    async with get_db() as db:
        stmt = delete(WatchedBrand).where(WatchedBrand.name == name).where(WatchedBrand.user_id == current_user.id)
        await db.execute(stmt)
        await db.commit()
        return {"status": "deleted", "brand": name}

@router.post("/scrape")
async def trigger_brand_scrape(region: str = "india", days: int = 1, current_user: TokenData = Depends(get_current_user)):
    """Trigger a scale using Celery."""
    async with get_db() as db:
        res = await db.execute(select(WatchedBrand).where(WatchedBrand.user_id == current_user.id))
        brands = res.scalars().all()
        if not brands:
            raise HTTPException(400, "No brands to scrape.")
        
        job_id = str(uuid.uuid4())
        date_to = date.today()
        date_from = date_to - timedelta(days=days)

        new_job = ScrapeJob(
            id=job_id, sector="brand_tracker", region=region,
            user_id=current_user.id, date_from=date_from, date_to=date_to,
            status='pending', started_at=datetime.now()
        )
        db.add(new_job)
        await db.commit()

        # Simplified for orchestrated engine
        celery_app.send_task(
            "scraper.tasks.run_scrape_task",
            args=[job_id, "brand_tracker", region, str(date_from), str(date_to), "broad", current_user.id]
        )
        return {"job_id": job_id, "status": "pending"}

@router.post("/scrape/{name}")
async def trigger_individual_brand_scrape(name: str, days: int = 1, current_user: TokenData = Depends(get_current_user)):
    """Trigger a scrape for a single brand."""
    async with get_db() as db:
        res = await db.execute(
            select(WatchedBrand)
            .where(func.lower(WatchedBrand.name) == func.lower(name))
            .where(WatchedBrand.user_id == current_user.id)
        )
        brand = res.scalar_one_or_none()
        if not brand:
            raise HTTPException(404, f"Brand node '{name}' not found for this user")
        
        job_id = str(uuid.uuid4())
        date_to = date.today()
        date_from = date_to - timedelta(days=days)

        new_job = ScrapeJob(
            id=job_id, sector=brand.name, region=brand.region or "india",
            user_id=current_user.id, date_from=date_from, date_to=date_to,
            status='pending', started_at=datetime.now()
        )
        db.add(new_job)
        await db.commit()

        celery_app.send_task(
            "scraper.tasks.run_scrape_task",
            args=[job_id, brand.name, brand.region or "india", str(date_from), str(date_to), "broad", current_user.id]
        )
        return {"job_id": job_id, "status": "pending"}

@router.get("/download/{name}")
async def download_brand_articles(
    name: str, 
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    current_user: TokenData = Depends(get_current_user)
):
    """Download brand articles as CSV."""
    async with get_db() as db:
        stmt = select(Article).where(Article.sector == name).where(Article.user_id == current_user.id)
        if date_from: stmt = stmt.where(Article.published_at >= date_from)
        if date_to: stmt = stmt.where(Article.published_at <= date_to)
        stmt = stmt.order_by(Article.published_at.desc())
        
        res = await db.execute(stmt)
        articles = res.scalars().all()
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Title", "URL", "Agency", "Published_At", "Summary"])
        
        for a in articles:
            writer.writerow([a.title, a.url, a.agency, a.published_at, a.summary])
            
        output.seek(0)
        return StreamingResponse(
            output, media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={name}_articles.csv"}
        )

@router.get("/download/{name}/excel")
async def download_brand_articles_excel(
    name: str, 
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    current_user: TokenData = Depends(get_current_user)
):
    """Download brand articles as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import io

    async with get_db() as db:
        stmt = (
            select(Article)
            .where(Article.sector == name)
            .where(Article.user_id == current_user.id)
            .order_by(Article.published_at.desc())
        )
        if date_from: stmt = stmt.where(Article.published_at >= date_from)
        if date_to: stmt = stmt.where(Article.published_at <= date_to)
        
        res = await db.execute(stmt)
        articles = res.scalars().all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Nexus_{name}"

        # Standard Premium Headers
        headers = ["Title", "Resolved URL", "Publisher/Agency", "Author", "Summary", "Published At", "Source Feed", "Keyword Matched"]
        ws.append(headers)

        # Formatting Header
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"
        alternate_fill = PatternFill(start_color="F0F4FA", end_color="F0F4FA", fill_type="solid")

        for i, a in enumerate(articles, start=2):
            published_str = a.published_at.strftime("%d %b %Y %H:%M") if a.published_at else ""
            row_data = [
                a.title,
                a.resolved_url or a.url,
                a.agency,
                a.author or "Staff Reporter",
                a.summary,
                published_str,
                a.source_feed or "brand_tracker",
                name
            ]
            ws.append(row_data)
            if i % 2 == 0:
                for cell in ws[i]: cell.fill = alternate_fill
            
            url_cell = ws.cell(row=i, column=2)
            url_cell.hyperlink = a.resolved_url or a.url
            url_cell.font = Font(color="0000FF", underline="single")

        # Auto-fit
        for col in range(1, len(headers) + 1):
            column = get_column_letter(col)
            ws.column_dimensions[column].width = 25 # Standard width for brand report

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"NEXUS_Brand_{name}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        return StreamingResponse(
            output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
