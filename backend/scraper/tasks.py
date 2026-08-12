import logging
from typing import Optional, List, Dict, Any
import json
import os
import httpx
import trafilatura
import random
from datetime import datetime
from celery_app import app as celery_app
from db.database import get_db_sync, Article, ScrapeJob, IrrelevantArticle, JobFunnelLog, Client
from scraper.browser import scrape_url
from sqlalchemy import select, update, insert, delete
from scraper.engine import normalize_url
from scraper.llm import get_redis_sync
from scraper.search_utils import verify_boolean_relevance

logger = logging.getLogger(__name__)

import openpyxl
import re

def is_cell_colored(cell):
    fill = cell.fill
    if not fill or not fill.fill_type:
        return False
    fg = fill.fgColor
    if fg and fg.type == 'rgb':
        return fg.rgb and fg.rgb != '00000000' and fg.rgb != 'FFFFFFFF'
    return False

def is_cell_black(cell):
    fill = cell.fill
    if not fill or not fill.fill_type:
        return False
    fg = fill.fgColor
    return fg and fg.type == 'rgb' and fg.rgb == 'FF000000'

def parse_keyword_string(s):
    if not s:
        return []
    s = s.replace('\u2018', '"').replace('\u2019', '"')
    s = s.replace('\u201c', '"').replace('\u201d', '"')
    s = s.replace('\u201b', '"').replace('\u201f', '"')
    s = s.replace('`', '"')
    
    quotes = re.findall(r'"([^"]*)"', s)
    if quotes:
        return [q.strip() for q in quotes if q.strip()]
    parts = []
    for p in re.split(r'[,\n]', s):
        p_clean = p.strip().strip('"\'')
        if p_clean:
            parts.append(p_clean)
    return parts

def is_valid_keyword_cell(val):
    if not val:
        return False
    val_str = str(val).strip()
    has_quotes = any(q in val_str for q in ('"', '“', '”', '\u201c', '\u201d'))
    has_boolean = " AND " in val_str or " OR " in val_str
    is_header_desc = "general format" in val_str.lower() or "use boolean" in val_str.lower() or "keywords industry" in val_str.lower() or "industry & competition" in val_str.lower()
    return (has_quotes or has_boolean) and not is_header_desc

def load_kws_from_excel(filepath, omit_black=False, use_only_colored=False):
    if not os.path.exists(filepath):
        logger.warning(f"Keywords file not found: {filepath}")
        return []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        structure = []
        current_master = None
        for r in range(1, ws.max_row + 1):
            c1 = ws.cell(r, 1)
            c2 = ws.cell(r, 2)
            val1 = c1.value
            val2 = c2.value
            if not val1 and not val2:
                continue
            if omit_black and (is_cell_black(c1) or is_cell_black(c2)):
                continue
                
            if use_only_colored:
                if val2 and not is_cell_colored(c2) and is_valid_keyword_cell(val2):
                    continue
                    
            val2_str = str(val2) if val2 else ""
            is_master = val1 and (not val2 or "keywords" in val2_str.lower() or "notes" in val2_str.lower() or not is_valid_keyword_cell(val2))
            if is_master:
                current_master = str(val1).strip()
            else:
                if val1 and val2 and current_master:
                    if is_valid_keyword_cell(val2):
                        structure.append({
                            "master": current_master,
                            "sub": str(val1).strip(),
                            "raw_cell_value": str(val2).strip()
                        })
        return structure
    except Exception as e:
        logger.error(f"Error loading keywords from {filepath}: {e}")
        return []

def _is_product_section(section_name: str) -> bool:
    """
    Returns True if the CSV section name belongs to the 'product' bucket
    (Product & consumer, Ads & monetization, Workspace).
    Everything else is treated as 'corporate'.
    """
    name_lower = section_name.lower().strip()
    return (
        name_lower.startswith("product & consumer")
        or name_lower.startswith("ads &")
        or name_lower.startswith("workspace")
    )


def parse_super_final_csv(filepath: str) -> dict:
    """
    Parses Google_keywords_super_final.csv and returns:
      {
        "corporate": [{"master": ..., "sub": ..., "raw_cell_value": ...}, ...],
        "product":   [{"master": ..., "sub": ..., "raw_cell_value": ...}, ...],
      }

    CSV structure:
      - A 'section header' row has a non-empty col1 and an empty (or header-descriptor) col2.
      - A 'data' row has a non-empty col1 (sub-label) AND a non-empty col2 (keyword string).
      - Rows where both col1 and col2 are empty are separators — ignored.

    The col2 keyword string is preserved verbatim as 'raw_cell_value' so that the
    existing parse_keyword_string / evaluate_headline_relevance pipeline can consume it.
    """
    import csv as _csv

    if not os.path.exists(filepath):
        logger.warning(f"[Heavy] Super-final CSV not found: {filepath}")
        return {"corporate": [], "product": []}

    corporate_entries = []
    product_entries   = []

    # Header-descriptor strings that signal a master-heading row rather than a data row
    _HEADER_HINTS = (
        "keywords", "notes", "use boolean", "general format",
        "keywords industry", "industry & competition",
    )

    try:
        with open(filepath, newline="", encoding="utf-8-sig") as f:
            reader = _csv.reader(f)
            current_master = None
            current_bucket = "corporate"  # default

            for row in reader:
                # Normalise: pad to at least 2 columns
                col1 = row[0].strip() if len(row) > 0 else ""
                col2 = row[1].strip() if len(row) > 1 else ""

                # Skip fully empty rows
                if not col1 and not col2:
                    continue

                # Determine if this is a section-header row:
                #   col1 has content, col2 is either empty OR a descriptor hint (not actual keywords)
                col2_lower = col2.lower()
                is_header_desc = any(h in col2_lower for h in _HEADER_HINTS)
                has_keyword_content = (
                    col2
                    and not is_header_desc
                    and any(q in col2 for q in ('"', '\u201c', '\u201d', "'"))
                )

                if col1 and (not col2 or is_header_desc or not has_keyword_content):
                    # This is a master section header row
                    current_master = col1.strip()
                    current_bucket = "product" if _is_product_section(current_master) else "corporate"
                    continue

                # Data row: col1 = sub-label, col2 = keyword string
                if col1 and col2 and current_master and has_keyword_content:
                    entry = {
                        "master": current_master,
                        "sub": col1.strip(),
                        "raw_cell_value": col2,
                    }
                    if current_bucket == "product":
                        product_entries.append(entry)
                    else:
                        corporate_entries.append(entry)

        logger.info(
            f"[Heavy] parse_super_final_csv: corporate={len(corporate_entries)} entries, "
            f"product={len(product_entries)} entries from {filepath}"
        )
    except Exception as e:
        logger.error(f"[Heavy] Failed to parse super-final CSV {filepath}: {e}", exc_info=True)

    return {"corporate": corporate_entries, "product": product_entries}


def match_keyword_item(k, headline_lower):
    k_lower = k.lower().strip()
    if not k_lower:
        return False, 0
        
    if " and " not in k_lower and " or " not in k_lower:
        clean_k = k_lower.strip('"\'() ')
        if not any(char.isalnum() for char in clean_k):
            return False, 0
        return clean_k in headline_lower, len(clean_k.split())
        
    if " or " in k_lower:
        parts = [p.strip().strip('"\'() ') for p in k_lower.split(" or ")]
        for p in parts:
            if p and any(c.isalnum() for c in p) and p in headline_lower:
                return True, len(p.split())
        return False, 0
        
    if " and " in k_lower:
        parts = [p.strip().strip('"\'() ') for p in re.split(r'\s+and\s+', k_lower)]
        parts_clean = []
        for p in parts:
            p_clean = re.sub(r'near/\d+', '', p).strip().strip('"\'() ')
            if p_clean and any(c.isalnum() for c in p_clean):
                parts_clean.append(p_clean)
        if parts_clean and all(p in headline_lower for p in parts_clean):
            words_count = sum(len(p.split()) for p in parts_clean)
            return True, words_count
        return False, 0
    return False, 0

def evaluate_headline_relevance(headline, keywords_list):
    if not headline:
        return 0, []
    headline_lower = headline.lower()
    matched_subs = []
    total_score = 0
    
    for entry in keywords_list:
        sub_heading = entry["sub"]
        master_heading = entry["master"]
        raw_cell = entry["raw_cell_value"]
        
        items = []
        if "," in raw_cell and '"' in raw_cell:
            items = parse_keyword_string(raw_cell)
        else:
            items = [raw_cell]
            
        cell_matched = False
        cell_score = 0
        matched_items = []
        
        for item in items:
            matched, score = match_keyword_item(item, headline_lower)
            if matched:
                cell_matched = True
                cell_score = max(cell_score, score)
                matched_items.append(item)
                
        if cell_matched:
            points = 2 if cell_score <= 1 else (5 if cell_score == 2 else 8)
            total_score += points
            matched_subs.append({
                "master": master_heading,
                "sub": sub_heading,
                "score": points,
                "matched_items": matched_items
            })
            
    confidence = min(total_score, 10)
    return confidence, matched_subs

def normalize_publication_name(name):
    if not name:
        return ""
    name = re.sub(r'\(.*?\)', '', name)
    name = name.lower().strip()
    if name.startswith("the "):
        name = name[4:]
    return "".join(c for c in name if c.isalnum())

def is_indian_article(agency_name: str, url: str) -> bool:
    name = (agency_name or "").strip().lower()
    url_lower = (url or "").strip().lower()
    if ".in/" in url_lower or url_lower.endswith(".in") or ".co.in" in url_lower or ".indiatimes.com" in url_lower:
        return True
    indian_keywords = [
        "timesofindia", "times of india", "delhitimes", "delhi times", "mumbaitimes", "mumbai times", "bombaytimes", "bombay times",
        "hindustantimes", "hindustan times", "ht brunch", "htbrunch", "indianexpress", "indian express", "thehindu", "the hindu",
        "economictimes", "economic times", "et now", "etnow", "et prime", "etprime", "et wealth", "etwealth", "ettravelworld",
        "indiatoday", "india today", "zeenews", "zee news", "news18", "zeebiz", "zee business", "cnbctv18", "cnbc tv18",
        "moneycontrol", "livemint", "business-standard", "business standard", "financialexpress", "financial express",
        "thehindubusinessline", "business line", "businessline", "forbesindia", "forbes india", "fortuneindia", "fortune india",
        "businesstoday", "business today", "theweek.in", "the week", "outlookmoney", "outlook money", "press trust of india",
        "pti.in", "cntraveller", "condé nast", "conde nast", "natgeotraveller", "travelandleisureindia", "curlytales", "curly tales",
        "outlooktraveller", "outlook traveller", "the-ken", "the ken", "yourstory", "inc42", "vccircle", "story18", "ians.in", "ians",
        "newindianexpress", "new indian express", "deccanherald", "deccan herald", "tribuneindia", "the tribune", "telegraphindia",
        "telegraph india", "the telegraph", "deccanchronicle", "deccan chronicle", "news9live", "news9", "travelandtourworld",
        "travelbizmonitor", "hotelierindia", "hotelier india", "bwhotelier", "indianretailer", "indian retailer", "tradebrains",
        "livefromalounge", "traveltrendstoday", "bottindia", "hospitalitybizindia", "heraldgoa", "oheraldo", "uniindia",
        "united news of india", "thehansindia", "the hans india"
    ]
    if any(k in name for k in indian_keywords) or any(k in url_lower for k in indian_keywords):
        return True
    return False

def load_priority_media_list(filepath):
    if not os.path.exists(filepath):
        logger.warning(f"Priority media list not found: {filepath}")
        return []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        pubs = []
        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, 2).value
            if val:
                val = str(val).strip()
                c1 = ws.cell(r, 1).value
                if c1 is not None:
                    pubs.append(val)
        return pubs
    except Exception as e:
        logger.error(f"Error loading priority media list: {e}")
        return []

def is_priority_publication(agency_name, priority_list):
    if not agency_name or not priority_list:
        return False
    norm_agency = normalize_publication_name(agency_name)
    for p in priority_list:
        norm_p = normalize_publication_name(p)
        if norm_p == norm_agency or norm_p in norm_agency or norm_agency in norm_p:
            return True
    return False

def group_articles_by_sections(articles: list[dict], company_name: str) -> dict[str, list[dict]]:
    TARGET_SECTIONS = [
        "Competition Brand Keywords, Company Keywords, Competition Keywords, Brand Names, Competition Brand keywords, Company keywords",
        "Company Keywords, Google Pay keywords, Brand Names",
        "Competition Brand keywords, Company Keywords, Brand Names",
        "Industry keywords",
        "Industry Keywords",
        "Youtube keywords, Competition Keywords, Mobile OS & Hardware, Competition Brand Keywords",
        "Company keywords, Brand Names, Company Keywords",
        "Company Keywords, Brand Names, Google Maps keywords",
        "Mobile OS & Hardware",
        "Brand Names"
    ]
    by_pillar = {}
    if "google" in company_name.lower():
        # Initialize all target sections to preserve the correct order
        for sec in TARGET_SECTIONS:
            by_pillar[sec] = []
        for art in articles:
            sub_category_str = art.get("_sub_category") or ""
            subs = [s.strip().lower() for s in sub_category_str.split(",") if s.strip()]
            
            mapped_section = None
            if "google pay keywords" in subs:
                mapped_section = "Company Keywords, Google Pay keywords, Brand Names"
            elif "google maps keywords" in subs:
                mapped_section = "Company Keywords, Brand Names, Google Maps keywords"
            elif "youtube keywords" in subs:
                mapped_section = "Youtube keywords, Competition Keywords, Mobile OS & Hardware, Competition Brand Keywords"
            elif "mobile os & hardware" in subs:
                mapped_section = "Mobile OS & Hardware"
            elif "competition brand keywords" in subs or "competition brand keywords" in subs:
                if any(x in subs for x in ["google - gemini app", "openai", "chatgpt", "claude", "deepmind"]):
                    mapped_section = "Competition Brand keywords, Company Keywords, Brand Names"
                else:
                    mapped_section = "Competition Brand Keywords, Company Keywords, Competition Keywords, Brand Names, Competition Brand keywords, Company keywords"
            elif "industry keywords" in subs:
                mapped_section = "Industry Keywords"
            elif "company keywords" in subs or "company keywords" in subs:
                mapped_section = "Company keywords, Brand Names, Company Keywords"
            elif "brand names" in subs:
                mapped_section = "Brand Names"
            else:
                if any(x in subs for x in ["google watch", "google earbuds", "wearables"]):
                    mapped_section = "Mobile OS & Hardware"
                elif any(x in subs for x in ["google - gemini app", "openai", "chatgpt", "claude", "deepmind"]):
                    mapped_section = "Competition Brand keywords, Company Keywords, Brand Names"
                elif any(x in subs for x in ["youtube"]):
                    mapped_section = "Youtube keywords, Competition Keywords, Mobile OS & Hardware, Competition Brand Keywords"
                else:
                    mapped_section = "Competition Brand Keywords, Company Keywords, Competition Keywords, Brand Names, Competition Brand keywords, Company keywords"
            
            if mapped_section:
                by_pillar[mapped_section].append(art)
            else:
                by_pillar["Brand Names"].append(art)
                
        # Filter out empty sections
        by_pillar = {k: v for k, v in by_pillar.items() if v}
    else:
        for art in articles:
            pillar = art.get("_pillar") or "Other"
            if pillar not in by_pillar:
                by_pillar[pillar] = []
            by_pillar[pillar].append(art)
            
    return by_pillar

PLAYWRIGHT_ONLY_DOMAINS = {"axios.com", "ndtv.com"}

# Rotate across several recent Chrome UA strings to avoid fingerprinting.
_UA_POOL = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.6533.120 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.6533.120 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.6478.229 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.6533.120 Safari/537.36",
]

def _pick_ua() -> str:
    return random.choice(_UA_POOL)

def should_use_playwright(url: str) -> bool:
    try:
        from urllib.parse import urlparse
        domain = urlparse(url).netloc.lower()
        if domain.startswith("www."):
            domain = domain[4:]
        return any(domain == d or domain.endswith("." + d) for d in PLAYWRIGHT_ONLY_DOMAINS)
    except Exception:
        return False

KNOWN_PAYWALLED_DOMAINS = {
    "the-ken.com", "livemint.com", "economictimes.indiatimes.com",
    "business-standard.com", "financialexpress.com",
    "wsj.com", "ft.com",
    "reuters.com", "bloomberg.com", "thehindu.com", "indianexpress.com",
    "moneycontrol.com", "timesofindia.indiatimes.com", "hindustantimes.com",
}

PAYWALL_HTML_INDICATORS = [
    "paywall", "premium-content", "subscription-wall", "subscribe-gate",
    "metered-content", "locked-content", "premium-article",
    "subscribe to continue reading", "exclusive to subscribers",
    "premium article", "this article is for subscribers",
    "you have reached your free article limit",
    "sign up to read this article", "become a member to read",
]

def detect_paywall(url: str, html_content: str, body_text: str) -> bool:
    """
    Detects if an article is behind a paywall using three methods:
    1. Known paywalled domain matching
    2. HTML content heuristic scanning for paywall indicators
    3. Body length anomaly (short body text but large HTML)
    """
    try:
        from urllib.parse import urlparse
        domain = urlparse(url).netloc.lower()
        if domain.startswith("www."):
            domain = domain[4:]
        # 1. Known domain check
        if any(domain == d or domain.endswith("." + d) for d in KNOWN_PAYWALLED_DOMAINS):
            return True
    except Exception:
        pass

    # 2. HTML heuristic check
    if html_content:
        html_lower = html_content[:20000].lower()  # Only scan first 20KB for performance
        for indicator in PAYWALL_HTML_INDICATORS:
            if indicator in html_lower:
                return True

    # 3. Body length anomaly: short extracted text from a large HTML page
    if html_content and body_text:
        if len(body_text) < 200 and len(html_content) > 5000:
            return True

    return False

def _mark_article_processed(job_id: str):
    """
    Safely increment total_scraped and mark job as completed if all articles processed.
    This MUST be called on EVERY code path, success or failure, so jobs always complete.
    """
    try:
        with get_db_sync() as db:
            db.execute(
                update(ScrapeJob)
                .where(ScrapeJob.id == job_id)
                .values(total_scraped=ScrapeJob.total_scraped + 1)
            )
            db.commit()

            job = db.execute(select(ScrapeJob).where(ScrapeJob.id == job_id)).scalar_one_or_none()
            if job and job.total_found > 0 and job.total_scraped >= job.total_found:
                db.execute(
                    update(ScrapeJob)
                    .where(ScrapeJob.id == job_id)
                    .values(status='completed', current_phase='Completed', completed_at=datetime.now())
                )
                db.commit()
                logger.info(f"Job {job_id} completed: {job.total_scraped}/{job.total_found}")
    except Exception as e:
        logger.error(f"Error marking article processed for job {job_id}: {e}")

def _increment_funnel_metric(job_id: str, field_name: str, amount: int = 1):
    if not job_id:
        return
    try:
        with get_db_sync() as db:
            log_entry = db.execute(select(JobFunnelLog).where(JobFunnelLog.job_id == job_id)).scalar_one_or_none()
            if not log_entry:
                log_entry = JobFunnelLog(job_id=job_id)
                db.add(log_entry)
                db.commit()
                db.refresh(log_entry)
            
            val = getattr(log_entry, field_name, 0) or 0
            setattr(log_entry, field_name, val + amount)
            db.commit()
    except Exception as e:
        logger.error(f"Error logging funnel metric for job {job_id}: {e}")


# ─── Orchestrator Task ────────────────────────────────────────────────────────

@celery_app.task(name="scraper.tasks.run_scrape_task", bind=True)
def run_scrape_task(self, job_id, sector, region, date_from, date_to, search_mode, user_id):
    """
    Orchestrator: Discovers URLs and dispatches independent scraping nodes.
    Now fully synchronous for gevent compatibility.
    Celery tasks run server-side and persist through user logout.
    """
    logger.info(f"Starting Orchestrator for job {job_id}")
    from scraper.engine import run_scrape_job
    try:
        run_scrape_job(
            job_id=job_id,
            sector=sector,
            region=region,
            date_from=date_from,
            date_to=date_to,
            search_mode=search_mode,
            user_id=user_id
        )
        logger.info(f"Discovery phase for job {job_id} completed.")
    except Exception as e:
        logger.error(f"Orchestrator failed for job {job_id}: {e}")
        raise e

# ─── Scraper Node (I/O Intensive) ─────────────────────────────────────────────

@celery_app.task(
    name="scraper.tasks.scrape_article_node",
    bind=True,
    rate_limit="100/m",
    max_retries=10,
    retry_backoff=15,    # Exponential backoff starting at 15s
    retry_jitter=True    # Jitter backoff offsets
)
def scrape_article_node(self, article_data, job_id, sector, region, user_id):
    """
    Task Node 1: Fetches HTML and extracts raw body. 
    Synchronous for gevent compatibility.
    Runs server-side independently of any user session.
    """
    from scraper.engine import scrape_only, is_job_cancelled
    from scraper.google_news import resolve_google_news_url_sync
    try:
        if is_job_cancelled(job_id):
            logger.info(f"Scrape task halted for job {job_id} [Reason: Job Cancelled/Global Stop]")
            _mark_article_processed(job_id)
            return None

        # Resolve Google News redirect in sync (httpx)
        url = article_data.get("url") or article_data.get("link")
        if not url:
            logger.warning(f"Task received article without URL: {article_data}")
            _mark_article_processed(job_id)
            return None
            
        resolved_url = resolve_google_news_url_sync(url)
        if not resolved_url:
            logger.warning(f"Could not resolve URL: {url}")
            _mark_article_processed(job_id)
            return None
        
        # ─── URL NORMALIZATION ───
        normalized_url = normalize_url(resolved_url)
        
        # ─── DB CACHE CHECK ───
        with get_db_sync() as db:
            # Check existing
            existing_article = db.execute(
                select(Article.id)
                .where(Article.url == normalized_url)
                .where(Article.user_id == user_id)
            ).first()
            if existing_article:
                logger.info(f"Celery DB Cache HIT (already exists): {normalized_url}. Skipping.")
                _mark_article_processed(job_id)
                _increment_funnel_metric(job_id, "cache_skipped")
                return None
                
            # Check irrelevant
            cached_irrelevant = db.execute(
                select(IrrelevantArticle)
                .where(IrrelevantArticle.url == normalized_url)
            ).scalar_one_or_none()
            if cached_irrelevant:
                age = datetime.now() - cached_irrelevant.last_seen_at
                if age.days < 30:
                    logger.info(f"Celery Irrelevant Cache HIT: {normalized_url}. Skipping.")
                    _mark_article_processed(job_id)
                    _increment_funnel_metric(job_id, "cache_skipped")
                    return None
                else:
                    db.delete(cached_irrelevant)
                    db.commit()

        # ─── LOCKING ───
        lock_key = f"lock:scrape:{normalized_url}"
        r = get_redis_sync()
        if not r.set(lock_key, job_id, nx=True, ex=600):
            logger.info(f"Task overlap detected for {normalized_url}. Skipping redundant node.")
            _mark_article_processed(job_id)
            return None

        # ─── DOMAIN RATE LIMITING (Throttling) ───
        from urllib.parse import urlparse
        domain = urlparse(resolved_url).netloc.lower()
        if domain.startswith("www."):
            domain = domain[4:]
            
        domain_lock_key = f"lock:domain:{domain}"
        if not r.set(domain_lock_key, job_id, nx=True, ex=2):
            logger.info(f"Domain rate limit triggered for {domain}. Rescheduling task.")
            # Retry after a short randomized jitter countdown (3-7s)
            raise self.retry(countdown=random.randint(3, 7))

        # ─── FAST-REJECT PRE-FILTER ───
        keywords = []
        with get_db_sync() as db:
            from db.database import WatchedBrand
            brand_obj = db.execute(
                select(WatchedBrand)
                .where(WatchedBrand.name == sector)
                .where(WatchedBrand.user_id == user_id)
            ).first()
            if brand_obj:
                brand_obj = brand_obj[0]
                if brand_obj.keywords:
                    keywords = [k.strip() for k in brand_obj.keywords.split(",") if k.strip()]
        
        if not keywords:
            from scraper.config import SECTOR_KEYWORDS
            keywords.extend(SECTOR_KEYWORDS.get(sector.lower(), []))
            
        title = article_data.get("title", "")
        description = article_data.get("description", "")
        text_to_check = f"{title} {description}"
        
        # Check publication category first to bypass fast-reject for Category A
        from scraper.search_utils import match_publication_category
        url_to_check = article_data.get("url") or article_data.get("link")
        agency_to_check = article_data.get("agency") or "News"
        pub_category = match_publication_category(agency_to_check, url_to_check)
        is_cat_a = (pub_category == "A")

        # Fast-reject keyword pre-filter is temporarily disabled for maximum recall
        is_pre_filtered_relevant = True
        # Original keyword filter (commented out to allow switching back):
        # if is_cat_a:
        #     is_pre_filtered_relevant = True
        # elif keywords:
        #     is_pre_filtered_relevant = verify_boolean_relevance(text_to_check, keywords)
        #     if not is_pre_filtered_relevant and sector.lower() in text_to_check.lower():
        #         is_pre_filtered_relevant = True
        #         
        # if not is_pre_filtered_relevant:
        #     logger.info(f"Fast-reject (Celery): article '{title}' failed keyword pre-filter. Skipping.")
        #     with get_db_sync() as db:
        #         db.merge(IrrelevantArticle(url=normalized_url, last_seen_at=datetime.now()))
        #         db.commit()
        #     _mark_article_processed(job_id)
        #     _increment_funnel_metric(job_id, "pre_filter_dropped")
        #     return None

        # --- FAST-TRACK SCRAPING (httpx + trafilatura) ---
        html = None
        if not should_use_playwright(resolved_url):
            try:
                headers = {
                    "User-Agent": _pick_ua(),
                    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
                    "Accept-Language": "en-US,en;q=0.9",
                    "Referer": "https://news.google.com/",
                    "Cache-Control": "no-cache",
                }
                with httpx.Client(timeout=httpx.Timeout(15.0, connect=8.0), follow_redirects=True) as client:
                    resp = client.get(resolved_url, headers=headers)
                    if resp.status_code == 429:
                        logger.info(f"Rate limited (429) by {resolved_url}, skipping to Playwright")
                    elif resp.status_code == 200:
                        text_content = trafilatura.extract(resp.text)
                        if text_content and len(text_content) > 400:
                            logger.info(f"Fast-track success for {resolved_url} ({len(text_content)} chars)")
                            html = resp.text
                        else:
                            logger.debug(f"Fast-track got HTML but <400 chars text ({len(text_content or '')} chars), falling to Playwright")
            except Exception as e:
                logger.debug(f"Fast-track failed for {resolved_url}: {e}")
        else:
            logger.info(f"Skipping fast-track HTTP client for known hostile domain: {resolved_url}")

        # --- FALLBACK: SUBPROCESS BROWSER (fully isolated from Gevent) ---
        if not html:
            logger.info(f"Falling back to Playwright for {resolved_url}")
            html = scrape_url(resolved_url)
            
        if not html:
            logger.warning(f"Scrape failed (both fast-track and browser) for {resolved_url}")
            _mark_article_processed(job_id)
            return None
            
        _increment_funnel_metric(job_id, "scraped_count")

        # Move processed data back to article_data for Engine
        article_data["resolved_url"] = resolved_url
        article_data["raw_html"] = html

        article_id = scrape_only(article_data, job_id, sector, region, user_id)
        if article_id:
            logger.info(f"Scraped raw content for article {article_id}. Triggering enrichment...")
            enrich_article_node.delay(article_id)
    except Exception as e:
        logger.error(f"Scrape node failed for {article_data.get('url')}: {e}")
        # On final retry failure, still increment so job doesn't hang
        if self.request.retries >= self.max_retries:
            _mark_article_processed(job_id)
        raise self.retry(exc=e)

# ─── Enrichment Node (Compute Intensive) ──────────────────────────────────────

@celery_app.task(name="scraper.tasks.enrich_article_node", bind=True, max_retries=3)
def enrich_article_node(self, article_id):
    """
    Task Node 2: Performs AI analysis (Grok/Groq).
    Runs server-side, completely independent of user session.
    """
    from scraper.llm import perform_full_enrichment_sync, check_relevance_with_groq
    from scraper.engine import is_job_cancelled
    
    with get_db_sync() as db:
        res = db.execute(select(Article).where(Article.id == article_id))
        article = res.scalar_one_or_none()
        if not article or not article.full_body: return
        
        if is_job_cancelled(article.scrape_job_id):
            logger.info(f"Enrichment cancelled for job {article.scrape_job_id}. Skipping article {article_id}")
            return

        # ─── RELEVANCE CHECK (120B Model) ───
        try:
            keywords = []
            from db.database import WatchedBrand
            brand_obj = db.execute(
                select(WatchedBrand)
                .where(WatchedBrand.name == article.sector)
                .where(WatchedBrand.user_id == article.user_id)
            ).first()
            if brand_obj:
                brand_obj = brand_obj[0]
                if brand_obj.keywords:
                    keywords = [k.strip() for k in brand_obj.keywords.split(",") if k.strip()]
            
            if not keywords:
                from scraper.config import SECTOR_KEYWORDS
                keywords.extend(SECTOR_KEYWORDS.get(article.sector.lower(), []))
            
            # Retrieve client context guidelines if applicable
            client_context = ""
            client_summary_length = 35
            client_name = article.sector
            if " - " in client_name:
                client_name = client_name.split(" - ")[0]
            client_obj = db.execute(
                select(Client).where(Client.name == client_name)
            ).scalars().first()
            if client_obj:
                client_context = client_obj.context or ""
                client_summary_length = client_obj.summary_length or 35
            
            # --- Similarity Pre-Filter ---
            from scraper.similarity import evaluate_similarity_pre_filter, SIM_DROP_THRESHOLD
            sim_score = evaluate_similarity_pre_filter(
                article.title, article.full_body, keywords, client_context
            )
            # Cosine similarity pre-filter drop is temporarily disabled for maximum recall
            logger.info(f"Cosine similarity pre-filter drop bypassed for article {article_id}: '{article.title}' (score: {sim_score:.4f} < {SIM_DROP_THRESHOLD}).")

            # --- LLM Relevance Check ---
            is_semantic_relevant, verdict, reason, score = check_relevance_with_groq(
                article.title, article.full_body, keywords, client_name, client_context=client_context
            )
        except Exception as rel_err:
            logger.error(f"Relevance check exception in Celery task: {rel_err}")
            is_semantic_relevant = True # Fallback
            verdict = "uncertain"
            reason = f"Exception: {rel_err}"
            score = 0.5
            
        if not is_semantic_relevant:
            logger.info(f"Article {article_id} rejected by relevance check (verdict: {verdict}, score: {score}). Deleting and caching.")
            from db.database import IrrelevantArticle
            db.merge(IrrelevantArticle(
                url=article.url,
                title=article.title,
                description=article.summary or (article.full_body[:200] if article.full_body else ""),
                rejection_reason=reason,
                relevance_score=score,
                last_seen_at=datetime.now()
            ))
            db.delete(article)
            db.commit()
            _increment_funnel_metric(article.scrape_job_id, "relevance_no")
            return
            
        _increment_funnel_metric(article.scrape_job_id, "relevance_yes")

        # ─── SUMMARIZATION & ENRICHMENT ───
        try:
            enriched_data = perform_full_enrichment_sync(
                article.full_body, 
                article.title, 
                article.resolved_url or article.url, 
                article.sector,
                context_agency=article.agency,
                extra_metadata=article.extra_metadata,
                summary_length=client_summary_length
            )
            
            article.summary = enriched_data.get("summary")
            article.sentiment = enriched_data.get("sentiment")
            article.tags = enriched_data.get("tags")
            if enriched_data.get("agency"): article.agency = enriched_data.get("agency")
            if enriched_data.get("author"): article.author = enriched_data.get("author")
            
            db.commit()
            _increment_funnel_metric(article.scrape_job_id, "summarized_count")
            logger.info(f"Successfully enriched article {article_id}")
        except Exception as e:
            logger.error(f"AI Enrichment failed for article {article_id}: {e}")
            raise self.retry(exc=e, countdown=15)


# ─── Stale Job Watchdog (runs every 5 minutes via Celery Beat) ────────────────

@celery_app.task(name="scraper.tasks.complete_stale_jobs")
def complete_stale_jobs():
    """
    Watchdog: Scans for jobs stuck in 'running' state and marks complete if all articles are scraped.
    Ensures jobs finish even if some tasks crash silently.
    Runs every 5 minutes via Celery Beat schedule.
    """
    from datetime import datetime, timedelta
    from db.database import ClientRunLog
    
    # Redis Lock to prevent duplicate runs from concurrent schedulers
    try:
        r = get_redis_sync()
        current_minute = datetime.now().minute
        minute_block = current_minute - (current_minute % 5)
        lock_key = f"lock:scheduler:stale:{datetime.now().strftime('%Y-%m-%d-%H')}-{minute_block}"
        if not r.set(lock_key, "1", nx=True, ex=240):
            logger.info("Stale jobs cleanup already processed by another instance. Skipping.")
            return
    except Exception as lock_err:
        logger.warning(f"Failed to acquire Redis stale-jobs lock: {lock_err}")
    try:
        with get_db_sync() as db:
            stale_cutoff = datetime.now() - timedelta(minutes=10)
            running_jobs = db.execute(
                select(ScrapeJob).where(
                    ScrapeJob.status == 'running',
                    ScrapeJob.started_at < stale_cutoff,
                    ScrapeJob.total_found > 0
                )
            ).scalars().all()

            for job in running_jobs:
                # Force-complete if total_scraped is near total_found (within 3 to handle edge cases)
                if job.total_scraped >= max(0, job.total_found - 3):
                    db.execute(
                        update(ScrapeJob).where(ScrapeJob.id == job.id).values(
                            status='completed',
                            current_phase='Completed',
                            total_scraped=job.total_found,  # Correct the counter
                            completed_at=datetime.now()
                        )
                    )
                    logger.info(f"Watchdog force-completed stale job {job.id} ({job.total_scraped}/{job.total_found})")
            
            # Clean up stuck client run logs (older than 120 minutes)
            client_stale_cutoff = datetime.now() - timedelta(minutes=120)
            stale_client_logs = db.execute(
                select(ClientRunLog).where(
                    ClientRunLog.status == 'running',
                    ClientRunLog.started_at < client_stale_cutoff
                )
            ).scalars().all()
            
            for log_entry in stale_client_logs:
                db.execute(
                    update(ClientRunLog).where(ClientRunLog.id == log_entry.id).values(
                        status='failed',
                        error_message='Task timed out or was interrupted (Watchdog recovery)',
                        completed_at=datetime.now()
                    )
                )
                logger.info(f"Watchdog force-failed stale client run log {log_entry.id}")
            
            db.commit()
    except Exception as e:
        logger.error(f"Stale job watchdog error: {e}")


@celery_app.task(name="scraper.tasks.run_client_report_task")
def run_client_report_task(client_id: int):
    """
    Main background task to generate daily monitoring briefing reports for a client.
    """
    from db.database import get_db_sync, Client, ClientSection, ClientKeyword, ClientRecipient, ClientRunLog
    from scraper.engine import discover_articles
    from scraper.google_news import resolve_google_news_url_sync
    from scraper.report_generator import generate_docx_report
    from utils.google_docs import upload_docx_to_google_doc
    from utils.email import send_report_email, send_error_alert_email
    from scraper.llm import perform_full_enrichment_sync
    from datetime import date, datetime, timedelta
    import pytz
    import traceback
    
    client_name = f"Client ID {client_id}"
    logger.info(f"Starting client report task for client_id {client_id}")
    
    # Create a running log entry
    run_log_id = None
    with get_db_sync() as db:
        client = db.execute(select(Client).where(Client.id == client_id)).scalar_one_or_none()
        if not client:
            logger.error(f"Client with ID {client_id} not found.")
            return False
            
        run_log = ClientRunLog(
            client_id=client_id,
            status="running",
            started_at=datetime.utcnow()
        )
        db.add(run_log)
        db.commit()
        db.refresh(run_log)
        run_log_id = run_log.id
        client_name = client.name
        db_template_path = client.template_path
        
        # Resolve path dynamically to ensure environment portability
        template_path = None
        if db_template_path:
            templates_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "templates")
            os.makedirs(templates_dir, exist_ok=True)
            template_path = os.path.join(templates_dir, os.path.basename(db_template_path))
            
            # Restore from database if missing on disk (critical for container rebuilding/ephemeral storage)
            if getattr(client, "template_data", None):
                try:
                    with open(template_path, "wb") as buffer:
                        buffer.write(client.template_data)
                    logger.info(f"Dynamically restored template file from database: {template_path}")
                except Exception as e:
                    logger.error(f"Failed to restore template from database for client {client.id}: {e}")
            
        client_context = client.context
        client_timezone = client.timezone or "Asia/Kolkata"
        priority_media_list = getattr(client, "priority_media_list", None)
        region_filter = getattr(client, "region_filter", "All")
        intl_exceptions = getattr(client, "intl_exceptions", None)
        
    try:
        def _update_progress(msg: str):
            try:
                timestamp = datetime.now().strftime("%H:%M:%S")
                log_line = f"[{timestamp}] {msg}"
                with get_db_sync() as db:
                    run_log = db.execute(select(ClientRunLog).where(ClientRunLog.id == run_log_id)).scalar_one_or_none()
                    if run_log:
                        current_log = run_log.progress_message or ""
                        updated_log = f"{current_log}\n{log_line}" if current_log else log_line
                        db.execute(
                            update(ClientRunLog)
                            .where(ClientRunLog.id == run_log_id)
                            .values(progress_message=updated_log)
                        )
                        db.commit()
            except Exception as dberr:
                logger.error(f"Failed to update progress log in DB: {dberr}")

        _update_progress("Initializing client report profile...")
        
        # Load sections and keywords
        sections_data = {}
        all_emails = []
        
        with get_db_sync() as db:
            sections = db.execute(select(ClientSection).where(ClientSection.client_id == client_id)).scalars().all()
            recipients = db.execute(select(ClientRecipient).where(ClientRecipient.client_id == client_id)).scalars().all()
            all_emails = [r.email for r in recipients]
            
            for section in sections:
                kwd_objs = db.execute(select(ClientKeyword).where(ClientKeyword.section_id == section.id)).scalars().all()
                sections_data[section.name] = [k.keyword for k in kwd_objs]
                
        if not sections_data:
            raise ValueError("No sections or keywords configured for this client.")
            
        # Discover, scrape, verify and enrich articles for each section
        report_data_filtered = {} # {section_name: [list of article dicts]}
        report_data_master = {} # {section_name: [list of article dicts]}
        
        # Determine the search window based on the day of the week in client's timezone
        client_tz = pytz.timezone(client_timezone)
        run_date = datetime.now(client_tz).date()
        
        search_dates = [run_date]
        if run_date.weekday() == 0:  # 0 is Monday
            # On Monday, consolidate Friday, Saturday, Sunday, and Monday
            search_dates.extend([
                run_date - timedelta(days=1),  # Sunday
                run_date - timedelta(days=2),  # Saturday
                run_date - timedelta(days=3)   # Friday
            ])
        else:
            # On other weekdays, search today and yesterday
            search_dates.append(run_date - timedelta(days=1))
            
        # ── Phase 1: Discover articles for ALL sections in parallel ──────────────
        from scraper.engine import discover_direct_feeds
        from concurrent.futures import ThreadPoolExecutor as _TPE, as_completed as _ac

        def _discover_section(section_name_kws):
            sec_name, kws = section_name_kws
            if not kws:
                return sec_name, []
            job_id_disc = f"client_{client_id}_sec_{sec_name}"
            try:
                # Init funnel log
                try:
                    with get_db_sync() as db:
                        db.execute(delete(JobFunnelLog).where(JobFunnelLog.job_id == job_id_disc))
                        db.execute(insert(JobFunnelLog).values(job_id=job_id_disc, rss_discovered=0))
                        db.commit()
                except Exception as e:
                    logger.error(f"Failed to initialize funnel log for section '{sec_name}': {e}")

                discovered = []
                seen_urls = set()
                for target_date in search_dates:
                    date_job = f"client_{client_id}_sec_{sec_name}_{target_date.strftime('%Y%m%d')}"
                    for art in discover_articles(keywords=kws, day=target_date, geo="IN", region_name="india",
                                                 job_id=date_job, sector=f"{client_name} - {sec_name}"):
                        if art["url"] not in seen_urls:
                            discovered.append(art)
                            seen_urls.add(art["url"])
                    for art in discover_direct_feeds(keywords=kws, day=target_date, job_id=date_job,
                                                     sector=f"{client_name} - {sec_name}"):
                        if art["url"] not in seen_urls:
                            discovered.append(art)
                            seen_urls.add(art["url"])

                try:
                    with get_db_sync() as db:
                        db.execute(update(JobFunnelLog).where(JobFunnelLog.job_id == job_id_disc)
                                   .values(rss_discovered=len(discovered)))
                        db.commit()
                except Exception:
                    pass
                return sec_name, discovered
            except Exception as disc_err:
                # A discovery failure for one section must not crash the whole report.
                logger.error(f"Discovery failed for section '{sec_name}': {disc_err}", exc_info=True)
                return sec_name, []

        # Emit per-section discovery messages upfront (matches old log format)
        section_discoveries: dict = {}
        active_sections = {sn: kw for sn, kw in sections_data.items() if kw}
        for sec_name in active_sections:
            _update_progress(f"Discovering articles for section '{sec_name}'...")
            logger.info(f"Discovering articles for section '{sec_name}' with keywords: {active_sections[sec_name]}")

        if active_sections:
            disc_workers = min(len(active_sections), 4)
            with _TPE(max_workers=disc_workers) as disc_exe:
                disc_futures = {disc_exe.submit(_discover_section, item): item[0] for item in active_sections.items()}
                for fut in _ac(disc_futures):
                    sec_name, disc = fut.result()
                    section_discoveries[sec_name] = disc
                    logger.info(f"Discovery done for '{sec_name}': {len(disc)} articles found.")

        # ── Phase 2: Process articles per section (ALL sections in parallel) ──
        from concurrent.futures import ThreadPoolExecutor, as_completed

        def _process_section(section_name, keywords):
            if not keywords:
                return section_name, [], []

            discovered = section_discoveries.get(section_name, [])

            # Deduplicate by url
            unique_discovered = []
            seen_urls = set()
            for art in discovered:
                if art["url"] not in seen_urls:
                    unique_discovered.append(art)
                    seen_urls.add(art["url"])

            filtered_section_articles = []
            master_section_articles = []

            # Resolve and scrape each article concurrently
            
            def _process_single_article(art_idx_tuple):
                art, idx = art_idx_tuple
                raw_url = art["url"]
                title = art["title"]
                desc = art.get("description") or ""
                # Strip HTML tags from RSS description (Google News RSS contains raw HTML markup)
                if desc and "<" in desc:
                    import re
                    desc = re.sub(r'<[^>]+>', '', desc).strip()
                    desc = re.sub(r'&nbsp;', ' ', desc)
                    desc = re.sub(r'\s{2,}', ' ', desc).strip()
                agency = art.get("agency") or "News"
                
                job_id = f"client_{client_id}_sec_{section_name}"
                
                try:
                    # 1. Normalize raw URL for early checks
                    normalized_raw_url = normalize_url(raw_url)
                    
                    # 2. Check early cache for existing/irrelevant using raw URL
                    with get_db_sync() as db:
                        from db.database import IrrelevantArticle
                        # Check existing
                        existing_article = db.execute(
                            select(Article)
                            .where(Article.url == normalized_raw_url)
                        ).scalars().first()
                        if existing_article and existing_article.full_body and existing_article.summary:
                            logger.info(f"Early DB Cache HIT for {normalized_raw_url}")
                            _increment_funnel_metric(job_id, "cache_skipped")
                            
                            from scraper.search_utils import match_publication_category
                            cached_meta = existing_article.extra_metadata or {}
                            pub_category = cached_meta.get("publication_category")
                            if not pub_category:
                                pub_category = match_publication_category(existing_article.agency or agency, existing_article.resolved_url or existing_article.url)
                            
                            return {
                                "art_data": {
                                    "title": existing_article.title,
                                    "url": existing_article.resolved_url or existing_article.url,
                                    "agency": existing_article.agency or agency,
                                    "summary": existing_article.summary,
                                    "publication_category": pub_category,
                                    "published_at": existing_article.published_at.isoformat() if existing_article.published_at else None
                                },
                                "is_relevant_kw": True,
                                "is_semantic_relevant": True
                            }
                        
                        # Check irrelevant
                        cached_ir = db.execute(
                            select(IrrelevantArticle)
                            .where(IrrelevantArticle.url == normalized_raw_url)
                        ).scalar_one_or_none()
                        if cached_ir:
                            age = datetime.now() - cached_ir.last_seen_at
                            if age.days < 30:
                                logger.info(f"Early Irrelevant Cache HIT for {normalized_raw_url}")
                                _increment_funnel_metric(job_id, "cache_skipped")
                                from scraper.search_utils import match_publication_category
                                return {
                                    "art_data": {
                                        "title": cached_ir.title or title,
                                        "url": raw_url,
                                        "agency": agency,
                                        "summary": cached_ir.description or desc or "Irrelevant article cached.",
                                        "publication_category": match_publication_category(agency, raw_url)
                                    },
                                    "is_relevant_kw": True,
                                    "is_semantic_relevant": False
                                }
                            else:
                                db.delete(cached_ir)
                                db.commit()
                                
                    # 3. Triage pre-filter (fast reject)
                    text_to_check = f"{title} {desc}"
                    
                    # Check publication category first to bypass fast-reject for Category A
                    from scraper.search_utils import match_publication_category
                    pub_category = match_publication_category(agency, raw_url)
                    is_cat_a = (pub_category == "A")

                    # Fast-reject keyword pre-filter is temporarily disabled for maximum recall
                    is_relevant_kw = True
                    # Original keyword filter (commented out to allow switching back):
                    # if is_cat_a:
                    #     is_relevant_kw = True
                    # else:
                    #     is_relevant_kw = verify_boolean_relevance(text_to_check, keywords)
                    #     if not is_relevant_kw and section_name.lower() in text_to_check.lower():
                    #         is_relevant_kw = True
                    #     
                    # if not is_relevant_kw:
                    #     logger.info(f"Fast-reject: article '{title}' failed pre-filter.")
                    #     # Cache as irrelevant in DB
                    #     with get_db_sync() as db:
                    #         db.merge(IrrelevantArticle(url=normalized_raw_url, last_seen_at=datetime.now()))
                    #         db.commit()
                    #     _increment_funnel_metric(job_id, "pre_filter_dropped")
                    #     return None
                        
                    # 4. Resolve URL
                    logger.info(f"Resolving Google News URL: {raw_url}")
                    resolved_url = resolve_google_news_url_sync(raw_url)
                    if not resolved_url:
                        # Resolution failed — URL is still on Google's domain (redirect page).
                        # Scraping it would return a redirect notice page and falsely trigger
                        # paywall detection. Skip the article entirely.
                        logger.warning(f"Skipping unresolvable Google News URL: {raw_url}")
                        _increment_funnel_metric(job_id, "resolution_failed")
                        return None
                    normalized_url = normalize_url(resolved_url)
                    
                    # 5. Check cache for canonical URL
                    with get_db_sync() as db:
                        # Check existing
                        existing_article = db.execute(
                            select(Article)
                            .where(Article.url == normalized_url)
                        ).scalars().first()
                        if existing_article and existing_article.full_body and existing_article.summary:
                            logger.info(f"Canonical DB Cache HIT for {normalized_url}")
                            _increment_funnel_metric(job_id, "cache_skipped")
                            
                            from scraper.search_utils import match_publication_category
                            cached_meta = existing_article.extra_metadata or {}
                            pub_category = cached_meta.get("publication_category")
                            if not pub_category:
                                pub_category = match_publication_category(existing_article.agency or agency, existing_article.resolved_url or existing_article.url)
                                
                            return {
                                "art_data": {
                                    "title": existing_article.title,
                                    "url": existing_article.resolved_url or existing_article.url,
                                    "agency": existing_article.agency or agency,
                                    "summary": existing_article.summary,
                                    "publication_category": pub_category,
                                    "is_paywalled": False,
                                    "published_at": existing_article.published_at.isoformat() if existing_article.published_at else None
                                },
                                "is_relevant_kw": True,
                                "is_semantic_relevant": True
                            }
                            
                        # Check irrelevant
                        cached_ir = db.execute(
                            select(IrrelevantArticle)
                            .where(IrrelevantArticle.url == normalized_url)
                        ).scalar_one_or_none()
                        if cached_ir:
                            age = datetime.now() - cached_ir.last_seen_at
                            if age.days < 30:
                                logger.info(f"Canonical Irrelevant Cache HIT for {normalized_url}")
                                _increment_funnel_metric(job_id, "cache_skipped")
                                from scraper.search_utils import match_publication_category
                                return {
                                    "art_data": {
                                        "title": cached_ir.title or title,
                                        "url": resolved_url,
                                        "agency": agency,
                                        "summary": cached_ir.description or desc or "Irrelevant article cached.",
                                        "publication_category": match_publication_category(agency, resolved_url),
                                        "is_paywalled": False
                                    },
                                    "is_relevant_kw": True,
                                    "is_semantic_relevant": False
                                }
                            else:
                                db.delete(cached_ir)
                                db.commit()

                    # 6. Scrape HTML content
                    logger.info(f"Scraping content from resolved URL: {resolved_url}")
                    html_content = ""
                    _fast_body_chars = 0
                    if not should_use_playwright(resolved_url):
                        try:
                            headers = {
                                "User-Agent": _pick_ua(),
                                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
                                "Accept-Language": "en-US,en;q=0.9",
                                "Referer": "https://news.google.com/",
                                "Cache-Control": "no-cache",
                            }
                            with httpx.Client(follow_redirects=True, timeout=httpx.Timeout(12.0, connect=8.0)) as client_http:
                                resp = client_http.get(resolved_url, headers=headers)
                                if resp.status_code == 429:
                                    logger.info(f"Rate limited (429) by {resolved_url}, falling back to Playwright")
                                elif resp.status_code == 200:
                                    html_content = resp.text
                                    _fast_body_chars = len(trafilatura.extract(html_content) or "")
                        except Exception as e:
                            logger.warning(f"Fast HTTP scrape failed for {resolved_url}: {e}")
                    else:
                        logger.info(f"Skipping fast HTTP scrape for known hostile domain: {resolved_url}")

                    # Trigger Playwright if httpx failed entirely, or if it returned HTML
                    # but trafilatura could only extract <300 chars — indicates a JS-rendered
                    # or subscription-wall page that needs a real browser to get content.
                    if not html_content or _fast_body_chars < 300:
                        try:
                            logger.info(f"Falling back to Playwright for {resolved_url} (fast-track body: {_fast_body_chars} chars)")
                            pw_html = scrape_url(resolved_url)
                            if pw_html:
                                html_content = pw_html  # Use Playwright result; better than wall HTML
                        except Exception as e:
                            logger.error(f"Browser scrape failed for {resolved_url}: {e}")
                            
                    if not html_content:
                        logger.warning(f"Could not fetch HTML content for {resolved_url}.")
                        from scraper.search_utils import match_publication_category
                        pub_cat = match_publication_category(agency, resolved_url)
                        is_pw = detect_paywall(resolved_url, "", "")
                        
                        # Category A zero-failure fallback
                        if pub_cat == "A":
                            logger.info(f"Scraper failed for Category A url {resolved_url}. Applying zero-failure fallback.")
                            is_relevant_fallback = True
                            # Use desc if available; fall back to title so LLM always has some signal
                            fallback_content = desc.strip() if (desc and len(desc.strip()) > 10) else title
                            fallback_summary = fallback_content if len(fallback_content) > 20 else f"[Paywalled/blocked] {title}"
                            try:
                                from scraper.llm import check_relevance_with_groq
                                is_relevant_fallback, _, _, _ = check_relevance_with_groq(
                                    title, fallback_content, keywords, client_name, client_context=client_context
                                )
                            except Exception as fallback_err:
                                logger.error(f"Fallback LLM relevance check failed for Category A: {fallback_err}")
                                is_relevant_fallback = True
                            if is_relevant_fallback:
                                return {
                                    "art_data": {
                                        "title": title,
                                        "url": resolved_url,
                                        "agency": agency,
                                        "summary": fallback_summary,
                                        "publication_category": pub_cat,
                                        "is_paywalled": True
                                    },
                                    "is_relevant_kw": True,
                                    "is_semantic_relevant": True
                                }

                        if is_pw and desc:
                            # Paywalled: run LLM relevance on headline + RSS description
                            logger.info(f"Paywall detected for {resolved_url}. Running LLM relevance on title+description.")
                            try:
                                from scraper.llm import check_relevance_with_groq
                                pw_relevant, _, _, _ = check_relevance_with_groq(
                                    title, desc, keywords, client_name, client_context=client_context
                                )
                            except Exception as pw_err:
                                logger.error(f"Paywall LLM relevance check failed: {pw_err}")
                                pw_relevant = False
                            
                            if pw_relevant:
                                logger.info(f"Paywalled article '{title}' deemed RELEVANT by LLM.")
                                return {
                                    "art_data": {
                                        "title": title,
                                        "url": resolved_url,
                                        "agency": agency,
                                        "summary": desc,
                                        "publication_category": pub_cat,
                                        "is_paywalled": True
                                    },
                                    "is_relevant_kw": True,
                                    "is_semantic_relevant": True
                                }
                        
                        return {
                            "art_data": {
                                "title": title,
                                "url": resolved_url,
                                "agency": agency,
                                "summary": desc or "Could not fetch HTML content.",
                                "publication_category": pub_cat,
                                "is_paywalled": is_pw
                            },
                            "is_relevant_kw": True,
                            "is_semantic_relevant": False
                        }
                        
                    _increment_funnel_metric(job_id, "scraped_count")
                        
                    # 7. Extract body text
                    from scraper.parser import extract_body
                    body_text = extract_body(html_content)
                        
                    if not body_text or len(body_text) < 100:
                        logger.warning(f"No meaningful text extracted for {resolved_url}.")
                        from scraper.search_utils import match_publication_category
                        pub_cat = match_publication_category(agency, resolved_url)
                        is_pw = detect_paywall(resolved_url, html_content, body_text or "")
                        
                        # Category A zero-failure fallback
                        if pub_cat == "A":
                            logger.info(f"Extraction failed for Category A url {resolved_url}. Applying zero-failure fallback.")
                            is_relevant_fallback = True
                            # Combine partial body + desc + title so LLM always has signal
                            fallback_content = " ".join(filter(None, [body_text or "", desc or ""])).strip() or title
                            fallback_summary = fallback_content if len(fallback_content) > 20 else f"[Paywalled/blocked] {title}"
                            try:
                                from scraper.llm import check_relevance_with_groq
                                is_relevant_fallback, _, _, _ = check_relevance_with_groq(
                                    title, fallback_content, keywords, client_name, client_context=client_context
                                )
                            except Exception as fallback_err:
                                logger.error(f"Fallback LLM relevance check failed for Category A: {fallback_err}")
                                is_relevant_fallback = True
                            if is_relevant_fallback:
                                return {
                                    "art_data": {
                                        "title": title,
                                        "url": resolved_url,
                                        "agency": agency,
                                        "summary": fallback_summary,
                                        "publication_category": pub_cat,
                                        "is_paywalled": True
                                    },
                                    "is_relevant_kw": True,
                                    "is_semantic_relevant": True
                                }

                        if is_pw:
                            # Paywalled: run LLM relevance on headline + available text + RSS description
                            fallback_text = " ".join(filter(None, [body_text or "", desc or ""])).strip() or title
                            if fallback_text:
                                logger.info(f"Paywall detected for {resolved_url}. Running LLM relevance on partial content.")
                                try:
                                    from scraper.llm import check_relevance_with_groq
                                    pw_relevant, _, _, _ = check_relevance_with_groq(
                                        title, fallback_text, keywords, client_name, client_context=client_context
                                    )
                                except Exception as pw_err:
                                    logger.error(f"Paywall LLM relevance check failed: {pw_err}")
                                    pw_relevant = False
                                
                                if pw_relevant:
                                    logger.info(f"Paywalled article '{title}' deemed RELEVANT by LLM.")
                                    return {
                                        "art_data": {
                                            "title": title,
                                            "url": resolved_url,
                                            "agency": agency,
                                            "summary": fallback_text,
                                            "publication_category": pub_cat,
                                            "is_paywalled": True
                                        },
                                        "is_relevant_kw": True,
                                        "is_semantic_relevant": True
                                    }
                        
                        return {
                            "art_data": {
                                "title": title,
                                "url": resolved_url,
                                "agency": agency,
                                "summary": desc or "No meaningful text extracted from article.",
                                "publication_category": pub_cat,
                                "is_paywalled": is_pw
                            },
                            "is_relevant_kw": True,
                            "is_semantic_relevant": False
                        }
                        
                    # 8. Cosine Similarity Pre-Filter
                    from scraper.similarity import evaluate_similarity_pre_filter, SIM_DROP_THRESHOLD
                    sim_score = evaluate_similarity_pre_filter(
                        title, body_text, keywords, client_context or ""
                    )
                    # Cosine similarity pre-filter drop is temporarily disabled for maximum recall
                    logger.info(f"Cosine similarity pre-filter drop bypassed for '{title}' (score: {sim_score:.4f} < {SIM_DROP_THRESHOLD}).")
                    # Original similarity filter (commented out to allow switching back):
                    # if sim_score < SIM_DROP_THRESHOLD:
                    #     logger.info(f"Cosine similarity pre-filter drop for '{title}' (score: {sim_score:.4f} < {SIM_DROP_THRESHOLD}). Skipping LLM.")
                    #     with get_db_sync() as db:
                    #         db.merge(IrrelevantArticle(
                    #             url=normalized_url, 
                    #             title=title, 
                    #             description=desc or body_text[:200],
                    #             rejection_reason=f"Similarity pre-filter drop (score: {sim_score:.4f} < {SIM_DROP_THRESHOLD})",
                    #             relevance_score=sim_score,
                    #             last_seen_at=datetime.now()
                    #         ))
                    #         db.commit()
                    #     _increment_funnel_metric(job_id, "pre_filter_dropped")
                    #     from scraper.search_utils import match_publication_category
                    #     return {
                    #         "art_data": {
                    #             "title": title,
                    #             "url": resolved_url,
                    #             "agency": agency,
                    #             "summary": desc or (body_text[:200] + "...") if body_text else "Similarity pre-filter drop.",
                    #             "publication_category": match_publication_category(agency, resolved_url),
                    #             "is_paywalled": False
                    #         },
                    #         "is_relevant_kw": True,
                    #         "is_semantic_relevant": False
                    #     }

                    # 9. Relevance Check (Asymmetric & Ensembling)
                    is_semantic_relevant = False
                    verdict = "uncertain"
                    reason = ""
                    score = 0.5
                    from scraper.llm import check_relevance_with_groq
                    try:
                        is_semantic_relevant, verdict, reason, score = check_relevance_with_groq(
                            title, body_text, keywords, client_name, client_context=client_context
                        )
                    except Exception as rel_err:
                        logger.error(f"Relevance verification error for '{title}': {rel_err}")
                        is_semantic_relevant = True # Fallback to True if API fails
                        verdict = "uncertain"
                        reason = f"Exception: {rel_err}"
                        score = 0.5
                        
                    if not is_semantic_relevant:
                        logger.info(f"Relevance check: article '{title}' is not relevant (verdict: {verdict}, score: {score}). Skipping.")
                        with get_db_sync() as db:
                            db.merge(IrrelevantArticle(
                                url=normalized_url,
                                title=title,
                                description=desc or body_text[:200],
                                rejection_reason=reason,
                                relevance_score=score,
                                last_seen_at=datetime.now()
                            ))
                            db.commit()
                        _increment_funnel_metric(job_id, "relevance_no")
                        from scraper.search_utils import match_publication_category
                        return {
                            "art_data": {
                                "title": title,
                                "url": resolved_url,
                                "agency": agency,
                                "summary": desc or (body_text[:200] + "...") if body_text else "Semantically irrelevant.",
                                "publication_category": match_publication_category(agency, resolved_url),
                                "is_paywalled": False
                            },
                            "is_relevant_kw": True,
                            "is_semantic_relevant": False
                        }
                        
                    _increment_funnel_metric(job_id, "relevance_yes")
                        
                    # 9. Summarize & Enrich
                    logger.info(f"Enriching and summarizing article: {title}")
                    summary_text = ""
                    author_name = None
                    extra_meta = {}
                    try:
                        from scraper.parser import extract_author_v2
                        author_metadata = extract_author_v2(html_content)
                        html_top, html_bottom = "", ""
                        try:
                            import re
                            body_start_match = re.search(r"<body.*?>", html_content[:15000], re.I)
                            body_start_idx = body_start_match.end() if body_start_match else 0
                            html_top = html_content[body_start_idx:body_start_idx + 3000]
                            html_bottom = html_content[-3000:]
                        except Exception as e_snip:
                            logger.warning(f"Snippeting failed inside client report process for {resolved_url}: {e_snip}")
                        
                        extra_meta = {
                            "author_metadata": author_metadata,
                            "html_snippets": {
                                "top": html_top,
                                "bottom": html_bottom
                            }
                        }
                        
                        enrichment = perform_full_enrichment_sync(
                            body=body_text,
                            title=title,
                            url=resolved_url,
                            sector=client_name,
                            context_agency=agency,
                            extra_metadata=extra_meta
                        )
                        if enrichment:
                            if enrichment.get("summary"):
                                summary_text = enrichment["summary"]
                            agency = enrichment.get("agency") or agency
                            author_name = enrichment.get("author")
                    except Exception as e:
                        logger.error(f"LLM Enrichment failed for '{title}': {e}")
                        
                    if not summary_text:
                        summary_text = body_text[:300] + "..."
                        
                    _increment_funnel_metric(job_id, "summarized_count")
                    
                    from scraper.search_utils import match_publication_category
                    pub_category = match_publication_category(agency, resolved_url)
                    extra_meta["publication_category"] = pub_category
                    
                    # Parse RSS published date
                    pub_at_dt = None
                    pub_at_str = art.get("published_at")
                    if pub_at_str:
                        try:
                            pub_at_dt = datetime.fromisoformat(pub_at_str.replace("Z", "+00:00")).replace(tzinfo=None)
                        except Exception:
                            pass
                    if not pub_at_dt:
                        pub_at_dt = datetime.now()
                        
                    # 10. Cache to DB
                    try:
                        val_dict = {
                            "title": title,
                            "url": normalized_url,
                            "resolved_url": resolved_url,
                            "full_body": body_text,
                            "summary": summary_text,
                            "author": author_name,
                            "agency": agency,
                            "published_at": pub_at_dt,
                            "sector": f"{client_name} - {section_name}",
                            "region": "india",
                            "user_id": f"client_{client_id}",
                            "scraped_at": datetime.now(),
                            "extra_metadata": extra_meta
                        }
                        with get_db_sync() as db:
                            existing = db.execute(
                                select(Article)
                                .where(Article.url == normalized_url)
                                .where(Article.user_id == f"client_{client_id}")
                            ).scalars().first()
                            if existing:
                                for k, v in val_dict.items():
                                    setattr(existing, k, v)
                            else:
                                db.add(Article(**val_dict))
                            db.commit()
                    except Exception as dberr:
                        logger.error(f"Failed to cache scraped article to database: {dberr}")
                        
                    art_data = {
                        "title": title,
                        "url": resolved_url,
                        "author": author_name,
                        "agency": agency,
                        "summary": summary_text,
                        "publication_category": pub_category,
                        "is_paywalled": False,
                        "published_at": pub_at_dt.isoformat()
                    }
                    
                    return {
                        "art_data": art_data,
                        "is_relevant_kw": True,
                        "is_semantic_relevant": True
                    }
                except Exception as art_err:
                    logger.error(f"Error processing article '{title}': {art_err}", exc_info=True)
                    return None

            total_to_process = len(unique_discovered)
            _update_progress(f"Processing and filtering {total_to_process} discovered articles in '{section_name}'...")
            logger.info(f"Starting concurrent processing of {total_to_process} articles for section '{section_name}'")
            
            art_tuples = [(art, idx) for idx, art in enumerate(unique_discovered)]
            processed_count = 0
            
            # Run up to 8 threads per section (sections now run in parallel, so 8 keeps total threads safe)
            seen_resolved_urls = set()
            with ThreadPoolExecutor(max_workers=8) as executor:
                futures = {executor.submit(_process_single_article, t): t for t in art_tuples}
                for future in as_completed(futures):
                    processed_count += 1
                    _update_progress(f"Processed article {processed_count}/{total_to_process} in '{section_name}'...")
                    try:
                        res = future.result()
                        if res:
                            art_data = res["art_data"]
                            is_relevant_kw = res["is_relevant_kw"]
                            is_semantic_relevant = res["is_semantic_relevant"]
                            
                            # Deduplicate resolved/normalized URL within this section
                            resolved_url = art_data.get("url")
                            norm_resolved = normalize_url(resolved_url) if resolved_url else ""
                            if norm_resolved:
                                if norm_resolved in seen_resolved_urls:
                                    logger.info(f"Skipping duplicate resolved article within section: {art_data['title']}")
                                    continue
                                seen_resolved_urls.add(norm_resolved)
                            
                            if is_relevant_kw:
                                master_section_articles.append(art_data)
                                if is_semantic_relevant:
                                    passed_region = True
                                    art_agency = art_data.get("agency") or "News"
                                    art_url = art_data.get("url") or ""
                                    is_indian = is_indian_article(art_agency, art_url)
                                    
                                    if region_filter == "Indian":
                                        if not is_indian:
                                            passed_exception = False
                                            if intl_exceptions:
                                                allowed_exceptions = [normalize_publication_name(p) for p in intl_exceptions.split(",") if p.strip()]
                                                art_pub_norm = normalize_publication_name(art_agency)
                                                if art_pub_norm in allowed_exceptions:
                                                    passed_exception = True
                                            if not passed_exception:
                                                passed_region = False
                                    elif region_filter == "International":
                                        if is_indian:
                                            passed_region = False
                                            
                                    passed_priority = True
                                    if priority_media_list:
                                        allowed_pubs = [normalize_publication_name(p) for p in priority_media_list.split(",") if p.strip()]
                                        art_pub_norm = normalize_publication_name(art_agency)
                                        if art_pub_norm not in allowed_pubs:
                                            passed_priority = False
                                            
                                    if passed_region and passed_priority:
                                        filtered_section_articles.append(art_data)
                    except Exception as future_err:
                        logger.error(f"Thread task exception: {future_err}")
            
            cat_counts = {"A": 0, "B": 0, "C": 0}
            paywalled_count = 0
            for art in filtered_section_articles:
                cat_counts[art.get("publication_category", "C")] += 1
                if art.get("is_paywalled"):
                    paywalled_count += 1

            _update_progress(
                f"Section '{section_name}' completed. "
                f"Discovered: {total_to_process}, "
                f"Relevant: {len(filtered_section_articles)} (Cat A: {cat_counts['A']}, Cat B: {cat_counts['B']}, Cat C: {cat_counts['C']}), "
                f"Paywalled: {paywalled_count}."
            )
            return section_name, filtered_section_articles, master_section_articles

        active_processing_sections = {sn: kw for sn, kw in sections_data.items() if kw}
        temp_filtered = {}
        temp_master = {}
        sec_workers = min(len(active_processing_sections), 5)
        with ThreadPoolExecutor(max_workers=sec_workers) as sec_exe:
            sec_futures = {
                sec_exe.submit(_process_section, sn, kw): sn
                for sn, kw in active_processing_sections.items()
            }
            for fut in as_completed(sec_futures):
                try:
                    sn, filtered, master = fut.result()
                    temp_filtered[sn] = filtered
                    temp_master[sn] = master
                except Exception as sec_err:
                    logger.error(f"Section processing failed: {sec_err}", exc_info=True)

        # Re-align report dictionaries to the original database configuration order
        report_data_filtered = {sn: temp_filtered[sn] for sn in sections_data if sn in temp_filtered}
        report_data_master = {sn: temp_master[sn] for sn in sections_data if sn in temp_master}

        # Check if we got any articles at all
        total_filtered_count = sum(len(articles) for articles in report_data_filtered.values())
        has_articles = total_filtered_count > 0
        if not has_articles:
            logger.info("No relevant articles found for any section. A briefing report indicating this will still be generated.")
            
        # 6. Generate DOCX and Excel files
        _update_progress("Compiling Word briefing documents and Excel briefings...")
        date_str = datetime.now().strftime("%d-%m-%Y")
        reports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports")
        os.makedirs(reports_dir, exist_ok=True)
        
        timestamp_suffix = f"{date_str}_{int(datetime.now().timestamp())}"
        
        docx_filename_filtered = f"{client_name.replace(' ', '_')}_Filtered_{timestamp_suffix}.docx"
        docx_path_filtered = os.path.join(reports_dir, docx_filename_filtered)
        
        docx_filename_master = f"{client_name.replace(' ', '_')}_Master_{timestamp_suffix}.docx"
        docx_path_master = os.path.join(reports_dir, docx_filename_master)
        
        excel_filename = f"{client_name.replace(' ', '_')}_{timestamp_suffix}.xlsx"
        excel_path = os.path.join(reports_dir, excel_filename)
        
        logger.info(f"Generating Filtered Word report: {docx_path_filtered}")
        generate_docx_report(
            client_name=client_name,
            date_str=datetime.now().strftime("%B %d, %Y"),
            data=report_data_filtered,
            output_path=docx_path_filtered,
            template_path=template_path
        )
        
        logger.info(f"Generating Master Word report: {docx_path_master}")
        generate_docx_report(
            client_name=f"{client_name} (Master)",
            date_str=datetime.now().strftime("%B %d, %Y"),
            data=report_data_master,
            output_path=docx_path_master,
            template_path=template_path
        )

        logger.info(f"Generating Client Excel report: {excel_path}")
        excel_grouped_data = {
            "Daily News Briefing": {sec: arts for sec, arts in report_data_filtered.items() if arts}
        }
        from scraper.report_generator import generate_excel_report
        generate_excel_report(
            client_name=client_name,
            report_type="Client Intelligence Report",
            date_str=datetime.now().strftime("%Y-%m-%d"),
            grouped_data=excel_grouped_data,
            output_path=excel_path
        )
        
        # 7. Upload to Google Drive/Docs & Share
        google_doc_url_filtered = None
        google_doc_url_master = None
        
        if os.path.exists(docx_path_filtered):
            _update_progress("Uploading Filtered report to Google Drive...")
            logger.info("Uploading Filtered report to Google Drive...")
            try:
                google_doc_url_filtered = upload_docx_to_google_doc(
                    docx_path=docx_path_filtered,
                    client_name=client_name,
                    date_str=datetime.now().strftime("%B %d, %Y"),
                    recipients=all_emails,
                    doc_suffix="",
                    template_path=template_path
                )
            except Exception as e:
                logger.error(f"Failed to upload Filtered report to Google Docs: {e}")
                
        if os.path.exists(docx_path_master):
            _update_progress("Uploading Master report to Google Drive...")
            logger.info("Uploading Master report to Google Drive...")
            try:
                google_doc_url_master = upload_docx_to_google_doc(
                    docx_path=docx_path_master,
                    client_name=client_name,
                    date_str=datetime.now().strftime("%B %d, %Y"),
                    recipients=all_emails,
                    doc_suffix=" (Master)",
                    template_path=template_path
                )
            except Exception as e:
                logger.error(f"Failed to upload Master report to Google Docs: {e}")
                
        # 7.5 Generate and Upload Cumulative Excel to Google Sheets
        cumulative_sheet_url = None
        if has_articles:
            _update_progress("Uploading Cumulative Spreadsheet tab to Google Drive...")
            logger.info("Uploading Cumulative Spreadsheet tab to Google Drive...")
            try:
                from utils.google_docs import upload_cumulative_excel_to_google_sheet
                date_str_tab = datetime.now().strftime("%Y-%m-%d")
                cumulative_sheet_url = upload_cumulative_excel_to_google_sheet(
                    grouped_data=report_data_filtered,
                    client_name=client_name,
                    date_str=date_str_tab,
                    recipients=all_emails
                )
            except Exception as e:
                logger.error(f"Failed to upload Cumulative report to Google Sheets: {e}")

        # 8. Send Email notification
        _update_progress("Sending daily briefing email...")
        logger.info(f"Sending report email to: {all_emails}")
        email_sent = send_report_email(
            recipient_emails=all_emails,
            client_name=client_name,
            docx_path_filtered=docx_path_filtered,
            docx_path_master=None,  # Do not attach Master DOCX
            excel_path_master=excel_path,  # Attach Excel Briefing report
            google_doc_url_filtered=google_doc_url_filtered,
            google_doc_url_master=google_doc_url_master,  # Keep Master Google Doc link
            has_articles=has_articles,
            cumulative_sheet_url=cumulative_sheet_url
        )
        
        # Read the generated report file into binary data to save in the database
        report_file_data = None
        if docx_path_filtered and os.path.exists(docx_path_filtered):
            try:
                with open(docx_path_filtered, "rb") as f:
                    report_file_data = f.read()
            except Exception as e:
                logger.error(f"Failed to read generated report file: {e}")

        excel_file_data = None
        if excel_path and os.path.exists(excel_path):
            try:
                with open(excel_path, "rb") as f:
                    excel_file_data = f.read()
            except Exception as e:
                logger.error(f"Failed to read generated excel file: {e}")

        if not email_sent:
            logger.warning("Report generated but email notification failed to send (SMTP connection error).")
            with get_db_sync() as db:
                run_log = db.execute(select(ClientRunLog).where(ClientRunLog.id == run_log_id)).scalar_one_or_none()
                current_log = run_log.progress_message or ""
                timestamp = datetime.now().strftime("%H:%M:%S")
                final_msg = f"[{timestamp}] Completed with warning: Email failed to send."
                updated_log = f"{current_log}\n{final_msg}" if current_log else final_msg
                db.execute(
                    update(ClientRunLog)
                    .where(ClientRunLog.id == run_log_id)
                    .values(
                        status="completed",
                        generated_file_path=google_doc_url_filtered or docx_path_filtered,
                        generated_file_data=report_file_data,
                        generated_excel_path=excel_path,
                        generated_excel_data=excel_file_data,
                        cumulative_sheet_url=cumulative_sheet_url,
                        progress_message=updated_log,
                        error_message="Email notification failed to send (SMTP connection timeout).",
                        completed_at=datetime.utcnow()
                    )
                )
                
                client_update_vals = {"last_run_at": datetime.utcnow()}
                if cumulative_sheet_url:
                    client_update_vals["cumulative_sheet_url"] = cumulative_sheet_url
                db.execute(
                    update(Client)
                    .where(Client.id == client_id)
                    .values(**client_update_vals)
                )
                db.commit()
            
            # Clean up local temporary files from server disk
            for temp_file in [docx_path_filtered, docx_path_master, excel_path]:
                if temp_file and os.path.exists(temp_file):
                    try:
                        os.remove(temp_file)
                        logger.info(f"Cleaned up local temporary report file from disk: {temp_file}")
                    except Exception as clean_err:
                        logger.warning(f"Failed to remove temporary file {temp_file}: {clean_err}")
            return True
            
        # Update run log status to completed
        with get_db_sync() as db:
            run_log = db.execute(select(ClientRunLog).where(ClientRunLog.id == run_log_id)).scalar_one_or_none()
            current_log = run_log.progress_message or ""
            timestamp = datetime.now().strftime("%H:%M:%S")
            final_msg = f"[{timestamp}] Completed successfully."
            updated_log = f"{current_log}\n{final_msg}" if current_log else final_msg
            db.execute(
                update(ClientRunLog)
                .where(ClientRunLog.id == run_log_id)
                .values(
                    status="completed",
                    generated_file_path=google_doc_url_filtered or docx_path_filtered,
                    generated_file_data=report_file_data,
                    generated_excel_path=excel_path,
                    generated_excel_data=excel_file_data,
                    cumulative_sheet_url=cumulative_sheet_url,
                    progress_message=updated_log,
                    completed_at=datetime.utcnow()
                )
            )
            
            client_update_vals = {"last_run_at": datetime.utcnow()}
            if cumulative_sheet_url:
                client_update_vals["cumulative_sheet_url"] = cumulative_sheet_url
            db.execute(
                update(Client)
                .where(Client.id == client_id)
                .values(**client_update_vals)
            )
            db.commit()
            
        # Clean up local temporary files from server disk
        for temp_file in [docx_path_filtered, docx_path_master, excel_path]:
            if temp_file and os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                    logger.info(f"Cleaned up local temporary report file from disk: {temp_file}")
                except Exception as clean_err:
                    logger.warning(f"Failed to remove temporary file {temp_file}: {clean_err}")

        logger.info(f"Client Report Task completed successfully for client {client_id}")
        return True
        
    except Exception as e:
        error_msg = str(e)
        full_tb = traceback.format_exc()
        logger.error(f"Client report task failed for client {client_id}: {error_msg}\n{full_tb}", exc_info=True)
        
        # Trigger fail-safe email alert immediately
        try:
            send_error_alert_email(
                client_name=client_name,
                error_details=f"Exception: {error_msg}\n\nTraceback:\n{full_tb}"
            )
        except Exception as alert_err:
            logger.error(f"Failed to send fail-safe email alert for client {client_id}: {alert_err}")
            
        # Update run log status to failed
        if run_log_id:
            try:
                with get_db_sync() as db:
                    run_log = db.execute(select(ClientRunLog).where(ClientRunLog.id == run_log_id)).scalar_one_or_none()
                    current_log = run_log.progress_message or ""
                    timestamp = datetime.now().strftime("%H:%M:%S")
                    final_msg = f"[{timestamp}] Failed: {error_msg}"
                    updated_log = f"{current_log}\n{final_msg}" if current_log else final_msg
                    db.execute(
                        update(ClientRunLog)
                        .where(ClientRunLog.id == run_log_id)
                        .values(
                            status="failed",
                            error_message=error_msg,
                            progress_message=updated_log,
                            completed_at=datetime.utcnow()
                        )
                    )
                    db.commit()
            except Exception as dberr:
                logger.error(f"Failed to write failure log to database: {dberr}")
                
        return False


@celery_app.task(name="scraper.tasks.check_client_schedules")
def check_client_schedules():
    """
    Celery Beat task to check client reporting schedules and trigger runs.
    Runs every 5 minutes.
    """
    from db.database import get_db_sync, Client, ClientRunLog
    from datetime import datetime
    import pytz
    from sqlalchemy import select, desc
    
    # Redis Lock to prevent duplicate runs from concurrent schedulers
    try:
        r = get_redis_sync()
        current_minute = datetime.now().minute
        minute_block = current_minute - (current_minute % 5)
        lock_key = f"lock:scheduler:check:{datetime.now().strftime('%Y-%m-%d-%H')}-{minute_block}"
        if not r.set(lock_key, "1", nx=True, ex=240):
            logger.info("Schedule check already processed by another instance. Skipping.")
            return
    except Exception as lock_err:
        logger.warning(f"Failed to acquire Redis scheduler lock: {lock_err}")
    
    logger.info("Checking client automation schedules...")
    
    with get_db_sync() as db:
        active_clients = db.execute(select(Client).where(Client.is_active == True)).scalars().all()
        
        for client in active_clients:
            try:
                # 1. Parse client timezone
                client_tz = pytz.timezone(client.timezone)
                from datetime import datetime, timedelta
                now_tz = datetime.now(client_tz)
                
                # Skip scheduled runs on weekends (Saturday=5, Sunday=6) in client's timezone
                if now_tz.weekday() in [5, 6]:
                    logger.info(f"Skipping schedule check for client '{client.name}' as it is the weekend ({now_tz.strftime('%A')}).")
                    continue
                
                # 2. Parse scheduled time (format HH:MM)
                sched_hour, sched_min = map(int, client.scheduled_time.split(":"))
                
                # 3. Calculate time differences to check if we are in the 10-minute window
                # We check target time for both today and yesterday to support midnight boundary crossings
                target_today = now_tz.replace(hour=sched_hour, minute=sched_min, second=0, microsecond=0)
                diff_today = (now_tz - target_today).total_seconds() / 60.0
                
                target_yesterday = target_today - timedelta(days=1)
                diff_yesterday = (now_tz - target_yesterday).total_seconds() / 60.0
                
                # Determine if current time falls within 10 minutes after scheduled target time
                is_in_window = (0 <= diff_today < 10) or (0 <= diff_yesterday < 10)
                
                if is_in_window:
                    # Decide which target date this trigger belongs to
                    target_date = target_today.date() if (0 <= diff_today < 10) else target_yesterday.date()
                    
                    # Check if a run has already occurred for this specific target date
                    last_log = db.execute(
                        select(ClientRunLog)
                        .where(ClientRunLog.client_id == client.id)
                        .order_by(desc(ClientRunLog.started_at))
                        .limit(1)
                    ).scalar_one_or_none()
                    
                    already_ran = False
                    if last_log:
                        last_started = last_log.started_at.replace(tzinfo=pytz.utc).astimezone(client_tz) if last_log.started_at.tzinfo else pytz.utc.localize(last_log.started_at).astimezone(client_tz)
                        if last_started.date() == target_date and last_log.status in ["completed", "running"]:
                            already_ran = True
                            
                    if not already_ran:
                        logger.info(f"Scheduling report run for client '{client.name}' (Target Date: {target_date}, Scheduled: {client.scheduled_time} in {client.timezone})")
                        celery_app.send_task(
                            "scraper.tasks.run_client_report_task",
                            args=[client.id]
                        )
            except Exception as e:
                logger.error(f"Error checking schedule for client '{client.name}': {e}")


def deduplicate_across_publications(articles_list):
    """
    Deduplicates articles based on normalized title similarity across all publications.
    Keeps only one article per unique story (first one found).
    """
    seen_titles = set()
    deduped = []
    for a in articles_list:
        title = a.get("title", "")
        if not title:
            continue
        # Normalize: convert to lowercase, alphanumeric only
        norm = "".join(ch for ch in title.lower() if ch.isalnum())
        # Check first 50 chars to filter out minor trailing publication suffix variations
        key = norm[:50]
        if key not in seen_titles:
            deduped.append(a)
            seen_titles.add(key)
    return deduped


def build_mailer_grouped_data(articles_list):
    """
    Groups matched articles into three main categories for the mailer:
    1. Google (Critical/Crisis, Corporate/Organizational, AI Specific/Research, Product, Pixel/Watches/Devices)
    2. Competition (Apple, Samsung, OpenAI/ChatGPT, Microsoft, Amazon, Meta (Facebook/Instagram/WhatsApp), Perplexity, Anthropic, MapMyIndia/Mappls, Paytm, Spokesperson Related, Crisis Related)
    3. Industry (10 specified topics)
    
    Allows articles to be placed in multiple sections if they match multiple criteria.
    """
    import re
    
    # Initialize the target categories in the exact order requested
    grouped = {
        "Google": {
            "Critical/Crisis": [],
            "Corporate/Organizational": [],
            "AI Specific/Research": [],
            "Product": [],
            "Pixel/Watches/Devices": []
        },
        "Competition": {
            "Apple": [],
            "Samsung": [],
            "OpenAI/ChatGPT": [],
            "Microsoft": [],
            "Amazon": [],
            "Meta (Facebook/Instagram/WhatsApp)": [],
            "Perplexity": [],
            "Anthropic": [],
            "MapMyIndia/Mappls": [],
            "Paytm": [],
            "Spokesperson Related": [],
            "Crisis Related": []
        },
        "Industry": {
            "Data center": [],
            "AI Sovereignty": [],
            "AI and Jobs": [],
            "Content moderation": [],
            "Social media ban": [],
            "Copyright": [],
            "Broad AI hot topics": [],
            "Digital Divide": [],
            "AI Bubble": [],
            "Trade deal": []
        }
    }
    
    # Compile regexes for faster matching
    re_google = re.compile(
        r"\b(google|alphabet|youtube|yt\b|android|pixel|nest|fitbit|gemini|waymo|sundar\b|pichai|preeti\b|lobana|dutta\b|chobey|gupta\b|swamy|sreedharan|mohan\b|digikavach)\b", 
        re.IGNORECASE
    )
    
    re_crisis = re.compile(
        r"\b(crisis|fine|fines|penalty|penalties|investigation|investigates|probe|probes|cci|antitrust|monopoly|court|lawsuit|sues|suit|ban|bans|regulatory|regulation|legal|layoffs|layoff|strike|hacked|breach|vuln|tax|dispute)\b",
        re.IGNORECASE
    )
    
    re_corp = re.compile(
        r"\b(appoint|leader|leadership|partnership|partner|investment|invest|funding|office|expansion|hiring|hire|event|csr|arts|culture|philanthropy|grant|startup|ecosystem|skills|skilling)\b",
        re.IGNORECASE
    )
    
    re_ai = re.compile(
        r"\b(ai\b|artificial intelligence|gemini|deepmind|bard|vertex|llm|generative|genai)\b",
        re.IGNORECASE
    )
    
    re_prod = re.compile(
        r"\b(search|maps|pay|gpay|wallet|ads|workspace|meet|docs|sheets|drive|chrome|classroom)\b",
        re.IGNORECASE
    )
    
    re_dev = re.compile(
        r"\b(pixel|watch|wear os|fitbit|earbuds|buds|chromecast|nest|tensor|hardware|device|devices|smartphone|smartwatch)\b",
        re.IGNORECASE
    )
    
    # Competition brands & topics
    re_apple = re.compile(r"\b(apple|iphone|ipad|macbook|airpods|ios|bionic|iwatch)\b", re.IGNORECASE)
    re_samsung = re.compile(r"\b(samsung|galaxy|exynos|one ui)\b", re.IGNORECASE)
    re_openai = re.compile(r"\b(openai|chatgpt)\b", re.IGNORECASE)
    re_microsoft = re.compile(r"\b(microsoft|msft\b|azure)\b", re.IGNORECASE)
    re_amazon = re.compile(r"\b(amazon|aws\b)\b", re.IGNORECASE)
    re_meta = re.compile(r"\b(meta\b|facebook|instagram|whatsapp|insta\b)\b", re.IGNORECASE)
    re_perplexity = re.compile(r"\bperplexity\b", re.IGNORECASE)
    re_anthropic = re.compile(r"\b(anthropic|claude)\b", re.IGNORECASE)
    re_mmi = re.compile(r"\b(mapmyindia|mappls)\b", re.IGNORECASE)
    re_paytm = re.compile(r"\bpaytm\b", re.IGNORECASE)
    re_spokesperson = re.compile(r"\b(spokesperson|spokesman|spokeswoman|representative|statement|declared|commented|replied)\b", re.IGNORECASE)
    
    # Industry critical topics
    re_datacenter = re.compile(r"\b(data\s+centers?|data\s+centres?|server\s+farms?|power\s+grids?|cooling\s+systems?)\b", re.IGNORECASE)
    re_sovereignty = re.compile(r"\b(sovereign\s+ai|ai\s+sovereignty|domestic\s+ai|localized\s+ai|local\s+ai|national\s+ai|meity|digital\s+public\s+infrastructure|dpi|dpdp)\b", re.IGNORECASE)
    re_jobs = re.compile(r"\b(jobs?|hiring|layoffs?|unemployment|workforce|skills?\s+gap|future\s+of\s+work|employment)\b", re.IGNORECASE)
    re_moderation = re.compile(r"\b(content\s+moderation|moderation|hate\s+speech|take\s+down|misinformation|fake\s+news|deepfakes?|fact\s+checks?|online\s+safety)\b", re.IGNORECASE)
    re_socban = re.compile(r"\b(social\s+media\s+bans?|ban\s+social\s+media|kids\s+bans?|youth\s+restrictions?|banning\s+tiktok)\b", re.IGNORECASE)
    re_copyright = re.compile(r"\b(copyrights?|fair\s+use|publisher\s+lawsuits?|licensing\s+deals?|intellectual\s+property|ip\s+infringement)\b", re.IGNORECASE)
    re_broadai = re.compile(r"\b(ai\b|artificial\s+intelligence|machine\s+learning|llms?|generative\s+ai|gen\s+ai|chatgpt|openai|claude|deepseek|perplexity|llama)\b", re.IGNORECASE)
    re_digitaldivide = re.compile(r"\b(digital\s+divide|rural\s+internet|skilling|digital\s+literacy|accessibility|affordable\s+devices|connectivity)\b", re.IGNORECASE)
    re_bubble = re.compile(r"\b(ai\s+bubble|bubble|tech\s+crash|overvalued|overhype|capital\s+spending|monetize\s+ai|profitability)\b", re.IGNORECASE)
    re_trade = re.compile(r"\b(trade\s+deals?|tariffs?|import\s+dut(y|ies)|cross-border|export\s+controls?|trade\s+agreements?)\b", re.IGNORECASE)

    for art in articles_list:
        title = art.get("title") or ""
        body = art.get("summary") or art.get("full_body") or ""
        text_to_check = f"{title} {body}"
        
        # Check matching keywords in _keyword_hits to help accuracy
        kw_hits = [str(k).lower() for k in art.get("_keyword_hits", [])]
        
        # Competition brands & topics
        is_apple = bool(re_apple.search(title)) or any("apple" in k or "iphone" in k or "ios" in k or "airpods" in k for k in kw_hits)
        is_samsung = bool(re_samsung.search(title)) or any("samsung" in k or "galaxy" in k for k in kw_hits)
        is_openai_chatgpt = bool(re_openai.search(title)) or any("openai" in k or "chatgpt" in k for k in kw_hits)
        is_msft = bool(re_microsoft.search(title)) or any("microsoft" in k or "azure" in k for k in kw_hits)
        is_amzn = bool(re_amazon.search(title)) or any("amazon" in k or "aws" in k for k in kw_hits)
        is_meta_fb = bool(re_meta.search(title)) or any("meta" in k or "facebook" in k or "instagram" in k or "whatsapp" in k for k in kw_hits)
        is_perp = bool(re_perplexity.search(title)) or any("perplexity" in k for k in kw_hits)
        is_anth = bool(re_anthropic.search(title)) or any("anthropic" in k or "claude" in k for k in kw_hits)
        is_mmi_map = bool(re_mmi.search(title)) or any("mapmyindia" in k or "mappls" in k for k in kw_hits)
        is_paytm_app = bool(re_paytm.search(title)) or any("paytm" in k for k in kw_hits)
        is_spokes = bool(re_spokesperson.search(title))
        is_comp_crisis = bool(re_crisis.search(title))

        is_competitor = (
            is_apple or is_samsung or is_openai_chatgpt or is_msft or is_amzn or
            is_meta_fb or is_perp or is_anth or is_mmi_map or is_paytm_app
        )

        is_google_branded = bool(re_google.search(text_to_check)) or any("google" in k or "youtube" in k or "android" in k or "gemini" in k or "pixel" in k or "sundar" in k for k in kw_hits)
        
        # --- 1. GOOGLE CATEGORY ---
        # Exclude competitor-related news to keep Google section direct.
        if is_google_branded and not is_competitor:
            added_to_google = False
            
            # Critical/Crisis (spokesperson & crisis related on top)
            is_crisis = bool(re_crisis.search(title)) or any("policy" in k or "regulation" in k or "legal" in k or "cci" in k for k in kw_hits)
            if is_crisis:
                grouped["Google"]["Critical/Crisis"].append(art)
                added_to_google = True
            
            # Corporate/Organizational
            is_corp = bool(re_corp.search(title)) or any("skilling" in k or "startup" in k or "arts" in k or "culture" in k for k in kw_hits)
            if is_corp:
                grouped["Google"]["Corporate/Organizational"].append(art)
                added_to_google = True
                
            # AI Specific/Research
            is_ai = bool(re_ai.search(title)) or any("gemini" in k or "deepmind" in k or "ai" in k for k in kw_hits)
            if is_ai:
                grouped["Google"]["AI Specific/Research"].append(art)
                added_to_google = True
                
            # Product
            is_prod = bool(re_prod.search(title)) or any("search" in k or "maps" in k or "pay" in k or "workspace" in k for k in kw_hits)
            if is_prod:
                grouped["Google"]["Product"].append(art)
                added_to_google = True
                
            # Pixel/Watches/Devices
            is_dev = bool(re_dev.search(title)) or any("pixel" in k or "watch" in k or "buds" in k or "wear os" in k or "tensor" in k for k in kw_hits)
            if is_dev:
                grouped["Google"]["Pixel/Watches/Devices"].append(art)
                added_to_google = True
                
            # Fallback if Google but no subcategory matched
            if not added_to_google:
                grouped["Google"]["Corporate/Organizational"].append(art)
                
        # --- 2. COMPETITION CATEGORY ---
        # We classify if any of these match and it contains competitor elements
        if is_apple:
            grouped["Competition"]["Apple"].append(art)
        if is_samsung:
            grouped["Competition"]["Samsung"].append(art)
        if is_openai_chatgpt:
            grouped["Competition"]["OpenAI/ChatGPT"].append(art)
        if is_msft:
            grouped["Competition"]["Microsoft"].append(art)
        if is_amzn:
            grouped["Competition"]["Amazon"].append(art)
        if is_meta_fb:
            grouped["Competition"]["Meta (Facebook/Instagram/WhatsApp)"].append(art)
        if is_perp:
            grouped["Competition"]["Perplexity"].append(art)
        if is_anth:
            grouped["Competition"]["Anthropic"].append(art)
        if is_mmi_map:
            grouped["Competition"]["MapMyIndia/Mappls"].append(art)
        if is_paytm_app:
            grouped["Competition"]["Paytm"].append(art)
        if is_spokes and is_competitor:
            grouped["Competition"]["Spokesperson Related"].append(art)
        if is_comp_crisis and is_competitor:
            grouped["Competition"]["Crisis Related"].append(art)
            
        # --- 3. INDUSTRY CATEGORY ---
        if re_datacenter.search(title):
            grouped["Industry"]["Data center"].append(art)
        if re_sovereignty.search(title) or any("policy" in k or "regulation" in k or "dpdp" in k for k in kw_hits):
            grouped["Industry"]["AI Sovereignty"].append(art)
        if re_jobs.search(title) and ("ai" in title.lower() or "ai" in body.lower()):
            grouped["Industry"]["AI and Jobs"].append(art)
        if re_moderation.search(title):
            grouped["Industry"]["Content moderation"].append(art)
        if re_socban.search(title):
            grouped["Industry"]["Social media ban"].append(art)
        if re_copyright.search(title) or any("copyright" in k for k in kw_hits):
            grouped["Industry"]["Copyright"].append(art)
        if re_broadai.search(title) or any("ai" in k or "chatgpt" in k or "openai" in k or "claude" in k for k in kw_hits):
            grouped["Industry"]["Broad AI hot topics"].append(art)
        if re_digitaldivide.search(title):
            grouped["Industry"]["Digital Divide"].append(art)
        if re_bubble.search(title):
            grouped["Industry"]["AI Bubble"].append(art)
        if re_trade.search(title):
            grouped["Industry"]["Trade deal"].append(art)

    # Clean up empty subcategories to prevent printing blank headers
    cleaned_grouped = {}
    for master, subs in grouped.items():
        cleaned_subs = {sub: arts for sub, arts in subs.items() if arts}
        if cleaned_subs:
            cleaned_grouped[master] = cleaned_subs
            
    return cleaned_grouped


# ──────────────────────────────────────────────────────────────────────────────
# Heavy Automation Tasks
# ──────────────────────────────────────────────────────────────────────────────

@celery_app.task(name="scraper.tasks.run_heavy_automation_task", time_limit=3600, soft_time_limit=3300)
def run_heavy_automation_task(company_id: int):
    """
    Main background task: fetch articles, dedup, filter, generate Master + Filtered reports.
    Phase 2 scope: fetch → dedup → Master Report.
    Phase 3 adds: filtering → Filtered Report.
    """
    from db.database import get_db_sync, HeavyCompany, HeavyRun, HeavyRunArticle, HeavyRecipient
    from scraper.heavy_filter import exact_dedup, near_dedup, bucket_articles
    from scraper.report_generator import generate_docx_report
    from utils.email import send_error_alert_email
    from datetime import date, datetime, timedelta
    import pytz

    company_name = f"Company ID {company_id}"
    run_id = None

    logger.info(f"[Heavy] Starting automation for company_id {company_id}")

    # Create run record
    with get_db_sync() as db:
        company = db.execute(select(HeavyCompany).where(HeavyCompany.id == company_id)).scalar_one_or_none()
        if not company:
            logger.error(f"[Heavy] Company {company_id} not found")
            return False

        run = HeavyRun(company_id=company_id, status="running", started_at=datetime.utcnow())
        db.add(run)
        db.commit()
        db.refresh(run)
        run_id = run.id
        company_name = company.name
        sector_match = company.sector_match
        window_hours = company.window_hours
        search_mode = company.search_mode if company.search_mode else "title"
        pooja_algo_enabled = getattr(company, "pooja_algo_enabled", False)
        pooja_folder_filtering_enabled = getattr(company, "pooja_folder_filtering_enabled", False)
        pooja_priority_conf = getattr(company, "pooja_priority_conf", 5)
        pooja_non_priority_conf = getattr(company, "pooja_non_priority_conf", 7)
        llm_judge_enabled = getattr(company, "llm_judge_enabled", False)


    try:
        def save_intermediate_csv(filename, articles_list, extra_headers=[]):
            import csv
            reports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports")
            os.makedirs(reports_dir, exist_ok=True)
            file_path = os.path.join(reports_dir, filename)
            headers = ["Title", "URL", "Agency", "Published_At"] + extra_headers
            try:
                with open(file_path, "w", encoding="utf-8", newline="") as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    for a in articles_list:
                        row = [
                            a.get("title", ""),
                            a.get("url", ""),
                            a.get("agency", ""),
                            str(a.get("published_at", ""))
                        ]
                        for h in extra_headers:
                            if h == "Matched_Keyword":
                                kw_hits = a.get("_keyword_hits", [])
                                row.append(", ".join(kw_hits) if isinstance(kw_hits, list) else str(kw_hits))
                            elif h == "Sector":
                                row.append(a.get("_pillar", ""))
                            elif h == "Sub_Category":
                                row.append(a.get("_sub_category", ""))
                            else:
                                row.append("")
                        writer.writerow(row)
                return filename
            except Exception as ex:
                logger.error(f"Failed to save intermediate CSV {filename}: {ex}")
                return None

        def _update_progress(msg: str):
            try:
                with get_db_sync() as db:
                    run_rec = db.execute(select(HeavyRun).where(HeavyRun.id == run_id)).scalar_one_or_none()
                    if run_rec:
                        run_rec.progress_message = (run_rec.progress_message or "") + msg + "\n"
                        db.commit()
            except Exception as e:
                logger.warning(f"[Heavy] Progress update failed: {e}")

        # ─ Fetch from Nexus Remote Feed ───────────────────────────────────────
        _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Fetching articles from Nexus remote feed...")

        import requests as _requests
        import time as _time

        nexus_base = os.getenv("NEXUS_FEED_URL", "http://34.142.240.96")
        nexus_key  = os.getenv("NEXUS_SERVICE_KEY", "nexus_sk_fb74eaae34cd3e53f6ac2031479337cb")
        FETCH_TIMEOUT = int(os.getenv("NEXUS_FETCH_TIMEOUT", "60"))   # seconds per request
        FETCH_RETRIES = int(os.getenv("NEXUS_FETCH_RETRIES", "3"))     # attempts before giving up
        FETCH_RETRY_DELAY = 2                                           # seconds between retries

        def normalize_sector(sector_str: str) -> str:
            """Normalise a raw sector string to its canonical lowercase form."""
            if not sector_str:
                return "other"
            s = sector_str.lower().strip()
            # Strip trailing numbers / spaces so 'google 2', 'google1' → 'google'
            import re as _re
            base = _re.sub(r'[\s_-]*\d+$', '', s).strip()
            return base if base else s

        sectors = [s.strip() for s in sector_match.split(",") if s.strip()]
        cutoff_dt = datetime.utcnow() - timedelta(hours=window_hours)
        cutoff_date = cutoff_dt.date().isoformat()

        # Pre-defined mapping for phonetic/spelling variations that simple wildcard matches might miss
        SECTOR_VARIANTS = {
            'tech': ['tech', 'Tech', 'TECH', 'Techhh', 'tech1', 'teccH', 'Tech1', 'TeccH'],
            'stock market': ['stock market', 'Stock Market'],
            'policies': ['policies', 'Policies'],
            'real estate': ['real estate', 'Real Estate'],
            'healthcare': ['healthcare', 'HEALTHCARE', 'HealthCare', 'Health'],
            'startups': ['startups', 'StartUp'],
            'foods and drinks': ['foods and drinks', 'FOODS AND DRINKS', 'Foods'],
            'ai': ['ai', 'AI', 'Ai'],
            'google': ['google', 'google 2', 'google 3', 'Google3'],
            'travel': ['travel', 'Travell'],
            'lifestyle': ['lifestyle', 'LifeStyle'],
            'consultancies': ['consultancies', 'Consultancies'],
            'fintech': ['fintech', 'Fintech', 'FinTech', 'FINTECH'],
            'automobile': ['automobile', 'Automobile', 'auto', 'Auto', 'AUTOMOBILE'],
            'media and entertainment': ['media and entertainment', 'Media and Entertainment', 'media', 'Media', 'entertainment', 'Entertainment'],
            'education': ['education', 'Education', 'EDUCATION']
        }

        # ── Dynamically discover ALL sector name variants stored in the local DB ──
        from db.database import Article as _Article
        from sqlalchemy import select as _select, text as _text
        all_remote_variants: list[str] = []
        with get_db_sync() as _db:
            for sec in sectors:
                sec_norm = normalize_sector(sec)
                # 1. Start with any static explicit variants
                explicit = SECTOR_VARIANTS.get(sec_norm, [])
                
                # 2. Add any wildcard prefix matches from the database (e.g. 'google 1', 'google 2')
                rows = _db.execute(
                    _text("SELECT DISTINCT sector FROM articles WHERE published_at >= :cutoff AND LOWER(sector) LIKE :pat"),
                    {"cutoff": cutoff_dt, "pat": f"{sec_norm}%"}
                ).fetchall()
                db_matches = [r[0] for r in rows if r[0]]
                
                # Combine both lists and deduplicate
                found = list(set(explicit + db_matches))
                if not found:
                    found = [sec]
                all_remote_variants.extend(found)
                _update_progress(f"  Sector '{sec}' → Resolved variants: {found}")

        # De-duplicate while preserving order
        seen_variants: set[str] = set()
        unique_variants: list[str] = []
        for v in all_remote_variants:
            if v.lower() not in seen_variants:
                seen_variants.add(v.lower())
                unique_variants.append(v)

        fetched: list[dict] = []
        seen_ids: set = set()

        def _fetch_with_retry(url: str, params: dict) -> dict | None:
            """GET with retry-backoff. Returns parsed JSON or None on total failure."""
            for attempt in range(1, FETCH_RETRIES + 1):
                try:
                    resp = _requests.get(url, params=params, timeout=FETCH_TIMEOUT)
                    resp.raise_for_status()
                    return resp.json()
                except Exception as ex:
                    if attempt < FETCH_RETRIES:
                        _update_progress(f"  ⚠ Attempt {attempt}/{FETCH_RETRIES} failed ({ex}). Retrying in {FETCH_RETRY_DELAY}s...")
                        _time.sleep(FETCH_RETRY_DELAY)
                    else:
                        _update_progress(f"  ✗ Sector fetch failed after {FETCH_RETRIES} attempts: {ex}")
                        return None

        for v in unique_variants:
            _update_progress(f"  → Fetching sector: {v}")
            page = 1
            while True:
                data = _fetch_with_retry(
                    f"{nexus_base}/api/feed",
                    {
                        "api_key":   nexus_key,
                        "sector":    v,
                        "date_from": cutoff_date,
                        "page":      page,
                        "page_size": 500,
                    },
                )
                if data is None:
                    break   # all retries exhausted — move to next variant

                for a in data.get("articles", []):
                    uid = a.get("id") or a.get("url")
                    if uid and uid not in seen_ids:
                        seen_ids.add(uid)
                        fetched.append({
                            "id":           a.get("id"),
                            "title":        a.get("title"),
                            "url":          a.get("resolved_url") or a.get("url"),
                            "published_at": a.get("published_at"),
                            "full_body":    a.get("full_body"),
                            "summary":      a.get("summary"),
                            "agency":       a.get("agency"),
                            "author":       a.get("author"),
                            "source_feed":  normalize_sector(a.get("sector") or v),
                        })

                total_pages = data.get("total_pages", 1)
                if page >= total_pages:
                    break
                page += 1

        _update_progress(f"Fetched {len(fetched)} articles from Nexus remote feed")


        if not fetched:
            with get_db_sync() as db:
                run_rec = db.execute(select(HeavyRun).where(HeavyRun.id == run_id)).scalar_one_or_none()
                if run_rec:
                    run_rec.status = "completed"
                    run_rec.fetched_count = 0
                    run_rec.deduped_count = 0
                    run_rec.relevant_count = 0
                    run_rec.finished_at = datetime.utcnow()
                    db.commit()
            logger.info(f"[Heavy] No articles fetched for {company_name}")
            return True

        # Step 1: Save fetched articles
        fetched_fn = f"Fetched_Articles_Run_{run_id}.csv"
        save_intermediate_csv(fetched_fn, fetched)
        _update_progress(f"Step 1: Fetched {len(fetched)} articles from Nexus feed. Downloadable output: /api/heavy-automation/reports/{fetched_fn}")

        # ─ Exact dedup ────────────────────────────────────────────────────────
        _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Exact dedup...")
        deduped_exact = exact_dedup(fetched)
        _update_progress(f"After exact dedup: {len(deduped_exact)} articles")

        # ─ Near-dup clustering ────────────────────────────────────────────────
        _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Near-dup clustering...")
        deduped = near_dedup(deduped_exact, threshold=0.80)
        
        # Step 2: Save deduped articles
        deduped_fn = f"Deduped_Articles_Run_{run_id}.csv"
        save_intermediate_csv(deduped_fn, deduped)
        _update_progress(f"Step 2: After deduplication: {len(deduped)} articles. Downloadable output: /api/heavy-automation/reports/{deduped_fn}")

        # ─ Filter articles using super-final CSV keywords + priority media list ────
        _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Filtering articles using Google_keywords_super_final.csv...")

        # Resolve Google_material path robustly for both local and container environments
        _tasks_dir = os.path.dirname(os.path.abspath(__file__))
        _backend_dir = os.path.dirname(_tasks_dir)
        _parent_of_backend = os.path.dirname(_backend_dir)
        
        # Try local dev path first
        super_final_csv_path = os.path.join(_parent_of_backend, "Google_material", "Google_keywords_super_final.csv")
        media_list_path = os.path.join(_parent_of_backend, "Google_material", "[Internal] Google Online Priority Media List 2026.xlsx")
        
        # Fallback to production container path if local path is missing
        if not os.path.exists(super_final_csv_path):
            super_final_csv_path = os.path.join(_backend_dir, "Google_material", "Google_keywords_super_final.csv")
        if not os.path.exists(media_list_path):
            media_list_path = os.path.join(_backend_dir, "Google_material", "[Internal] Google Online Priority Media List 2026.xlsx")

        _update_progress("Loading super-final keywords CSV and priority publication list...")
        kw_buckets = parse_super_final_csv(super_final_csv_path)
        corp_keywords    = kw_buckets["corporate"]
        product_keywords = kw_buckets["product"]
        priority_publications = load_priority_media_list(media_list_path)
        _update_progress(
            f"Loaded {len(corp_keywords)} corporate keyword entries and "
            f"{len(product_keywords)} product keyword entries from CSV."
        )

        corporate_articles = []
        product_articles   = []
        relevant_map       = {}

        if pooja_folder_filtering_enabled:
            _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Filtering articles using Pooja's folder filtering logic...")
            folder_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Pooja_filtering_Logic_for_heavy_automation_final")
            media_list_path = os.path.join(folder_dir, "[Internal] Google Online Priority Media List 2026.xlsx")
            keywords_xlsx_path = os.path.join(folder_dir, "keywords.xlsx")
            
            import sys
            if folder_dir not in sys.path:
                sys.path.append(folder_dir)
            
            from Pooja_filtering_Logic_for_heavy_automation_final.filter_priority_media import load_priority_publications, build_match_keywords, is_priority as pooja_is_priority
            from Pooja_filtering_Logic_for_heavy_automation_final.filter_by_keywords import parse_keywords_file, build_keyword_index, match_title as pooja_match_title
            
            publications = load_priority_publications(media_list_path)
            match_pairs = build_match_keywords(publications)
            
            sectors_data = parse_keywords_file(keywords_xlsx_path)
            keyword_index = build_keyword_index(sectors_data)
            
            for art in deduped:
                agency = art.get("agency") or ""
                title = art.get("title") or ""
                
                is_pri, matched_pub = pooja_is_priority(agency, match_pairs)
                art["_is_priority"] = is_pri
                
                if is_pri:
                    kw, sector, sub_cat = pooja_match_title(title, keyword_index)
                    if kw:
                        art_p = dict(art)
                        art_p["confidence_score"] = 10
                        art_p["matches"] = [{"master": sector or "Other", "sub": sub_cat or "General", "matched_items": [kw]}]
                        art_p["_pillar"] = sector or "Other"
                        art_p["_sub_category"] = sub_cat or "General"
                        art_p["_keyword_hits"] = [kw]
                        art_p["_relevance_score"] = 10
                        
                        relevant_map[art["url"]] = art_p
                        
                        is_product = "product" in (sector or "").lower()
                        if is_product:
                            art_p["_source_type"] = "product"
                            product_articles.append(art_p)
                        else:
                            art_p["_source_type"] = "corporate"
                            corporate_articles.append(art_p)
            relevant = list(relevant_map.values())
            
            # Step 3: Save Pooja-filtered articles (before Claude check)
            pooja_filtered_fn = f"Pooja_Filtered_Articles_Run_{run_id}.csv"
            save_intermediate_csv(pooja_filtered_fn, relevant, ["Matched_Keyword", "Sector", "Sub_Category"])
            _update_progress(f"Step 3: Pooja filtered matches (before Claude check): {len(relevant)} articles. Downloadable output: /api/heavy-automation/reports/{pooja_filtered_fn}")
            
            # Send to Claude for keyword relevance check
            claude_log_fn = f"Claude_Verification_Log_Run_{run_id}.txt"
            
            if llm_judge_enabled:
                _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Sending {len(relevant)} articles to Claude to verify keyword relevance...")
                from scraper.heavy_llm import verify_article_keyword_with_claude
                
                claude_verified = []
                claude_log_lines = []
                
                for art in relevant:
                    title = art.get("title") or ""
                    kw_hits = art.get("_keyword_hits", [])
                    keyword = kw_hits[0] if (isinstance(kw_hits, list) and kw_hits) else str(kw_hits)
                    
                    is_valid = verify_article_keyword_with_claude(title, keyword)
                    if is_valid:
                        claude_verified.append(art)
                        claude_log_lines.append(f"[KEEP] Matched keyword '{keyword}' in title: {title}")
                    else:
                        claude_log_lines.append(f"[DISCARD] Matched keyword '{keyword}' is irrelevant in title: {title}")
                
                # Save Claude verification trace logs separately
                reports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports")
                with open(os.path.join(reports_dir, claude_log_fn), "w", encoding="utf-8") as f:
                    f.write("\n".join(claude_log_lines))
                
                _update_progress(f"Step 3.5: Claude keyword verification complete. Discarded {len(relevant) - len(claude_verified)} irrelevant matches. Downloadable output: /api/heavy-automation/reports/{claude_log_fn}")
            else:
                _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] LLM Judge disabled. Skipping Claude keyword verification (keeping all matches).")
                claude_verified = relevant
                claude_log_lines = ["Claude keyword verification bypassed (LLM Judge toggle is OFF)."]
                reports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports")
                with open(os.path.join(reports_dir, claude_log_fn), "w", encoding="utf-8") as f:
                    f.write("\n".join(claude_log_lines))
                
                _update_progress(f"Step 3.5: Claude keyword verification bypassed. Downloadable output: /api/heavy-automation/reports/{claude_log_fn}")

            
            # Rebuild relevant_map, corporate_articles, product_articles based on Claude verification
            relevant_map = {a["url"]: a for a in claude_verified}
            relevant = list(relevant_map.values())
            
            corporate_articles = [a for a in corporate_articles if a["url"] in relevant_map]
            product_articles = [a for a in product_articles if a["url"] in relevant_map]
            
            # Step 4: Save Claude-verified articles
            claude_verified_fn = f"Claude_Verified_Articles_Run_{run_id}.csv"
            save_intermediate_csv(claude_verified_fn, relevant, ["Matched_Keyword", "Sector", "Sub_Category"])
            _update_progress(f"Step 4: Claude verified relevant articles: {len(relevant)} articles. Downloadable output: /api/heavy-automation/reports/{claude_verified_fn}")
        else:
            # Counters for priority media list breakdown
            priority_total       = 0
            non_priority_total   = 0
            priority_corp_pass   = 0
            priority_prod_pass   = 0
            non_priority_corp_pass = 0
            non_priority_prod_pass = 0

            _update_progress(f"Running relevancy match in mode: {search_mode}")

            for art in deduped:
                title  = art.get("title", "") or ""
                agency = art.get("agency", "") or ""

                # Check based on search_mode
                if search_mode == "full_body":
                    body = art.get("summary") or art.get("full_body") or ""
                    query_text = f"{title}\n{body}"
                else:
                    query_text = title

                is_priority = is_priority_publication(agency, priority_publications)
                art["_is_priority"] = is_priority

                if is_priority:
                    priority_total += 1
                else:
                    non_priority_total += 1

                corp_conf,    corp_matches    = evaluate_headline_relevance(query_text, corp_keywords)
                product_conf, product_matches = evaluate_headline_relevance(query_text, product_keywords)

                # When Pooja Algo is enabled: priority pubs use a lower bar (pooja_priority_conf),
                # non-priority pubs use a stricter bar (pooja_non_priority_conf).
                # When disabled: all articles use pooja_non_priority_conf uniformly.
                if pooja_algo_enabled:
                    corp_min    = pooja_priority_conf    if is_priority else pooja_non_priority_conf
                    product_min = pooja_priority_conf    if is_priority else pooja_non_priority_conf
                else:
                    corp_min    = pooja_non_priority_conf
                    product_min = pooja_non_priority_conf

                # Corporate bucket
                if corp_conf >= corp_min:
                    art_corp = dict(art)
                    art_corp["confidence_score"] = corp_conf
                    art_corp["matches"]          = corp_matches
                    art_corp["_source_type"]     = "corporate"
                    corporate_articles.append(art_corp)
                    relevant_map[art["url"]] = art
                    if is_priority:
                        priority_corp_pass += 1
                    else:
                        non_priority_corp_pass += 1

                # Product bucket
                if product_conf >= product_min:
                    art_prod = dict(art)
                    art_prod["confidence_score"] = product_conf
                    art_prod["matches"]          = product_matches
                    art_prod["_source_type"]     = "product"
                    product_articles.append(art_prod)
                    relevant_map[art["url"]] = art
                    if is_priority:
                        priority_prod_pass += 1
                    else:
                        non_priority_prod_pass += 1

            relevant = list(relevant_map.values())
            _update_progress(
                f"Priority media list breakdown: "
                f"{priority_total} articles from priority pubs "
                f"({priority_corp_pass} corp + {priority_prod_pass} product passed), "
                f"{non_priority_total} from non-priority pubs "
                f"({non_priority_corp_pass} corp + {non_priority_prod_pass} product passed)"
            )
            _update_progress(
                f"Relevance matching: corporate matches={len(corporate_articles)}, "
                f"product matches={len(product_articles)}, unique relevant={len(relevant)}"
            )

        # Build audit trail parameters on unique relevant articles
        for art_url, art in relevant_map.items():
            corp_matches_found = []
            corp_c = 0
            for a in corporate_articles:
                if a["url"] == art_url:
                    corp_matches_found = a["matches"]
                    corp_c = a["confidence_score"]
                    break

            prod_matches_found = []
            prod_c = 0
            for a in product_articles:
                if a["url"] == art_url:
                    prod_matches_found = a["matches"]
                    prod_c = a["confidence_score"]
                    break

            all_matches  = corp_matches_found + prod_matches_found
            pillars      = list(set(m["master"] for m in all_matches))
            subs         = list(set(m["sub"]    for m in all_matches))
            keyword_hits = list(set(
                item
                for m in all_matches
                for item in m.get("matched_items", [])
            ))

            art["_relevance_score"] = max(corp_c, prod_c)
            art["_pillar"]          = ", ".join(pillars) if pillars else "Other"
            art["_sub_category"]    = ", ".join(subs)    if subs    else "General"
            art["_keyword_hits"]    = keyword_hits

            if corp_c > 0 and prod_c > 0:
                art["_bucket"] = "both"
            elif corp_c > 0:
                art["_bucket"] = "corporate"
            else:
                art["_bucket"] = "product"

        # ─ Summaries: generate fresh 30-40 word summaries using Groq ─────────────
        _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Generating 30-40 word summaries using Groq...")
        from scraper.heavy_llm import summarize_article
        for idx, art in enumerate(relevant, start=1):
            title = art.get("title") or ""
            body = art.get("full_body") or art.get("summary") or ""
            summary_text = summarize_article(title, body)
            if not summary_text:
                summary_text = art.get("summary") or body or ""
            art["_summary"] = summary_text
            art["summary"] = summary_text

        # Map summaries back to the per-bucket article lists
        for art in corporate_articles:
            orig = relevant_map.get(art["url"])
            if orig:
                art["summary"] = orig.get("_summary") or orig.get("summary") or orig.get("full_body") or ""

        for art in product_articles:
            orig = relevant_map.get(art["url"])
            if orig:
                art["summary"] = orig.get("_summary") or orig.get("summary") or orig.get("full_body") or ""

        # ─ Group articles by master/sub heading for the report ────────────────────
        def group_articles_for_report(articles_list):
            grouped = {}
            for a in articles_list:
                for match in a.get("matches", []):
                    master = match["master"]
                    sub    = match["sub"]
                    grouped.setdefault(master, {}).setdefault(sub, [])
                    if a["url"] not in [x["url"] for x in grouped[master][sub]]:
                        grouped[master][sub].append(a)
            return grouped

        corp_grouped    = group_articles_for_report(corporate_articles)
        product_grouped = group_articles_for_report(product_articles)

        # Merge both buckets into a single combined dict (corporate sections first,
        # then product sections) to produce ONE unified report file.
        all_grouped: dict = {}
        for src in (corp_grouped, product_grouped):
            for master, subs in src.items():
                if master not in all_grouped:
                    all_grouped[master] = {}
                for sub, arts in subs.items():
                    if sub not in all_grouped[master]:
                        all_grouped[master][sub] = []
                    # Deduplicate by URL when merging
                    existing_urls = {x["url"] for x in all_grouped[master][sub]}
                    for a in arts:
                        if a["url"] not in existing_urls:
                            all_grouped[master][sub].append(a)
                            existing_urls.add(a["url"])

        # Ensure database schema is up-to-date with google_doc_url and mailer_doc_path
        try:
            from sqlalchemy import text
            with get_db_sync() as db:
                db.execute(text("ALTER TABLE heavy_runs ADD COLUMN google_doc_url TEXT"))
                db.commit()
        except Exception:
            pass
        try:
            from sqlalchemy import text
            with get_db_sync() as db:
                db.execute(text("ALTER TABLE heavy_runs ADD COLUMN mailer_doc_path TEXT"))
                db.commit()
        except Exception:
            pass

        # ─ Generate single combined DOCX report ───────────────────────────────────
        _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Generating combined Google Report DOCX...")
        today_str = date.today().strftime("%Y-%m-%d")
        google_doc_filename = f"Google_{company_name}_{today_str}_{run_id}.docx"
        master_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports", google_doc_filename
        )
        from scraper.report_generator import generate_organized_docx_report
        generate_organized_docx_report(company_name, "Google Intelligence Report", today_str, all_grouped, master_path)
        _update_progress(f"Combined Google Report DOCX saved: {google_doc_filename}")

        # ─ Generate single combined Excel report ──────────────────────────────────
        _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Generating combined Google Report Excel...")
        google_excel_filename = f"Google_{company_name}_{today_str}_{run_id}.xlsx"
        master_excel_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports", google_excel_filename
        )
        from scraper.report_generator import generate_excel_report
        generate_excel_report(company_name, "Google Intelligence Report", today_str, all_grouped, master_excel_path)
        _update_progress(f"Combined Google Report Excel saved: {google_excel_filename}")

        # Aliases so the rest of the task (email, run record) stays consistent
        filtered_path       = None              # no separate filtered doc anymore
        filtered_excel_path = None              # no separate filtered excel anymore

        # Group relevant by sections for email
        by_pillar_email = group_articles_by_sections(relevant, company_name)

        # ─ Store per-article audit trail (Phase 5) ────────────────────────────
        _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Storing audit trail for {len(relevant)} articles...")

        try:
            with get_db_sync() as db:
                for art in relevant:
                    pub_date = art.get("published_at")
                    if isinstance(pub_date, str):
                        try:
                            pub_date = datetime.fromisoformat(pub_date.replace("Z", "+00:00"))
                        except Exception:
                            try:
                                pub_date = datetime.strptime(pub_date.split(".")[0], "%Y-%m-%d %H:%M:%S")
                            except Exception:
                                pub_date = datetime.utcnow()
                    elif not pub_date:
                        pub_date = datetime.utcnow()

                    audit = HeavyRunArticle(
                        run_id=run_id,
                        source_article_id=art.get("id"),
                        title=art.get("title"),
                        url=art.get("url"),
                        published_at=pub_date,
                        dedup_cluster_id=art.get("cluster_id"),
                        relevance_score=art.get("_relevance_score"),
                        included_in_brief=True,
                        pillar=art.get("_pillar"),
                        sub_category=art.get("_sub_category"),
                        matched_keywords=json.dumps(art.get("_keyword_hits", [])),
                        bucket=art.get("_bucket", "clear_keep"),
                    )
                    db.add(audit)
                db.commit()
            _update_progress(f"Audit trail stored.")
        except Exception as e:
            logger.warning(f"[Heavy] Audit trail storage failed: {e}")

        # Generate Executive Summary + Strategic Takeaways (Phase 4)
        if llm_judge_enabled:
            _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Generating executive summary...")

            from scraper.heavy_llm import generate_executive_summary, generate_strategic_takeaways
            
            # Deduplicate articles across publications for mailer content
            mailer_relevant = deduplicate_across_publications(relevant)
            
            exec_summary = generate_executive_summary(mailer_relevant, company_name=company_name)
            takeaways = generate_strategic_takeaways(mailer_relevant, company_name=company_name)

            if exec_summary:
                _update_progress(f"Executive Summary: {exec_summary[:100]}...")
            if takeaways:
                _update_progress(f"Strategic Takeaways generated.")
        else:
            _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] LLM Judge disabled. Skipping executive summary generation.")
            exec_summary = "Executive Summary skipped (LLM Judge / Claude disabled for this run)."
            takeaways = "Strategic Takeaways skipped (LLM Judge / Claude disabled for this run)."
            mailer_relevant = relevant


        # Generate Mailer DOCX (Phase 4 Extension)
        _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Generating Mailer DOCX...")
        google_mailer_filename = f"Mailer_Doc_{company_name}_{today_str}_{run_id}.docx"
        mailer_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports", google_mailer_filename
        )
        
        # Build new 3-category grouped structure for the mailer
        mailer_grouped = build_mailer_grouped_data(mailer_relevant)
        
        from scraper.report_generator import generate_mailer_docx_report
        generate_mailer_docx_report(company_name, "Google Daily Brief", today_str, exec_summary, takeaways, mailer_grouped, mailer_path)
        _update_progress(f"Mailer DOCX saved: {google_mailer_filename}")


        # Collect email recipients by role to use for sharing the Google Doc
        with get_db_sync() as db:
            brief_recips = db.execute(
                select(HeavyRecipient).where(
                    HeavyRecipient.company_id == company_id,
                    HeavyRecipient.role == "brief"
                )
            ).scalars().all()
            master_recips = db.execute(
                select(HeavyRecipient).where(
                    HeavyRecipient.company_id == company_id,
                    HeavyRecipient.role == "master_doc"
                )
            ).scalars().all()
        brief_emails = [r.email for r in brief_recips]
        master_emails = [r.email for r in master_recips]
        all_recipient_emails = list(set(brief_emails + master_emails))

        # Upload Mailer Doc to Google Docs (Google Drive)
        google_doc_url = None
        if os.environ.get("GOOGLE_CREDENTIALS_JSON") and all_recipient_emails:
            _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Uploading Mailer Doc to Google Drive as Google Doc...")
            try:
                from utils.google_docs import upload_docx_to_google_doc
                google_doc_url = upload_docx_to_google_doc(
                    docx_path=mailer_path,
                    client_name=company_name,
                    date_str=today_str,
                    recipients=all_recipient_emails,
                    doc_suffix="Mailer"
                )
                _update_progress(f"Google Doc link: {google_doc_url}")
            except Exception as upload_err:
                logger.error(f"[Heavy] Failed to upload Mailer Doc: {upload_err}")
                _update_progress(f"Warning: Google Doc upload failed: {upload_err}")
        else:
            _update_progress("Google Drive credentials not set or no recipients, skipping Google Docs upload.")

        # Read mailer doc bytes
        mailer_data = None
        if mailer_path and os.path.exists(mailer_path):
            try:
                with open(mailer_path, "rb") as f:
                    mailer_data = f.read()
            except Exception as e:
                logger.error(f"[Heavy] Failed to read mailer document for db backup: {e}")

        # Save paths and summaries to database
        with get_db_sync() as db:
            run_rec = db.execute(select(HeavyRun).where(HeavyRun.id == run_id)).scalar_one_or_none()
            if run_rec:
                run_rec.master_doc_path = master_path
                run_rec.master_excel_path = master_excel_path
                run_rec.google_doc_url = google_doc_url
                run_rec.mailer_doc_path = mailer_path
                run_rec.mailer_doc_data = mailer_data
                run_rec.executive_summary = exec_summary
                run_rec.takeaways = takeaways
                db.commit()

        # Update takeaways Google Sheet in a non-blocking try-except block
        if takeaways and takeaways.strip():
            _update_progress("Updating strategic takeaways Google Sheet...")
            try:
                from utils.google_docs import append_daily_takeaways_to_sheet
                sheet_url = append_daily_takeaways_to_sheet(company.name, date.today(), takeaways)
                if sheet_url:
                    _update_progress(f"Takeaways Google Sheet updated: {sheet_url}")
                    # Update company takeaways_sheet_url in DB
                    with get_db_sync() as db:
                        comp_rec = db.execute(select(HeavyCompany).where(HeavyCompany.id == company.id)).scalar_one_or_none()
                        if comp_rec:
                            comp_rec.takeaways_sheet_url = sheet_url
                            db.commit()
                else:
                    _update_progress("Warning: Takeaways Google Sheet was not updated.")
            except Exception as sheet_err:
                logger.error(f"[Heavy] Failed to update takeaways sheet: {sheet_err}", exc_info=True)
                _update_progress(f"Warning: Takeaways sheet update failed: {sheet_err}")

        # Build and send email
        _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Sending intelligence brief email...")

        from utils.email import send_report_email

        # 1. Read send configurations
        send_reports = getattr(company, "email_send_reports", True)
        send_html = getattr(company, "email_send_html", False)

        # 2. Render HTML body if enabled
        html_body = None
        if send_html and relevant:
            try:
                BLUE   = "#4285F4"
                RED    = "#EA4335"
                YELLOW = "#FBBC04"
                GREEN  = "#34A853"

                unique_subs = list(set(art.get("_sub_category") for art in mailer_relevant if art.get("_sub_category")))
                if "General" in unique_subs and len(unique_subs) > 1:
                    unique_subs.remove("General")
                top_tags = [sub for sub in unique_subs[:10]]

                # Parse bullets from executive summary and enforce 15 words limit strictly
                exec_bullets = []
                if exec_summary:
                    for line in exec_summary.split("\n"):
                        line = line.strip()
                        if not line:
                            continue
                        for prefix in ("-", "*", "•"):
                            if line.startswith(prefix):
                                clean_line = line[len(prefix):].strip()
                                words = clean_line.split()
                                if len(words) > 15:
                                    clean_line = " ".join(words[:15]) + "..."
                                exec_bullets.append(clean_line)
                                break
                if not exec_bullets and exec_summary:
                    words = exec_summary.split()
                    if len(words) > 15:
                        exec_bullets.append(" ".join(words[:15]) + "...")
                    else:
                        exec_bullets.append(exec_summary)

                sections = []
                accent_colors = [BLUE, RED, YELLOW, GREEN]
                
                # Limit inline email body articles to 30 to prevent Gmail clipping
                email_articles = mailer_relevant[:30]
                email_grouped = build_mailer_grouped_data(email_articles)
                
                idx = 0
                for cat_name in ("Google", "Competition", "Industry"):
                    if cat_name in email_grouped:
                        # Collect all articles under all subcategories of this master category, deduplicated by URL
                        cat_arts = []
                        seen_urls = set()
                        for sub_arts in email_grouped[cat_name].values():
                            for art in sub_arts:
                                if art["url"] not in seen_urls:
                                    cat_arts.append(art)
                                    seen_urls.add(art["url"])
                        
                        if cat_arts:
                            color = accent_colors[idx % len(accent_colors)]
                            sections.append({
                                "name": cat_name.upper(),
                                "accent": color,
                                "articles": cat_arts
                            })
                            idx += 1

                brief_data = {
                    "brand": company_name,
                    "subtitle": "DAILY INTELLIGENCE BRIEF",
                    "date_str": date.today().strftime("%d %B %Y").upper(),
                    "top_tags": top_tags,
                    "exec_intro": f"Key headline developments shaping the landscape today:",
                    "exec_bullets": exec_bullets,
                    "sections": sections,
                    "signoff_name": "THE MAVERICKS Intelligence Desk",
                    "signoff_sub": f"Daily News Coverage — {date.today().strftime('%d %B %Y')}",
                    "sections_covered": " | ".join(s["name"] for s in sections),
                    "disclaimer": f"This daily intelligence brief has been compiled from publicly available media sources for informational purposes only. The content herein represents the views of the cited third-party publications and does not constitute the official position of {company_name} or Alphabet Inc. All brand names and trademarks remain the property of their respective owners.",
                    "topic_tags": [f"#{t.replace(' ', '')}" for t in unique_subs[:10]],
                }
                from utils.mailer import render_brief_html
                html_body = render_brief_html(brief_data)
                _update_progress("HTML mailer rendered successfully.")
            except Exception as render_err:
                logger.error(f"[Heavy] Failed to render HTML brief: {render_err}", exc_info=True)
                _update_progress(f"Warning: HTML mailer rendering failed: {render_err}")

        email_status = "skipped"
        if brief_emails or master_emails:
            if company.mail_send_mode == "Immediate":
                _update_progress(f"Sending email immediately...")
                try:
                    success_brief = True
                    success_master = True
                    
                    # ── Daily Brief email (now can send HTML or plain text) ──
                    if brief_emails:
                        _update_progress(f"Sending Daily Brief email to {len(brief_emails)} recipients...")
                        success_brief = send_report_email(
                            recipient_emails=brief_emails,
                            client_name=company_name,
                            docx_path_filtered=mailer_path,
                            docx_path_master=None,
                            has_articles=bool(mailer_relevant),
                            brief_content=exec_summary,
                            # google_doc_url_filtered=google_doc_url,  # [COMMENTED OUT] Disabled sending Google Doc links in emails
                            html_body=html_body if send_html else None,
                        )

                    if master_emails:
                        _update_progress(f"Sending Corporate & Keywords Reports email to {len(master_emails)} recipients...")
                        excel_master = master_excel_path if send_reports else None
                        success_master = send_report_email(
                            recipient_emails=master_emails,
                            client_name=company_name,
                            docx_path_filtered=mailer_path,
                            docx_path_master=None,  # Do not send the 3rd doc (Master DOCX report)
                            has_articles=bool(mailer_relevant),
                            brief_content=exec_summary,
                            excel_path_filtered=filtered_excel_path,
                            excel_path_master=excel_master,
                            # google_doc_url_filtered=google_doc_url,  # [COMMENTED OUT] Disabled sending Google Doc links in emails
                            html_body=html_body if send_html else None,
                        )
                        
                    email_status = "sent" if (success_brief and success_master) else "failed"
                except Exception as e:
                    logger.error(f"[Heavy] Email send failed: {e}")
                    email_status = "failed"
                    try:
                        send_error_alert_email(company_name, f"Email send failed: {str(e)}")
                    except:
                        pass
                _update_progress(f"Email status: {email_status}")
            else:
                # Scheduled mode: store send_at for Phase 5 beat task
                _update_progress(f"Email scheduled for {company.mail_send_time} (mode: {company.mail_send_mode})")
                email_status = "pending"

        # Read the generated report files into binary data to save in the database
        master_doc_data = None
        if master_path and os.path.exists(master_path):
            try:
                with open(master_path, "rb") as f:
                    master_doc_data = f.read()
            except Exception as e:
                logger.error(f"Failed to read master doc report: {e}")

        filtered_doc_data = None
        if filtered_path and os.path.exists(filtered_path):
            try:
                with open(filtered_path, "rb") as f:
                    filtered_doc_data = f.read()
            except Exception as e:
                logger.error(f"Failed to read filtered doc report: {e}")

        master_excel_data = None
        if master_excel_path and os.path.exists(master_excel_path):
            try:
                with open(master_excel_path, "rb") as f:
                    master_excel_data = f.read()
            except Exception as e:
                logger.error(f"Failed to read master excel report: {e}")

        filtered_excel_data = None
        if filtered_excel_path and os.path.exists(filtered_excel_path):
            try:
                with open(filtered_excel_path, "rb") as f:
                    filtered_excel_data = f.read()
            except Exception as e:
                logger.error(f"Failed to read filtered excel report: {e}")

        # ─ Update run record ──────────────────────────────────────────────────
        with get_db_sync() as db:
            run_rec = db.execute(select(HeavyRun).where(HeavyRun.id == run_id)).scalar_one_or_none()
            if run_rec:
                run_rec.status = "completed"
                run_rec.fetched_count = len(fetched)
                run_rec.deduped_count = len(deduped)
                run_rec.relevant_count = len(relevant)
                run_rec.master_doc_path = master_path
                run_rec.filtered_doc_path = filtered_path
                run_rec.master_excel_path = master_excel_path
                run_rec.filtered_excel_path = filtered_excel_path
                run_rec.master_doc_data = master_doc_data
                run_rec.filtered_doc_data = filtered_doc_data
                run_rec.master_excel_data = master_excel_data
                run_rec.filtered_excel_data = filtered_excel_data
                run_rec.email_status = email_status
                run_rec.executive_summary = exec_summary
                run_rec.takeaways = takeaways
                run_rec.finished_at = datetime.utcnow()
                db.commit()

        logger.info(f"[Heavy] Task completed for {company_name}: fetched={len(fetched)}, deduped={len(deduped)}")
        return True

    except Exception as e:
        logger.error(f"[Heavy] Task failed for {company_name}: {e}", exc_info=True)
        try:
            with get_db_sync() as db:
                run_rec = db.execute(select(HeavyRun).where(HeavyRun.id == run_id)).scalar_one_or_none()
                if run_rec:
                    run_rec.status = "failed"
                    run_rec.error = str(e)
                    run_rec.finished_at = datetime.utcnow()
                    db.commit()
            send_error_alert_email(company_name, str(e))
        except Exception as alert_err:
            logger.error(f"[Heavy] Failed to update error state: {alert_err}")
        return False


@celery_app.task(name="scraper.tasks.check_heavy_automation_schedules")
def check_heavy_automation_schedules():
    """
    Beat task: checks all enabled HeavyCompany configs every 5 min.
    If current time matches frequency/days/fetch_time, dispatches run_heavy_automation_task.
    """
    from db.database import get_db_sync, HeavyCompany, HeavyRun
    from sqlalchemy import desc
    import pytz
    from datetime import timedelta

    logger.info("[Heavy] Checking automation schedules...")

    try:
        with get_db_sync() as db:
            companies = db.execute(select(HeavyCompany).where(HeavyCompany.enabled == True)).scalars().all()

            for company in companies:
                try:
                    tz = pytz.timezone(company.timezone or "Asia/Kolkata")
                    now_tz = datetime.now(tz)

                    # ─── SECTION A: Check and trigger daily/weekly/monthly scraper runs ───
                    sched_hour, sched_min = map(int, (company.fetch_time or "07:00").split(":"))
                    target = now_tz.replace(hour=sched_hour, minute=sched_min, second=0, microsecond=0)

                    # Check if we're within 10 minutes after daily/weekly schedule
                    diff_min = (now_tz - target).total_seconds() / 60.0
                    if 0 <= diff_min < 10:
                        # Check frequency/days
                        weekday = now_tz.strftime("%A")[:3].upper()
                        month_day = now_tz.day
                        days_list = [] if not company.days else json.loads(company.days)

                        should_run = False
                        if company.frequency == "Daily":
                            should_run = True
                        elif company.frequency == "Weekly" and weekday in days_list:
                            should_run = True
                        elif company.frequency == "Monthly" and str(month_day) in days_list:
                            should_run = True
                        elif company.frequency == "Custom" and weekday in days_list:
                            should_run = True

                        if should_run:
                            # Check if already ran today
                            last_run = db.execute(
                                select(HeavyRun)
                                .where(HeavyRun.company_id == company.id)
                                .order_by(desc(HeavyRun.started_at))
                                .limit(1)
                            ).scalar_one_or_none()

                            run_today = False
                            if last_run:
                                run_start = last_run.started_at or last_run.finished_at
                                if run_start:
                                    # database time is UTC, convert it to company's local timezone to compare dates correctly
                                    run_start_local = run_start.replace(tzinfo=pytz.utc).astimezone(tz)
                                    last_run_date = run_start_local.date()
                                else:
                                    last_run_date = None

                                if last_run_date == now_tz.date():
                                    run_today = True

                            if not run_today:
                                logger.info(f"[Heavy] Triggering daily/weekly run for {company.name}")
                                celery_app.send_task("scraper.tasks.run_heavy_automation_task", args=[company.id])

                    # ─── SECTION B: Check and send monthly takeaways spreadsheet email ───
                    send_monthly = getattr(company, "send_monthly_takeaways_enabled", False)
                    if send_monthly:
                        m_day = getattr(company, "monthly_takeaways_day", 1) or 1
                        m_time = getattr(company, "monthly_takeaways_time", "09:00") or "09:00"
                        
                        try:
                            m_hour, m_min = map(int, m_time.split(":"))
                            m_target = now_tz.replace(day=m_day, hour=m_hour, minute=m_min, second=0, microsecond=0)
                            
                            m_diff_min = (now_tz - m_target).total_seconds() / 60.0
                            if now_tz.day == m_day and (0 <= m_diff_min < 10):
                                target_month_str = now_tz.strftime("%Y-%m")
                                
                                last_sent = company.last_monthly_takeaways_sent_at
                                already_sent = False
                                if last_sent:
                                    last_sent_tz = last_sent.replace(tzinfo=pytz.utc).astimezone(tz)
                                    if last_sent_tz.strftime("%Y-%m") == target_month_str:
                                        already_sent = True
                                        
                                if not already_sent:
                                    logger.info(f"[Heavy] Triggering monthly takeaways report send for {company.name}")
                                    celery_app.send_task("scraper.tasks.send_monthly_takeaways_report_task", args=[company.id])
                        except Exception as monthly_sched_err:
                            logger.error(f"[Heavy] Monthly takeaways schedule check error for {company.name}: {monthly_sched_err}")

                except Exception as e:
                    logger.error(f"[Heavy] Schedule check error for {company.name}: {e}")

    except Exception as e:
        logger.error(f"[Heavy] Schedule check failed: {e}", exc_info=True)


@celery_app.task(name="scraper.tasks.check_heavy_scheduled_sends")
def check_heavy_scheduled_sends():
    """
    Phase 5: Beat task (every 5 min) — check for pending emails at their scheduled send time.
    If current time >= mail_send_time, send the queued reports.
    """
    from db.database import get_db_sync, HeavyRun, HeavyCompany, HeavyRecipient, HeavyRunArticle
    from utils.email import send_report_email, send_error_alert_email
    import pytz

    logger.info("[Heavy] Checking scheduled sends...")

    try:
        with get_db_sync() as db:
            # Find runs with email_status='pending' and current time >= mail_send_time
            pending_runs = db.execute(
                select(HeavyRun).where(HeavyRun.email_status == "pending")
            ).scalars().all()

            for run in pending_runs:
                try:
                    company = db.execute(
                        select(HeavyCompany).where(HeavyCompany.id == run.company_id)
                    ).scalar_one_or_none()

                    if not company or company.mail_send_mode != "Scheduled":
                        continue

                    # Check if current time >= mail_send_time in company's timezone
                    tz = pytz.timezone(company.timezone or "Asia/Kolkata")
                    now_tz = datetime.now(tz)
                    sched_hour, sched_min = map(int, (company.mail_send_time or "08:00").split(":"))
                    send_target = now_tz.replace(hour=sched_hour, minute=sched_min, second=0, microsecond=0)

                    if now_tz < send_target:
                        continue  # Not yet time

                    # Time to send — fetch recipients and send
                    brief_recips = db.execute(
                        select(HeavyRecipient).where(
                            HeavyRecipient.company_id == company.id,
                            HeavyRecipient.role == "brief"
                        )
                    ).scalars().all()

                    master_recips = db.execute(
                        select(HeavyRecipient).where(
                            HeavyRecipient.company_id == company.id,
                            HeavyRecipient.role == "master_doc"
                        )
                    ).scalars().all()

                    brief_emails = [r.email for r in brief_recips]
                    master_emails = [r.email for r in master_recips]

                    if not (brief_emails or master_emails):
                        logger.info(f"[Heavy] Run {run.id}: no recipients configured")
                        run.email_status = "skipped"
                        db.commit()
                        continue

                    # Extract configurations
                    send_reports = getattr(company, "email_send_reports", True)
                    send_html = getattr(company, "email_send_html", False)

                    # Reconstruct HTML briefing if enabled
                    html_body = None
                    if send_html:
                        try:
                            run_arts = db.execute(
                                select(HeavyRunArticle).where(HeavyRunArticle.run_id == run.id)
                            ).scalars().all()
                            
                            relevant_list = []
                            unique_subs = []
                            
                            for ra in run_arts:
                                art_dict = {
                                    "title": ra.title,
                                    "url": ra.url,
                                    "published_at": ra.published_at,
                                    "summary": ra.llm_summary or "",
                                    "_summary": ra.llm_summary or "",
                                    "publication": "wire",
                                    "source": "wire",
                                    "scope": "DOMESTIC",
                                    "_pillar": ra.pillar or "Other",
                                    "_sub_category": ra.sub_category or "General"
                                }
                                
                                # Resolve original publication/agency metadata from Article table
                                from db.database import Article as _Article
                                if ra.source_article_id:
                                    orig_art = db.execute(
                                        select(_Article).where(_Article.id == ra.source_article_id)
                                    ).scalar_one_or_none()
                                    if orig_art:
                                        art_dict["agency"] = orig_art.agency
                                        art_dict["author"] = orig_art.author
                                        art_dict["publication"] = orig_art.agency
                                        art_dict["source"] = orig_art.agency

                                relevant_list.append(art_dict)
                                if ra.sub_category and ra.sub_category not in unique_subs:
                                    unique_subs.append(ra.sub_category)

                            by_pillar_email_reconstructed = group_articles_by_sections(relevant_list, company.name)

                            BLUE   = "#4285F4"
                            RED    = "#EA4335"
                            YELLOW = "#FBBC04"
                            GREEN  = "#34A853"

                            if "General" in unique_subs and len(unique_subs) > 1:
                                unique_subs.remove("General")
                            top_tags = [sub for sub in unique_subs[:10]]

                            exec_cards = []
                            card_colors = [BLUE, GREEN, RED, YELLOW]
                            for idx, art in enumerate(relevant_list[:6]):
                                exec_cards.append({
                                    "label": (art.get("_sub_category") or "ALERT").upper().split(",")[0].strip(),
                                    "color": card_colors[idx % len(card_colors)],
                                    "text": art.get("title")
                                })

                            sections = []
                            accent_colors = [BLUE, RED, YELLOW, GREEN]
                            for idx, (pillar, arts) in enumerate(by_pillar_email_reconstructed.items()):
                                color = accent_colors[idx % len(accent_colors)]
                                sections.append({
                                    "name": pillar.upper(),
                                    "accent": color,
                                    "articles": arts
                                })

                            takeaway_items = []
                            run_takeaways = run.takeaways or ""
                            if run_takeaways:
                                lines = [l.strip().lstrip("-*•").strip() for l in run_takeaways.split("\n") if l.strip()]
                                for line in lines:
                                    parts = line.split("—", 1)
                                    if len(parts) == 2:
                                        title_part, text_part = parts[0].strip(), parts[1].strip()
                                    else:
                                        parts = line.split(":", 1)
                                        if len(parts) == 2:
                                            title_part, text_part = parts[0].strip(), parts[1].strip()
                                        else:
                                            title_part, text_part = "Key Takeaway", line
                                    takeaway_items.append({
                                        "title": title_part,
                                        "text": text_part
                                    })

                            brief_data = {
                                "brand": company.name,
                                "subtitle": "DAILY INTELLIGENCE BRIEF",
                                "date_str": run.started_at.strftime("%d %B %Y").upper(),
                                "top_tags": top_tags,
                                "exec_intro": f"{run.relevant_count} headline developments shaping {company.name}'s strategic landscape today:",
                                "exec_cards": exec_cards,
                                "sections": sections,
                                "takeaways_intro": "Key intelligence insights from today's coverage:",
                                "takeaways": takeaway_items,
                                "signoff_name": "THE MAVERICKS Intelligence Desk",
                                "signoff_sub": f"Daily News Coverage — {run.started_at.strftime('%d %B %Y')}",
                                "sections_covered": " | ".join(by_pillar_email_reconstructed.keys()),
                                "disclaimer": f"This daily intelligence brief has been compiled from publicly available media sources for informational purposes only. The content herein represents the views of the cited third-party publications and does not constitute the official position of {company.name} or Alphabet Inc. All brand names and trademarks remain the property of their respective owners.",
                                "topic_tags": [f"#{t.replace(' ', '')}" for t in unique_subs[:10]],
                            }
                            from utils.mailer import render_brief_html
                            html_body = render_brief_html(brief_data)
                        except Exception as render_err:
                            logger.error(f"[Heavy] Failed to render scheduled HTML brief for run {run.id}: {render_err}", exc_info=True)

                    # Send the reports
                    try:
                        success_brief = True
                        success_master = True
                        
                        if brief_emails:
                            logger.info(f"[Heavy] Run {run.id}: Sending scheduled Daily Brief email to {len(brief_emails)} recipients...")
                            success_brief = send_report_email(
                                recipient_emails=brief_emails,
                                client_name=company.name,
                                docx_path_filtered=None,
                                docx_path_master=None,
                                has_articles=run.relevant_count > 0,
                                brief_content=run.executive_summary,
                                html_body=html_body if send_html else None,
                            )
                            
                        if master_emails:
                            logger.info(f"[Heavy] Run {run.id}: Sending scheduled Reports email to {len(master_emails)} recipients...")
                            docx_path = run.master_doc_path if send_reports else None
                            excel_path = run.master_excel_path if send_reports else None
                            success_master = send_report_email(
                                recipient_emails=master_emails,
                                client_name=company.name,
                                docx_path_filtered=None,
                                docx_path_master=docx_path,
                                has_articles=run.relevant_count > 0,
                                brief_content=run.executive_summary,
                                excel_path_filtered=None,
                                excel_path_master=excel_path,
                                html_body=html_body if send_html else None,
                            )
                            
                        run.email_status = "sent" if (success_brief and success_master) else "failed"
                        logger.info(f"[Heavy] Run {run.id}: email sent (status: {run.email_status})")
                    except Exception as send_err:
                        logger.error(f"[Heavy] Run {run.id}: email send failed: {send_err}")
                        run.email_status = "failed"
                        try:
                            send_error_alert_email(company.name, f"Scheduled email send failed: {str(send_err)}")
                        except:
                            pass

                    db.commit()

                except Exception as e:
                    logger.error(f"[Heavy] Scheduled send error for run {run.id}: {e}")

    except Exception as e:
        logger.error(f"[Heavy] Scheduled sends check failed: {e}", exc_info=True)


@celery_app.task(name="scraper.tasks.send_monthly_takeaways_report_task")
def send_monthly_takeaways_report_task(company_id: int):
    """
    Downloads the takeaways Excel spreadsheet and emails it to all the company's recipients.
    Updates company.last_monthly_takeaways_sent_at on success.
    """
    from db.database import get_db_sync, HeavyCompany, HeavyRecipient
    from utils.google_docs import download_takeaways_sheet_file
    from utils.email import send_monthly_takeaways_report_email
    from datetime import datetime
    
    logger.info(f"[Heavy] Running send_monthly_takeaways_report_task for company ID {company_id}...")
    
    try:
        with get_db_sync() as db:
            company = db.execute(select(HeavyCompany).where(HeavyCompany.id == company_id)).scalar_one_or_none()
            if not company:
                logger.error(f"[Heavy] Company {company_id} not found.")
                return False
                
            recipients = db.execute(select(HeavyRecipient).where(HeavyRecipient.company_id == company.id)).scalars().all()
            recipient_emails = [r.email for r in recipients if r.email]
            
            if not recipient_emails:
                logger.info(f"[Heavy] No recipients configured for company {company.name}. Skipping monthly takeaways send.")
                return False
                
            sheet_url = company.takeaways_sheet_url
            
            # Download takeaways excel sheet
            excel_path = download_takeaways_sheet_file(company.name)
            if not excel_path or not os.path.exists(excel_path):
                logger.error(f"[Heavy] Takeaways spreadsheet could not be downloaded for {company.name}.")
                return False
                
            # Send monthly takeaways email
            success = send_monthly_takeaways_report_email(
                recipient_emails=recipient_emails,
                client_name=company.name,
                excel_path=excel_path,
                google_sheet_url=sheet_url
            )
            
            # Clean up local file
            try:
                os.remove(excel_path)
            except Exception:
                pass
                
            if success:
                # Update sent timestamp
                company.last_monthly_takeaways_sent_at = datetime.utcnow()
                db.commit()
                logger.info(f"[Heavy] Successfully completed monthly takeaways send for {company.name}")
                return True
            else:
                logger.error(f"[Heavy] Failed to send monthly takeaways email for {company.name}")
                return False
                
    except Exception as e:
        logger.error(f"[Heavy] Error sending monthly takeaways for company {company_id}: {e}", exc_info=True)
        return False


# ==============================================================================
# ─── ROBUST AUTOMATION CELERY TASKS ───────────────────────────────────────────
# ==============================================================================

@celery_app.task(name="scraper.tasks.check_robust_automation_schedules")
def check_robust_automation_schedules():
    """
    Beat task: Checks all enabled RobustCompany configurations every 5 min.
    If timezone, frequency, and time match, triggers run_robust_automation_task.
    """
    from db.database import get_db_sync, RobustCompany, RobustRun
    from sqlalchemy import desc
    import pytz
    from datetime import datetime, timedelta

    logger.info("[Robust] Checking automation schedules...")
    try:
        with get_db_sync() as db:
            companies = db.execute(select(RobustCompany).where(RobustCompany.enabled == True)).scalars().all()
            for company in companies:
                try:
                    tz = pytz.timezone(company.timezone or "Asia/Kolkata")
                    now_tz = datetime.now(tz)

                    # Trigger daily/weekly/monthly runs
                    sched_hour, sched_min = map(int, (company.fetch_time or "07:00").split(":"))
                    target = now_tz.replace(hour=sched_hour, minute=sched_min, second=0, microsecond=0)
                    
                    diff_min = (now_tz - target).total_seconds() / 60.0
                    if 0 <= diff_min < 10:
                        weekday = now_tz.strftime("%A")[:3].upper()
                        month_day = now_tz.day
                        days_list = [] if not company.days else json.loads(company.days)

                        should_run = False
                        if company.frequency == "Daily":
                            should_run = True
                        elif company.frequency == "Weekly" and weekday in days_list:
                            should_run = True
                        elif company.frequency == "Monthly" and str(month_day) in days_list:
                            should_run = True
                        elif company.frequency == "Custom" and weekday in days_list:
                            should_run = True

                        if should_run:
                            # Check if already ran today
                            last_run = db.execute(
                                select(RobustRun)
                                .where(RobustRun.company_id == company.id)
                                .order_by(desc(RobustRun.started_at))
                                .limit(1)
                            ).scalar_one_or_none()

                            run_today = False
                            if last_run:
                                run_start = last_run.started_at or last_run.finished_at
                                if run_start:
                                    run_start_local = run_start.replace(tzinfo=pytz.utc).astimezone(tz)
                                    if run_start_local.date() == now_tz.date():
                                        run_today = True

                            if not run_today:
                                logger.info(f"[Robust] Triggering schedule run for {company.name}")
                                celery_app.send_task("scraper.tasks.run_robust_automation_task", args=[company.id])

                    # Trigger monthly takeaways spreadsheet email
                    send_monthly = getattr(company, "send_monthly_takeaways_enabled", False)
                    if send_monthly:
                        m_day = getattr(company, "monthly_takeaways_day", 1) or 1
                        m_time = getattr(company, "monthly_takeaways_time", "09:00") or "09:00"
                        try:
                            m_hour, m_min = map(int, m_time.split(":"))
                            m_target = now_tz.replace(day=m_day, hour=m_hour, minute=m_min, second=0, microsecond=0)
                            m_diff_min = (now_tz - m_target).total_seconds() / 60.0
                            if now_tz.day == m_day and (0 <= m_diff_min < 10):
                                target_month_str = now_tz.strftime("%Y-%m")
                                last_sent = company.last_monthly_takeaways_sent_at
                                already_sent = False
                                if last_sent:
                                    last_sent_tz = last_sent.replace(tzinfo=pytz.utc).astimezone(tz)
                                    if last_sent_tz.strftime("%Y-%m") == target_month_str:
                                        already_sent = True
                                        
                                if not already_sent:
                                    logger.info(f"[Robust] Triggering monthly takeaways send for {company.name}")
                                    celery_app.send_task("scraper.tasks.send_robust_monthly_takeaways_report_task", args=[company.id])
                        except Exception as monthly_sched_err:
                            logger.error(f"[Robust] Monthly takeaways check error for {company.name}: {monthly_sched_err}")
                except Exception as e:
                    logger.error(f"[Robust] Schedule check error for {company.name}: {e}")
    except Exception as e:
        logger.error(f"[Robust] Schedule check failed: {e}", exc_info=True)


@celery_app.task(name="scraper.tasks.check_robust_scheduled_sends")
def check_robust_scheduled_sends():
    """
    Beat task: checks for pending robust runs scheduled for delayed send.
    """
    from db.database import get_db_sync, RobustRun, RobustCompany, RobustRecipient
    from utils.email import send_report_email, send_error_alert_email
    import pytz
    from datetime import datetime

    logger.info("[Robust] Checking scheduled sends...")
    try:
        with get_db_sync() as db:
            pending_runs = db.execute(
                select(RobustRun).where(RobustRun.email_status == "pending")
            ).scalars().all()

            for run in pending_runs:
                try:
                    company = db.execute(
                        select(RobustCompany).where(RobustCompany.id == run.company_id)
                    ).scalar_one_or_none()

                    if not company or company.mail_send_mode != "Scheduled":
                        continue

                    tz = pytz.timezone(company.timezone or "Asia/Kolkata")
                    now_tz = datetime.now(tz)
                    sched_hour, sched_min = map(int, (company.mail_send_time or "08:00").split(":"))
                    send_target = now_tz.replace(hour=sched_hour, minute=sched_min, second=0, microsecond=0)

                    if now_tz < send_target:
                        continue  # Not time yet

                    # Send email
                    recipients = db.execute(
                        select(RobustRecipient).where(RobustRecipient.company_id == company.id)
                    ).scalars().all()
                    brief_emails = [r.email for r in recipients if r.role == "brief"]
                    master_emails = [r.email for r in recipients if r.role == "master_doc"]

                    success_brief = True
                    success_master = True

                    # Generate inline HTML mailer if toggled
                    html_body = None
                    if company.send_html_mailer:
                        # Render html body from run object
                        html_body = render_robust_html_body(run, company.name)

                    if company.send_email and (brief_emails or master_emails):
                        if brief_emails:
                            success_brief = send_report_email(
                                recipient_emails=brief_emails,
                                client_name=company.name,
                                docx_path_filtered=run.mailer_doc_path if company.send_mailer_doc else None,
                                docx_path_master=None,
                                has_articles=run.relevant_count > 0,
                                brief_content=run.executive_summary,
                                html_body=html_body,
                            )
                        if master_emails:
                            success_master = send_report_email(
                                recipient_emails=master_emails,
                                client_name=company.name,
                                docx_path_filtered=run.mailer_doc_path if company.send_mailer_doc else None,
                                docx_path_master=run.master_doc_path if company.send_report_doc else None,
                                has_articles=run.relevant_count > 0,
                                brief_content=run.executive_summary,
                                excel_path_master=run.master_excel_path if company.send_report_excel else None,
                                html_body=html_body,
                            )

                    run.email_status = "sent" if (success_brief and success_master) else "failed"
                    db.commit()
                    logger.info(f"[Robust] Scheduled email sent for run {run.id} status: {run.email_status}")
                except Exception as e:
                    logger.error(f"[Robust] Scheduled email send failed for run {run.id}: {e}")
                    run.email_status = "failed"
                    run.error = str(e)
                    db.commit()
    except Exception as e:
        logger.error(f"[Robust] Scheduled sends check failed: {e}", exc_info=True)


@celery_app.task(name="scraper.tasks.send_robust_monthly_takeaways_report_task")
def send_robust_monthly_takeaways_report_task(company_id: int):
    """
    Emails the generated takeaways sheet of the company.
    """
    from db.database import get_db_sync, RobustCompany, RobustRecipient
    from utils.email import send_report_email
    from scraper.tasks import download_takeaways_sheet_file, send_monthly_takeaways_report_email
    
    try:
        with get_db_sync() as db:
            company = db.execute(select(RobustCompany).where(RobustCompany.id == company_id)).scalar_one_or_none()
            if not company or not company.takeaways_sheet_url:
                logger.error(f"[Robust] Company {company_id} takeaways sheet URL not found.")
                return False

            recipients = db.execute(
                select(RobustRecipient).where(RobustRecipient.company_id == company_id)
            ).scalars().all()
            recipient_emails = list(set(r.email for r in recipients))
            if not recipient_emails:
                logger.error(f"[Robust] No recipients for company {company_id}.")
                return False

            excel_path = download_takeaways_sheet_file(company.name)
            if not excel_path or not os.path.exists(excel_path):
                logger.error(f"[Robust] Takeaways spreadsheet download failed for {company.name}.")
                return False

            success = send_monthly_takeaways_report_email(
                recipient_emails=recipient_emails,
                client_name=company.name,
                excel_path=excel_path,
                google_sheet_url=company.takeaways_sheet_url
            )
            try:
                os.remove(excel_path)
            except Exception: pass

            if success:
                company.last_monthly_takeaways_sent_at = datetime.utcnow()
                db.commit()
                return True
            return False
    except Exception as e:
        logger.error(f"[Robust] Error in monthly takeaways send for company {company_id}: {e}", exc_info=True)
        return False


def _call_robust_llm_provider(messages: list, provider: str, max_tokens: int = 150, temperature: float = 0.2, system_prompt: str = None) -> Optional[str]:
    """Helper to route LLM queries to the requested provider."""
    from scraper.heavy_llm import _call_groq, _call_claude, _call_llm
    if not provider or provider.lower() == "none":
        return None
    try:
        if provider.lower() == "claude":
            return _call_claude(messages, system_prompt=system_prompt, max_tokens=max_tokens, temperature=temperature)
        elif provider.lower() == "groq":
            if system_prompt:
                messages = [{"role": "system", "content": system_prompt}] + messages
            return _call_groq(messages, max_tokens=max_tokens, temperature=temperature)
    except Exception as e:
        logger.error(f"[Robust LLM] Provider {provider} call failed: {e}")
    return None


@celery_app.task(name="scraper.tasks.run_robust_automation_task", time_limit=3600, soft_time_limit=3300)
def run_robust_automation_task(company_id: int):
    """
    Robust pipeline execution task: direct DB articles polling, in-memory openpyxl,
    conditional LLMs, customizable outputs.
    """
    from db.database import get_db_sync, RobustCompany, RobustRun, RobustRunArticle, RobustRecipient, Article as DBArticle
    from scraper.heavy_filter import exact_dedup, near_dedup
    from scraper.report_generator import generate_organized_docx_report, generate_excel_report, generate_mailer_docx_report
    from utils.email import send_report_email, send_error_alert_email
    import io
    import openpyxl
    import pytz
    from datetime import date, datetime, timedelta

    company_name = f"Company ID {company_id}"
    run_id = None

    with get_db_sync() as db:
        company = db.execute(select(RobustCompany).where(RobustCompany.id == company_id)).scalar_one_or_none()
        if not company:
            logger.error(f"[Robust] Company {company_id} not found.")
            return False

        run = RobustRun(company_id=company_id, status="running", started_at=datetime.utcnow())
        db.add(run)
        db.commit()
        db.refresh(run)

        run_id = run.id
        company_name = company.name
        sector_match = company.sector_match
        window_hours = company.window_hours

    try:
        def _update_progress(msg: str):
            try:
                with get_db_sync() as db:
                    run_rec = db.execute(select(RobustRun).where(RobustRun.id == run_id)).scalar_one_or_none()
                    if run_rec:
                        run_rec.progress_message = (run_rec.progress_message or "") + msg + "\n"
                        db.commit()
            except Exception as e:
                logger.warning(f"[Robust] Progress update failed: {e}")

        def save_intermediate_csv(filename, articles_list, extra_headers=[]):
            import csv
            reports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports")
            os.makedirs(reports_dir, exist_ok=True)
            file_path = os.path.join(reports_dir, filename)
            headers = ["Title", "URL", "Agency", "Published_At"] + extra_headers
            try:
                with open(file_path, "w", encoding="utf-8", newline="") as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    for a in articles_list:
                        row = [
                            a.get("title", ""),
                            a.get("url", ""),
                            a.get("agency", ""),
                            str(a.get("published_at", ""))
                        ]
                        for h in extra_headers:
                            if h == "Matched_Keyword":
                                kw_hits = a.get("_keyword_hits", [])
                                row.append(", ".join(kw_hits) if isinstance(kw_hits, list) else str(kw_hits))
                            elif h == "Sector":
                                row.append(a.get("_pillar", ""))
                            elif h == "Sub_Category":
                                row.append(a.get("_sub_category", ""))
                            else:
                                row.append("")
                        writer.writerow(row)
                return filename
            except Exception as ex:
                logger.error(f"[Robust] Failed to save intermediate CSV {filename}: {ex}")
                return None

        # 1. Fetch from Nexus Remote Feed
        _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Fetching articles from Nexus remote feed...")
        sectors = [s.strip().lower() for s in sector_match.split(",") if s.strip()]
        cutoff_dt = datetime.utcnow() - timedelta(hours=window_hours)
        cutoff_date = cutoff_dt.date().isoformat()

        import requests as _requests
        import time as _time

        nexus_base = os.getenv("NEXUS_FEED_URL", "http://34.142.240.96")
        nexus_key  = os.getenv("NEXUS_SERVICE_KEY", "nexus_sk_fb74eaae34cd3e53f6ac2031479337cb")
        FETCH_TIMEOUT = int(os.getenv("NEXUS_FETCH_TIMEOUT", "60"))
        FETCH_RETRIES = int(os.getenv("NEXUS_FETCH_RETRIES", "3"))
        FETCH_RETRY_DELAY = 2

        SECTOR_DB_MAPPING = {
            "travel": ["travel", "travell"],
            "tech": ["tech", "techhh"],
            "foods and drinks": ["foods and drinks", "foods"],
        }

        # Resolve variants
        query_variants = []
        for sec in sectors:
            variants = SECTOR_DB_MAPPING.get(sec, [sec])
            query_variants.extend(variants)

        # De-duplicate while preserving order
        unique_variants = []
        seen = set()
        for v in query_variants:
            if v.lower() not in seen:
                seen.add(v.lower())
                unique_variants.append(v)

        fetched = []
        seen_ids = set()

        def _fetch_with_retry(url: str, params: dict) -> dict | None:
            for attempt in range(1, FETCH_RETRIES + 1):
                try:
                    resp = _requests.get(url, params=params, timeout=FETCH_TIMEOUT)
                    resp.raise_for_status()
                    return resp.json()
                except Exception as ex:
                    if attempt < FETCH_RETRIES:
                        _time.sleep(FETCH_RETRY_DELAY)
                    else:
                        logger.error(f"[Robust] Fetch failed for {url} with params {params}: {ex}")
                        return None

        for v in unique_variants:
            _update_progress(f"  → Fetching sector: {v}")
            page = 1
            while True:
                data = _fetch_with_retry(
                    f"{nexus_base}/api/feed",
                    {
                        "api_key":   nexus_key,
                        "sector":    v,
                        "date_from": cutoff_date,
                        "page":      page,
                        "page_size": 500,
                    },
                )
                if data is None:
                    break

                for a in data.get("articles", []):
                    # We match article's published_at date/time against cutoff_dt
                    pub_at_str = a.get("published_at")
                    pub_dt = None
                    if pub_at_str:
                        try:
                            from dateutil.parser import parse as parse_date
                            pub_dt = parse_date(pub_at_str).replace(tzinfo=None)
                            if pub_dt < cutoff_dt:
                                continue
                        except Exception:
                            pass

                    uid = a.get("id") or a.get("url")
                    if uid and uid not in seen_ids:
                        seen_ids.add(uid)
                        fetched.append({
                            "id":           a.get("id"),
                            "title":        a.get("title"),
                            "url":          a.get("resolved_url") or a.get("url"),
                            "published_at": pub_dt,
                            "full_body":    a.get("full_body"),
                            "summary":      a.get("summary"),
                            "agency":       a.get("agency"),
                            "author":       a.get("author"),
                            "source_feed":  a.get("sector") or v,
                        })

                total_pages = data.get("total_pages", 1)
                if page >= total_pages:
                    break
                page += 1

        # Deduplicate fetched list by URL
        seen_urls = set()
        unique_fetched = []
        for a in fetched:
            if a["url"] not in seen_urls:
                seen_urls.add(a["url"])
                unique_fetched.append(a)
        fetched = unique_fetched

        _update_progress(f"Fetched {len(fetched)} articles matching sectors: {', '.join(sectors)}")
        fetched_fn = f"Robust_Fetched_Articles_Run_{run_id}.csv"
        save_intermediate_csv(fetched_fn, fetched)
        _update_progress(f"Step 1: Fetched {len(fetched)} articles from Nexus feed. Downloadable output: /api/robust-automation/reports/{fetched_fn}")

        if not fetched:
            with get_db_sync() as db:
                run_rec = db.execute(select(RobustRun).where(RobustRun.id == run_id)).scalar_one_or_none()
                if run_rec:
                    run_rec.status = "completed"
                    run_rec.finished_at = datetime.utcnow()
                    db.commit()
            return True

        # 2. Exact & Near Deduplication within agency
        _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Exact deduplication...")
        deduped_exact = exact_dedup(fetched)
        _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Near-duplicate TF-IDF clustering...")
        deduped = near_dedup(deduped_exact, threshold=0.80)

        deduped_fn = f"Robust_Deduplicated_Articles_Run_{run_id}.csv"
        save_intermediate_csv(deduped_fn, deduped)
        _update_progress(f"Step 2: After deduplication: {len(deduped)} articles. Downloadable output: /api/robust-automation/reports/{deduped_fn}")

        # 3. Dynamic Excel Parsing
        keyword_index = None
        priority_publications = None

        if company.keywords_file_data:
            _update_progress("Parsing uploaded keywords Excel sheet...")
            try:
                wb = openpyxl.load_workbook(io.BytesIO(company.keywords_file_data), data_only=True)
                ws = wb.active
                SUBCATEGORY_LABELS = {
                    "brand names", "india-specific", "common misspellings",
                    "leadership/personnel", "company keywords", "competition keywords",
                    "competition brand keywords", "industry keywords", "competitor keywords",
                }
                
                # Pooja's keywords sheet parser
                parsed_sectors = {}
                current_sec = None
                for row in ws.iter_rows(min_row=1, values_only=True):
                    col_a = str(row[0]).strip() if row[0] else ""
                    col_b = str(row[1]).strip() if row[1] else ""
                    if not col_a and not col_b:
                        continue
                    col_a_lower = col_a.lower().strip()
                    
                    is_sector_header = False
                    if col_a and col_a_lower not in SUBCATEGORY_LABELS:
                        if not col_b or col_b.lower() == "keywords" or col_b.lower().startswith("keywords ("):
                            is_sector_header = True
                            
                    if is_sector_header:
                        current_sec = col_a.strip()
                        if current_sec not in parsed_sectors:
                            parsed_sectors[current_sec] = {}
                        continue
                        
                    if col_a_lower in SUBCATEGORY_LABELS and col_b:
                        # Extract keywords
                        from Pooja_filtering_Logic_for_heavy_automation_final.filter_by_keywords import extract_keywords
                        kws = extract_keywords(col_b)
                        if current_sec and kws:
                            parsed_sectors[current_sec].setdefault(col_a.strip(), []).extend(kws)
                
                # Build keyword index
                keyword_index = []
                for sec_name, sub_cats in parsed_sectors.items():
                    for sub_cat, kw_list in sub_cats.items():
                        for kw in kw_list:
                            keyword_index.append((kw, sec_name, sub_cat))
                keyword_index.sort(key=lambda x: -len(x[0]))
                _update_progress(f"Loaded {len(keyword_index)} custom keyword rules.")
            except Exception as e:
                logger.error(f"[Robust] Keywords Excel parse error: {e}")
                _update_progress(f"Warning: Keywords file parsing failed: {e}")

        # Parse manual keywords list if available
        if company.manual_keywords:
            _update_progress("Parsing manually entered keywords...")
            try:
                import re as _re
                normalized_text = company.manual_keywords.replace("\n", ",")
                raw_kws = [k.strip() for k in normalized_text.split(",") if k.strip()]
                
                if keyword_index is None:
                    keyword_index = []
                
                for raw_kw in raw_kws:
                    cleaned_kw = _re.sub(r'^\d+[\s\.\)-]*', '', raw_kw).strip()
                    if cleaned_kw:
                        # Append as keyword
                        keyword_index.append((cleaned_kw, company.name, "Manual Keywords"))
                
                # Deduplicate and sort by length descending
                keyword_index = list(set(keyword_index))
                keyword_index.sort(key=lambda x: -len(x[0]))
                _update_progress(f"Loaded {len(keyword_index)} total keyword rules (including manual).")
            except Exception as e:
                logger.error(f"[Robust] Manual keywords parse error: {e}")
                _update_progress(f"Warning: Manual keywords parsing failed: {e}")

        if company.priority_media_file_data:
            _update_progress("Parsing uploaded priority publications Excel sheet...")
            try:
                wb = openpyxl.load_workbook(io.BytesIO(company.priority_media_file_data), read_only=True, data_only=True)
                ws = wb.active
                strict, lenient = [], []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    cells = list(row) + [None, None]
                    serial, pub_name = cells[0], cells[1]
                    if not pub_name:
                        continue
                    cname = str(pub_name).strip().split("\n")[0].strip()
                    if cname.lower() in ("", "name", "publication", "publications", "media") or len(cname) > 120:
                        continue
                    lenient.append(cname)
                    if isinstance(serial, (int, float)):
                        strict.append(cname)
                priority_publications = strict if strict else lenient
                # deduplicate
                seen_p = set()
                unique_p = []
                for p in priority_publications:
                    if p.lower() not in seen_p:
                        seen_p.add(p.lower())
                        unique_p.append(p)
                priority_publications = unique_p
                _update_progress(f"Loaded {len(priority_publications)} custom priority publications.")
            except Exception as e:
                logger.error(f"[Robust] Priority media Excel parse error: {e}")
                _update_progress(f"Warning: Priority media list parsing failed: {e}")

        # Helper to match publication against loaded list
        def is_agency_priority(agency: str, publications: list[str]) -> bool:
            if not agency or not publications:
                return False
            a = agency.lower().strip()
            from Pooja_filtering_Logic_for_heavy_automation_final.filter_priority_media import KNOWN_ALIASES
            seen = set()
            pairs = []
            
            def add_pat(pattern, pub):
                p = pattern.lower().strip()
                if p and p not in seen:
                    seen.add(p)
                    pairs.append((p, pub))
                    
            for pub in publications:
                add_pat(pub, pub)
                for alias in KNOWN_ALIASES.get(pub, []):
                    add_pat(alias, pub)
                    
            import re
            for p, pub in pairs:
                esc = re.escape(p)
                left = r"(?<!\w)" if p[:1].isalnum() else r""
                right = r"(?!\w)" if p[-1:].isalnum() else r""
                rx = re.compile(left + esc + right, re.IGNORECASE)
                if rx.search(a):
                    return True
            return False

        def match_title_against_index(title: str, index: list) -> tuple:
            title_lower = title.lower()
            import re
            for kw, sec, sub in index:
                kw_lower = kw.lower()
                # Support "+" logic (AND match)
                if "+" in kw:
                    parts = [p.strip().lower() for p in kw.split("+") if p.strip()]
                    if parts and all(p in title_lower for p in parts):
                        return kw, sec, sub
                    continue

                words = kw_lower.split()
                if len(words) > 1:
                    if all(w in title_lower for w in words):
                        return kw, sec, sub
                else:
                    if len(kw_lower) <= 3:
                        if re.search(r'\b' + re.escape(kw_lower) + r'\b', title_lower):
                            return kw, sec, sub
                    else:
                        if kw_lower in title_lower:
                            return kw, sec, sub
            return None, None, None

        # 4. Pooja's Filtering & Keyword Relevance Matches
        relevant_list = []
        discard_list = []

        _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Applying custom filtering rules...")
        for art in deduped:
            # If pooja_algo_enabled is False, bypass all filtering checks and keep all articles
            if not getattr(company, "pooja_algo_enabled", True):
                art["_pillar"] = "General"
                art["_sub_category"] = "General"
                if getattr(company, "group_by_source_sector", False):
                    source_sec = art.get("source_feed")
                    art["_pillar"] = source_sec.title() if source_sec else "General News"
                    art["_sub_category"] = "General"
                art["_keyword_hits"] = []
                art["_is_priority"] = True
                art["_relevance_score"] = 1.0
                art["confidence_score"] = 10
                art["_bucket"] = "clear_keep"
                relevant_list.append(art)
                continue

            # If priority publications list is uploaded, match against it
            is_pri = True
            if priority_publications is not None:
                is_pri = is_agency_priority(art.get("agency") or "", priority_publications)

            # If keywords file is uploaded, match against it
            has_kw = True
            matched_kw, pillar, sub_cat = None, None, None
            if keyword_index is not None:
                match_text = art.get("title") or ""
                if getattr(company, "search_mode", "title") == "body":
                    match_text = f"{match_text}\n{art.get('full_body') or ''}"
                matched_kw, pillar, sub_cat = match_title_against_index(match_text, keyword_index)
                has_kw = matched_kw is not None

            # Combined rules: must match both filters if both are present
            if is_pri and has_kw:
                art["_pillar"] = pillar or "General"
                art["_sub_category"] = sub_cat or "General"
                if getattr(company, "group_by_source_sector", False):
                    source_sec = art.get("source_feed")
                    art["_pillar"] = source_sec.title() if source_sec else "General News"
                    art["_sub_category"] = "General"
                art["_keyword_hits"] = [matched_kw] if matched_kw else []
                art["_is_priority"] = is_pri
                art["_relevance_score"] = 1.0
                art["confidence_score"] = 10
                art["_bucket"] = "clear_keep"
                relevant_list.append(art)
            else:
                discard_list.append(art)

        _update_progress(f"Matches before LLM validation: {len(relevant_list)} relevant, {len(discard_list)} discarded")
        pooja_filtered_fn = f"Robust_Pooja_Filtered_Articles_Run_{run_id}.csv"
        save_intermediate_csv(pooja_filtered_fn, relevant_list, ["Matched_Keyword", "Sector", "Sub_Category"])
        _update_progress(f"Step 3: Pooja filtered matches (before LLM verification): {len(relevant_list)} articles. Downloadable output: /api/robust-automation/reports/{pooja_filtered_fn}")

        # 5. Conditional LLM Verification
        llm_verify = company.llm_verification_provider
        if llm_verify and llm_verify.lower() != "none" and relevant_list:
            _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Validating article relevance using {llm_verify}...")
            verified_list = []
            for art in relevant_list:
                title = art.get("title") or ""
                kw_hits = art.get("_keyword_hits", [])
                keyword = kw_hits[0] if kw_hits else ""
                
                # Check with LLM
                system_prompt = "You are a precise news filtering assistant. Decide if the news article title is genuinely relevant to the matched keyword."
                prompt = f"""Article Title: {title}\nMatched Keyword: {keyword}\n\nDecide if this article is relevant. Respond ONLY with "yes" or "no"."""
                resp = _call_robust_llm_provider([{"role": "user", "content": prompt}], llm_verify, max_tokens=10, temperature=0.1, system_prompt=system_prompt)
                
                is_valid = True
                if resp:
                    is_valid = "yes" in resp.lower()
                
                if is_valid:
                    verified_list.append(art)
                else:
                    discard_list.append(art)
            relevant_list = verified_list
            _update_progress(f"Matches after LLM validation: {len(relevant_list)} verified")

        # 6. Conditional Summarization
        llm_summary = company.llm_summary_provider
        if llm_summary and llm_summary.lower() != "none" and relevant_list:
            _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Generating article summaries using {llm_summary}...")
            for art in relevant_list:
                title = art.get("title") or ""
                body = art.get("full_body") or art.get("summary") or ""
                prompt = f"Summarize this news article in 30-40 words.\n\nTitle: {title}\nBody:\n{body[:2000]}\n\nRespond with ONLY the summary."
                resp = _call_robust_llm_provider([{"role": "user", "content": prompt}], llm_summary, max_tokens=100, temperature=0.3)
                if resp:
                    art["_summary"] = resp.strip()
                    art["summary"] = resp.strip()
                else:
                    art["_summary"] = art.get("summary") or body[:200]
        else:
            for art in relevant_list:
                art["_summary"] = art.get("summary") or (art.get("full_body") or "")[:200]

        # Save Stage 4 verified list
        llm_verified_fn = f"Robust_LLM_Verified_Articles_Run_{run_id}.csv"
        save_intermediate_csv(llm_verified_fn, relevant_list, ["Matched_Keyword", "Sector", "Sub_Category"])
        _update_progress(f"Step 4: Verified relevant articles: {len(relevant_list)} articles. Downloadable output: /api/robust-automation/reports/{llm_verified_fn}")

        # 7. Executive Synthesis (Summary & Takeaways)
        exec_summary = None
        takeaways = None
        llm_exec = company.llm_executive_provider
        if llm_exec and llm_exec.lower() != "none" and relevant_list:
            _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Generating Executive Summary & Takeaways using {llm_exec}...")
            
            # Sort priority media first
            sorted_arts = sorted(relevant_list, key=lambda x: 1 if x.get("_is_priority") else 0, reverse=True)
            text_lines = []
            for idx, a in enumerate(sorted_arts[:15], start=1):
                text_lines.append(f"#{idx} [{a.get('agency') or 'Normal'}] Title: {a.get('title')}\nSummary: {a.get('_summary')}\n")
            context_text = "\n".join(text_lines)

            # Executive summary prompt
            system_summary = f"You are a premium news intelligence editor. Write a concise briefing summary for the daily {company.name} news."
            prompt_summary = f"""Based on these news articles, generate an executive summary as a bulleted list of 4-5 key developments.
Each bullet point MUST be strictly under 15 words. Keep it concise.
Articles:
{context_text}
Return ONLY the bulleted list (each line starting with `-`). No introduction."""
            exec_summary = _call_robust_llm_provider([{"role": "user", "content": prompt_summary}], llm_exec, max_tokens=300, temperature=0.2, system_prompt=system_summary)

            # Takeaways prompt
            system_takeaways = f"You are a premium strategic intelligence advisor. Extract takeaways for {company.name} from news coverage."
            prompt_takeaways = f"""Based on these news articles, formulate exactly six key strategic takeaways/insights for {company.name}.
Each takeaway must start with a bold key concept title, followed by a dash (—) and a 1-2 sentence analytical implication.
Articles:
{context_text}
Return ONLY the six takeaways. No introduction."""
            takeaways = _call_robust_llm_provider([{"role": "user", "content": prompt_takeaways}], llm_exec, max_tokens=1000, temperature=0.2, system_prompt=system_takeaways)
        else:
            exec_summary = "Executive Summary skipped (LLM executive synthesis is disabled)."
            takeaways = "Strategic Takeaways skipped (LLM executive synthesis is disabled)."

        # 8. Report Compilation
        _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Compiling reports...")
        
        # Group articles for reports
        def group_articles_for_report(articles_list):
            grouped = {}
            for a in articles_list:
                # Group by pillar and sub_category
                master = a.get("_pillar") or "General News"
                sub = a.get("_sub_category") or "General"
                grouped.setdefault(master, {}).setdefault(sub, [])
                if a["url"] not in [x["url"] for x in grouped[master][sub]]:
                    grouped[master][sub].append(a)
            return grouped

        all_grouped = group_articles_for_report(relevant_list)
        
        today_str = date.today().strftime("%Y-%m-%d")
        reports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports")
        os.makedirs(reports_dir, exist_ok=True)

        master_doc_filename = f"Robust_{company.name}_Master_{today_str}_{run_id}.docx"
        master_doc_path = os.path.join(reports_dir, master_doc_filename)
        generate_organized_docx_report(company.name, "Master Intelligence Report", today_str, all_grouped, master_doc_path)

        master_excel_filename = f"Robust_{company.name}_Master_{today_str}_{run_id}.xlsx"
        master_excel_path = os.path.join(reports_dir, master_excel_filename)
        generate_excel_report(company.name, "Master Excel Report", today_str, all_grouped, master_excel_path)

        mailer_doc_filename = f"Robust_{company.name}_Mailer_{today_str}_{run_id}.docx"
        mailer_doc_path = os.path.join(reports_dir, mailer_doc_filename)
        generate_mailer_docx_report(company.name, "Daily Brief Mailer", today_str, exec_summary, takeaways, all_grouped, mailer_doc_path)

        # Upload Mailer Doc to Google Docs
        google_doc_url = None
        with get_db_sync() as db:
            recips = db.execute(select(RobustRecipient).where(RobustRecipient.company_id == company.id)).scalars().all()
        all_emails = list(set(r.email for r in recips))

        if company.upload_to_google_drive and os.environ.get("GOOGLE_CREDENTIALS_JSON") and all_emails:
            _update_progress("Uploading briefing doc to Google Drive...")
            try:
                from utils.google_docs import upload_docx_to_google_doc
                google_doc_url = upload_docx_to_google_doc(
                    docx_path=mailer_path,
                    client_name=company.name,
                    date_str=today_str,
                    recipients=all_emails,
                    doc_suffix="Briefing"
                )
            except Exception as e:
                logger.error(f"[Robust] Google Drive upload failed: {e}")

        # Update Takeaways Sheet
        if company.update_takeaways_sheet and takeaways:
            _update_progress("Updating Takeaways Google Sheet...")
            try:
                from utils.google_docs import append_daily_takeaways_to_sheet
                sheet_url = append_daily_takeaways_to_sheet(company.name, date.today(), takeaways)
                if sheet_url:
                    company.takeaways_sheet_url = sheet_url
            except Exception as e:
                logger.error(f"[Robust] Takeaways sheet update failed: {e}")

        # Save binary outputs
        with open(master_doc_path, "rb") as f: master_doc_data = f.read()
        with open(master_excel_path, "rb") as f: master_excel_data = f.read()
        with open(mailer_doc_path, "rb") as f: mailer_doc_data = f.read()

        # Save intermediate CSV binary outputs
        fetched_csv_data = None
        fetched_csv_path = os.path.join(reports_dir, f"Robust_Fetched_Articles_Run_{run_id}.csv")
        if os.path.exists(fetched_csv_path):
            with open(fetched_csv_path, "rb") as f: fetched_csv_data = f.read()

        deduped_csv_data = None
        deduped_csv_path = os.path.join(reports_dir, f"Robust_Deduplicated_Articles_Run_{run_id}.csv")
        if os.path.exists(deduped_csv_path):
            with open(deduped_csv_path, "rb") as f: deduped_csv_data = f.read()

        pooja_csv_data = None
        pooja_csv_path = os.path.join(reports_dir, f"Robust_Pooja_Filtered_Articles_Run_{run_id}.csv")
        if os.path.exists(pooja_csv_path):
            with open(pooja_csv_path, "rb") as f: pooja_csv_data = f.read()

        verified_csv_data = None
        verified_csv_path = os.path.join(reports_dir, f"Robust_LLM_Verified_Articles_Run_{run_id}.csv")
        if os.path.exists(verified_csv_path):
            with open(verified_csv_path, "rb") as f: verified_csv_data = f.read()

        with get_db_sync() as db:
            run_rec = db.execute(select(RobustRun).where(RobustRun.id == run_id)).scalar_one_or_none()
            if run_rec:
                run_rec.master_doc_path = master_doc_path
                run_rec.master_excel_path = master_excel_path
                run_rec.mailer_doc_path = mailer_doc_path
                run_rec.google_doc_url = google_doc_url
                run_rec.master_doc_data = master_doc_data
                run_rec.master_excel_data = master_excel_data
                run_rec.mailer_doc_data = mailer_doc_data
                run_rec.fetched_csv_data = fetched_csv_data
                run_rec.deduped_csv_data = deduped_csv_data
                run_rec.pooja_csv_data = pooja_csv_data
                run_rec.verified_csv_data = verified_csv_data
                run_rec.executive_summary = exec_summary
                run_rec.takeaways = takeaways
                db.commit()

        # Audit Trail
        _update_progress("Saving article audit trail...")
        with get_db_sync() as db:
            for art in relevant_list:
                audit = RobustRunArticle(
                    run_id=run_id,
                    source_article_id=art.get("id"),
                    title=art.get("title"),
                    url=art.get("url"),
                    published_at=art.get("published_at"),
                    relevance_score=art.get("_relevance_score"),
                    included_in_brief=True,
                    pillar=art.get("_pillar"),
                    sub_category=art.get("_sub_category"),
                    matched_keywords=json.dumps(art.get("_keyword_hits", [])),
                    llm_summary=art.get("_summary"),
                    bucket=art.get("_bucket")
                )
                db.add(audit)
            db.commit()

        # 9. Email Dispatch
        email_status = "skipped"
        if company.send_email and all_emails:
            if company.mail_send_mode == "Immediate":
                _update_progress("Sending daily news email briefing...")
                
                # Render HTML mailer if enabled
                html_body = None
                if company.send_html_mailer:
                    html_body = render_robust_html_body(run_rec, company.name)

                brief_emails = [r.email for r in recips if r.role == "brief"]
                master_emails = [r.email for r in recips if r.role == "master_doc"]

                success_brief = True
                success_master = True

                if brief_emails:
                    success_brief = send_report_email(
                        recipient_emails=brief_emails,
                        client_name=company.name,
                        docx_path_filtered=mailer_doc_path if company.send_mailer_doc else None,
                        docx_path_master=None,
                        has_articles=len(relevant_list) > 0,
                        brief_content=exec_summary,
                        html_body=html_body,
                    )

                if master_emails:
                    success_master = send_report_email(
                        recipient_emails=master_emails,
                        client_name=company.name,
                        docx_path_filtered=mailer_doc_path if company.send_mailer_doc else None,
                        docx_path_master=master_doc_path if company.send_report_doc else None,
                        has_articles=len(relevant_list) > 0,
                        brief_content=exec_summary,
                        excel_path_master=master_excel_path if company.send_report_excel else None,
                        html_body=html_body,
                    )

                email_status = "sent" if (success_brief and success_master) else "failed"
            else:
                _update_progress(f"Email scheduled for delayed dispatch at {company.mail_send_time}")
                email_status = "pending"

        with get_db_sync() as db:
            run_rec = db.execute(select(RobustRun).where(RobustRun.id == run_id)).scalar_one_or_none()
            if run_rec:
                run_rec.status = "completed"
                run_rec.fetched_count = len(fetched)
                run_rec.deduped_count = len(deduped)
                run_rec.relevant_count = len(relevant_list)
                run_rec.email_status = email_status
                run_rec.finished_at = datetime.utcnow()
                db.commit()

        # Update last run timestamp in company record
        with get_db_sync() as db:
            comp_rec = db.execute(select(RobustCompany).where(RobustCompany.id == company.id)).scalar_one_or_none()
            if comp_rec:
                comp_rec.last_run_at = datetime.utcnow()
                db.commit()

        _update_progress(f"[{datetime.now().strftime('%H:%M:%S')}] Run complete!")
        return True

    except Exception as e:
        logger.error(f"[Robust] Execution failed: {e}", exc_info=True)
        try:
            with get_db_sync() as db:
                run_rec = db.execute(select(RobustRun).where(RobustRun.id == run_id)).scalar_one_or_none()
                if run_rec:
                    run_rec.status = "failed"
                    run_rec.error = str(e)
                    run_rec.finished_at = datetime.utcnow()
                    db.commit()
            send_error_alert_email(company_name, str(e))
        except Exception as alert_err:
            logger.error(f"[Robust] Failed to log failure details: {alert_err}")
        return False


def render_robust_html_body(run, company_name: str) -> str:
    """Helper to generate HTML mailer content."""
    try:
        from utils.mailer import render_brief_html
        from datetime import date
        
        # Pull audit articles to display
        from db.database import get_db_sync, RobustRunArticle
        with get_db_sync() as db:
            articles = db.execute(
                select(RobustRunArticle).where(RobustRunArticle.run_id == run.id)
            ).scalars().all()
            
        exec_bullets = []
        if run.executive_summary:
            for line in run.executive_summary.split("\n"):
                line = line.strip()
                if line.startswith("-") or line.startswith("*") or line.startswith("•"):
                    clean = line.strip("-*• ").strip()
                    # Limit to 15 words strictly
                    words = clean.split()
                    if len(words) > 15:
                        clean = " ".join(words[:15]) + "..."
                    exec_bullets.append(clean)
                    
        sections = []
        # Group by pillar and select top 30 to prevent Gmail clipping
        by_pillar = {}
        for a in articles[:30]:
            by_pillar.setdefault(a.pillar or "General", []).append({
                "title": a.title,
                "url": a.url,
                "agency": a.pillar or "News", # fallback
                "summary": a.llm_summary or ""
            })
            
        colors = ["#4285F4", "#EA4335", "#FBBC04", "#34A853"]
        for idx, (pillar, arts) in enumerate(by_pillar.items()):
            sections.append({
                "name": pillar.upper(),
                "accent": colors[idx % len(colors)],
                "articles": arts
            })
            
        brief_data = {
            "brand": company_name,
            "subtitle": "DAILY BRIEF BRIEFING",
            "date_str": date.today().strftime("%d %B %Y").upper(),
            "top_tags": list(by_pillar.keys())[:5],
            "exec_intro": "Key headline developments compiled today:",
            "exec_bullets": exec_bullets,
            "sections": sections,
            "signoff_name": "THE MAVERICKS Intelligence Desk",
            "signoff_sub": f"Daily Brief Coverage — {date.today().strftime('%d %B %Y')}",
            "sections_covered": " | ".join(by_pillar.keys()),
            "disclaimer": "This briefing is compiled from public media sources. All trademarks remain property of their respective owners.",
            "topic_tags": [f"#{p.replace(' ', '')}" for p in by_pillar.keys()][:5]
        }
        return render_brief_html(brief_data)
    except Exception as e:
        logger.error(f"[Robust] Failed to render HTML brief: {e}", exc_info=True)
        return f"<p>Daily briefing compilation complete. Executive Summary:<br>{run.executive_summary}</p>"

